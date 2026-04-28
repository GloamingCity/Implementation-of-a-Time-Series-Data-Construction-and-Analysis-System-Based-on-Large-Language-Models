import os
import json
import re
import time
import base64
import random
import threading
import argparse
from contextlib import contextmanager
from concurrent.futures import ThreadPoolExecutor, as_completed
from collections import deque
from pathlib import Path
from tqdm import tqdm
import numpy as np
from scipy import stats
import requests

try:
    import openpyxl
except ImportError:
    openpyxl = None

# ==============================================================================
#                                 配置区域
# ==============================================================================

# 1. API配置（智谱 + 硅基流动 + 豆包多路）
ZHIPU_API_KEY = "6c9f1aeae1ca442089e7b3ef3244eb7a.BTfbpqsH9lklwbKQ"
ZHIPU_BASE_URL = "https://open.bigmodel.cn/api/paas/v4/chat/completions"
SILICONFLOW_BASE_URL = os.getenv("SILICONFLOW_BASE_URL", "https://api.siliconflow.cn/v1/chat/completions")
SILICONFLOW_API_KEY = os.getenv(
    "SILICONFLOW_API_KEY",
    "sk-mpinrhqllllhjrkyjaqkosjupzedfnjputxscoqlveieofba",
)
DOUBAO_BASE_URL = os.getenv("DOUBAO_BASE_URL", "https://ark.cn-beijing.volces.com/api/v3/chat/completions")
DOUBAO_API_KEY_LITE = os.getenv("DOUBAO_API_KEY_LITE", "74b3b583-9d7f-4d81-a2f3-499883260104")
DOUBAO_API_KEY_PRO = os.getenv("DOUBAO_API_KEY_PRO", "e4b26c47-0dd1-4ef1-a386-5d467e53a0a3").strip()
DOUBAO_API_KEY_MINI = os.getenv("DOUBAO_API_KEY_MINI", "20970c68-36b3-4ed0-b499-6849d27cd75f").strip()
DOUBAO_API_KEY_SEED_18_251228 = os.getenv("DOUBAO_API_KEY_SEED_18_251228", "29e541ef-ed91-4dde-8d74-34241da97c14").strip()
DOUBAO_API_KEY_SEED_16_250615 = os.getenv("DOUBAO_API_KEY_SEED_16_250615", "0c790fde-df4d-4ff7-9e04-333f6fe58f0a").strip()
DOUBAO_API_KEY_SEED_16_FLASH_250615 = os.getenv("DOUBAO_API_KEY_SEED_16_FLASH_250615", "85427503-f454-49f1-8c75-a50478f02a31").strip()
DOUBAO_API_KEY_SEED_16_VISION_250815 = os.getenv("DOUBAO_API_KEY_SEED_16_VISION_250815", "515ebb48-f47c-452b-9da6-d32f78e864e1").strip()

# 豆包模型名固定使用公开模型名；key 使用对应接口凭据。
DOUBAO_MODEL_LITE = os.getenv("DOUBAO_MODEL_LITE", "doubao-seed-2-0-lite-260215")
DOUBAO_MODEL_PRO = os.getenv("DOUBAO_MODEL_PRO", "doubao-seed-2-0-pro-260215")
DOUBAO_MODEL_MINI = os.getenv("DOUBAO_MODEL_MINI", "doubao-seed-2-0-mini-260215")
DOUBAO_MODEL_SEED_18_251228 = os.getenv("DOUBAO_MODEL_SEED_18_251228", "doubao-seed-1-8-251228")
DOUBAO_MODEL_SEED_16_250615 = os.getenv("DOUBAO_MODEL_SEED_16_250615", "doubao-seed-1.6-250615")
DOUBAO_MODEL_SEED_16_FLASH_250615 = os.getenv("DOUBAO_MODEL_SEED_16_FLASH_250615", "doubao-seed-1-6-flash-250615")
DOUBAO_MODEL_SEED_16_VISION_250815 = os.getenv("DOUBAO_MODEL_SEED_16_VISION_250815", "doubao-seed-1-6-vision-250815")

DOUBAO_QUOTA_FALLBACK_TO_ZHIPU = True
DOUBAO_QUOTA_HTTP_STATUS = {402, 403, 429}
DOUBAO_QUOTA_ERROR_KEYWORDS = [
    "insufficient balance",
    "insufficient quota",
    "quota exhausted",
    "exceeded current quota",
    "credit is not enough",
    "account balance is not enough",
    "accountoverdueerror",
    "overdue",
    "余额不足",
    "额度不足",
    "额度已用尽",
    "配额不足",
    "欠费",
    "逾期",
    "token quota",
    "remaining quota",
    "setlimitexceeded",
    "safe experience mode",
    "model service has been paused",
    "reached the set inference limit",
    "model activation page",
]
DOUBAO_QUOTA_RECHECK_INTERVAL_SEC = 3600
DOUBAO_QUOTA_RECHECK_TIMEOUT_SEC = 30

# 模型级熔断（适用于所有 provider）：
# 命中确定性失败（如 404 模型不存在/无权限、鉴权失败、额度/配额不足）后，
# 将该模型临时下线一段时间，避免每个批次反复路由到同一失败模型拖慢整体速度。
MODEL_CIRCUIT_BREAKER_ENABLED = True
MODEL_CIRCUIT_NOT_FOUND_RECHECK_INTERVAL_SEC = int(os.getenv("MODEL_CIRCUIT_NOT_FOUND_RECHECK_INTERVAL_SEC", "3600"))
MODEL_CIRCUIT_AUTH_RECHECK_INTERVAL_SEC = int(os.getenv("MODEL_CIRCUIT_AUTH_RECHECK_INTERVAL_SEC", "3600"))
MODEL_CIRCUIT_QUOTA_RECHECK_INTERVAL_SEC = int(os.getenv("MODEL_CIRCUIT_QUOTA_RECHECK_INTERVAL_SEC", "3600"))
MODEL_CIRCUIT_RATE_LIMIT_RECHECK_INTERVAL_SEC = int(os.getenv("MODEL_CIRCUIT_RATE_LIMIT_RECHECK_INTERVAL_SEC", "90"))
MODEL_CIRCUIT_SKIP_LOG_INTERVAL_SEC = int(os.getenv("MODEL_CIRCUIT_SKIP_LOG_INTERVAL_SEC", "120"))
MODEL_CIRCUIT_MANAGED_PROVIDERS = {
    s.strip().lower()
    for s in os.getenv("MODEL_CIRCUIT_MANAGED_PROVIDERS", "doubao").split(",")
    if s.strip()
}
ACTIVE_MODEL_FORCE_FALLBACK_LOG_INTERVAL_SEC = int(os.getenv("ACTIVE_MODEL_FORCE_FALLBACK_LOG_INTERVAL_SEC", "60"))
MODEL_CIRCUIT_NOT_FOUND_KEYWORDS = [
    "invalidendpointormodel.notfound",
    "model or endpoint",
    "does not exist or you do not have access",
    "model not found",
    "endpoint not found",
    "no such model",
    "unknown model",
]
MODEL_CIRCUIT_AUTH_KEYWORDS = [
    "unauthorized",
    "forbidden",
    "permission denied",
    "authentication",
    "invalid api key",
    "api key",
    "access denied",
    "无权限",
    "鉴权",
]
MODEL_CIRCUIT_QUOTA_KEYWORDS = [
    "quota",
    "balance",
    "credit",
    "insufficient",
    "setlimitexceeded",
    "safe experience",
    "service has been paused",
    "inference limit",
    "activation page",
    "额度",
    "余额",
    "配额",
]

# 硅基流动动态降并发：当单轮重试次数过高时，当前任务临时降到1并发。
SILICONFLOW_DYNAMIC_DEGRADE_ENABLED = True
SILICONFLOW_RETRY_DEGRADE_THRESHOLD = 10

# 主模型调度链（优先填满智谱+硅基，剩余任务溢出到豆包）
MODEL_FALLBACKS = [
    {
        "id": "zhipu_primary_flash",
        "provider": "zhipu",
        "name": "glm-4.1v-thinking-flash",
        "base_url": ZHIPU_BASE_URL,
        "api_key": ZHIPU_API_KEY,
        "rpm": 1000,
        "tpm": 50000,
        "max_concurrency": 4,
        "timeout_sec_single": 120,
        "timeout_sec_batch": 180,
    },
    {
        "id": "siliconflow_glm_41v_9b_thinking",
        "provider": "siliconflow",
        "name": "THUDM/GLM-4.1V-9B-Thinking",
        "base_url": SILICONFLOW_BASE_URL,
        "api_key": SILICONFLOW_API_KEY,
        "rpm": 1000,
        "tpm": 50000,
        "max_concurrency": 2,
        "timeout_sec_single": 150,
        "timeout_sec_batch": 210,
        "disable_failover_on_error": True,
    },
    {
        "id": "doubao_seed_20_lite_260215",
        "provider": "doubao",
        "name": "doubao-seed-2-0-lite-260215",
        "model": DOUBAO_MODEL_LITE,
        "base_url": DOUBAO_BASE_URL,
        "api_key": DOUBAO_API_KEY_LITE,
        "rpm": 30000,
        "tpm": 5000000,
        "max_concurrency": 8,
        "timeout_sec_single": 90,
        "timeout_sec_batch": 150,
    },
    {
        "id": "doubao_seed_20_pro_260215",
        "provider": "doubao",
        "name": "doubao-seed-2-0-pro-260215",
        "model": DOUBAO_MODEL_PRO,
        "base_url": DOUBAO_BASE_URL,
        "api_key": DOUBAO_API_KEY_PRO,
        "rpm": 30000,
        "tpm": 5000000,
        "max_concurrency": 8,
        "timeout_sec_single": 90,
        "timeout_sec_batch": 150,
    },
    {
        "id": "doubao_seed_20_mini_260215",
        "provider": "doubao",
        "name": "doubao-seed-2-0-mini-260215",
        "model": DOUBAO_MODEL_MINI,
        "base_url": DOUBAO_BASE_URL,
        "api_key": DOUBAO_API_KEY_MINI,
        "rpm": 30000,
        "tpm": 5000000,
        "max_concurrency": 8,
        "timeout_sec_single": 90,
        "timeout_sec_batch": 150,
    },
    {
        "id": "doubao_seed_18_251228",
        "provider": "doubao",
        "name": "doubao-seed-1-8-251228",
        "model": DOUBAO_MODEL_SEED_18_251228,
        "base_url": DOUBAO_BASE_URL,
        "api_key": DOUBAO_API_KEY_SEED_18_251228,
        "rpm": 30000,
        "tpm": 5000000,
        "max_concurrency": 8,
        "timeout_sec_single": 90,
        "timeout_sec_batch": 150,
    },
    {
        "id": "doubao_seed_16_250615",
        "provider": "doubao",
        "name": "doubao-seed-1.6-250615",
        "model": DOUBAO_MODEL_SEED_16_250615,
        "base_url": DOUBAO_BASE_URL,
        "api_key": DOUBAO_API_KEY_SEED_16_250615,
        "rpm": 30000,
        "tpm": 5000000,
        "max_concurrency": 8,
        "timeout_sec_single": 90,
        "timeout_sec_batch": 150,
    },
    {
        "id": "doubao_seed_16_flash_250615",
        "provider": "doubao",
        "name": "doubao-seed-1-6-flash-250615",
        "model": DOUBAO_MODEL_SEED_16_FLASH_250615,
        "base_url": DOUBAO_BASE_URL,
        "api_key": DOUBAO_API_KEY_SEED_16_FLASH_250615,
        "rpm": 30000,
        "tpm": 5000000,
        "max_concurrency": 8,
        "timeout_sec_single": 90,
        "timeout_sec_batch": 150,
    },
    {
        "id": "doubao_seed_16_vision_250815",
        "provider": "doubao",
        "name": "doubao-seed-1-6-vision-250815",
        "model": DOUBAO_MODEL_SEED_16_VISION_250815,
        "base_url": DOUBAO_BASE_URL,
        "api_key": DOUBAO_API_KEY_SEED_16_VISION_250815,
        "rpm": 30000,
        "tpm": 5000000,
        "max_concurrency": 8,
        "timeout_sec_single": 90,
        "timeout_sec_batch": 150,
    },
]

# 硅基流动与智谱官方共用提示样本池（阈值在本脚本中本身即全局共享）。
REFERENCE_POOL_SHARED_MODEL_GROUPS = {
    "zhipu_primary_flash": "zhipu_with_siliconflow",
    "siliconflow_glm_41v_9b_thinking": "zhipu_with_siliconflow",
    # 各路豆包分别维护独立提示样本池，避免跨模型污染。
    "doubao_seed_20_lite_260215": "doubao_lite_only",
    "doubao_seed_20_pro_260215": "doubao_pro_only",
    "doubao_seed_20_mini_260215": "doubao_mini_only",
    "doubao_seed_18_251228": "doubao_seed_18_251228_only",
    "doubao_seed_16_250615": "doubao_seed_16_250615_only",
    "doubao_seed_16_flash_250615": "doubao_seed_16_flash_250615_only",
    "doubao_seed_16_vision_250815": "doubao_seed_16_vision_250815_only",
}

# 2. 生成配置（OpenAI兼容请求体）
generation_config = {
    "temperature": 0.1,
    "top_p": 1.0,
    "max_tokens": 2048,
}

# 3. 默认速率限制（仅在未命中模型专属配置时使用）
RPM_LIMIT = 1000
TPM_LIMIT = 50000

# 3. 构建发送给模型的指令 (Prompt)
#    这是最关键的部分，它指导模型如何对图片进行打分。
#    指令清晰地描述了评分标准，并要求模型以JSON格式返回结果，便于程序解析。
PROMPT = """
你是一个严格、精确、专业的“时间序列图文一致性”评审员。
你的任务是根据下方的评分细则，对提供给你的图片（左侧为时间序列曲线，右侧为中文描述）进行打分。

====================
评分规则
====================

1.  **S1 趋势与形态匹配（0-40分）**：评估描述中的“上升/下降/先升后降/先降后升/震荡/平台”等核心趋势是否与时间序列曲线的实际形态严格一致。
2.  **S2 关键变化定位（0-30分）**：评估描述中提到的关键变化点（如阶段切换点、峰值、谷值、突变或异常点）在时间轴上的位置是否与曲线上的实际位置大致相符。
3.  **S3 幅度与波动强弱（0-20分）**：评估描述中使用的形容词（如“波动剧烈”、“小幅上涨”、“窄幅震荡”、“高位整理”）是否准确地反映了曲线的实际波动幅度和所处的价格/数值区间。
4.  **S4 局部与整体一致性（0-10分）**：评估分段描述之间是否存在矛盾，或者局部描述是否与整体总结相悖。
5.  **P 幻觉惩罚（0-10分）**：如果描述中出现了曲线完全不支持的、凭空捏造的趋势、事件或数据（例如，曲线明明在下降，描述却说在上升），则进行扣分。

最终分数 = S1 + S2 + S3 + S4 - P，分数范围为0到100。

====================
硬约束（必须遵守）
====================
- 如果出现明显方向性错误（上升/下降说反）、明显阶段错配、把微小噪声当成"高频震荡/周期性"，必须显著扣分（通常不应高于60）。
- 若局部描述与整体总结自相矛盾，S4不得给高分。
- 只有在"几乎无事实性错误、关键拐点位置也基本正确"时，才允许给85分以上。
- 若提供了参考样例，可能同时包含高分正例与低分反例：
    与低分反例高度相似的待评图不得给高分，与高分正例相似才可给高分。
- 宁可保守，不可宽松高估。

====================
输出要求
====================
-   你可以完整输出你的思考与分析过程。
-   建议在开头先给最终分数标记，再写其它内容。可使用以下任一格式：
    1) {"score": 85}
    2) SCORE: 85
    3) 最终分数: 85
-   若输出了思考过程，也必须保证第一行先给分数。
-   最终分数必须是0到100的整数。
-   如果你无法判断，也必须给出一个保守分数，并按上述格式之一输出。

示例输出:
{
  "score": 85
}
"""

BATCH_PROMPT = """
你是严格的时间序列图文一致性评审员。
评分规则: S1(0-40)+S2(0-30)+S3(0-20)+S4(0-10)-P(0-10), 最终分数0-100整数。
硬约束:
- 若存在明显方向性错误/阶段错配/把微小噪声当强周期，分数通常不应高于60。
- 只有几乎无事实错误且关键拐点基本正确时，才允许85分以上。
- 若给了参考样例，样例中可能同时有高分正例和低分反例；与低分反例相似的待评图不得给高分。
输出要求:
- 第一行必须直接输出JSON，不要前言，不要<think>标签，不要解释、思考、代码块。
- 格式必须是: {"scores":[{"id":"img1","score":85},{"id":"img2","score":73}]}
- 每个id必须出现一次，score必须是0-100整数。
"""

SINGLE_COMPACT_PROMPT = """
你是时间序列图文一致性评审员。
请只输出一个0-100整数分数，严格按以下任一格式且必须在第一行给出：
1) {"score": 75}
2) SCORE: 75
禁止输出解释、思考、分析过程、代码块。
"""

BATCH_MAX_TOKENS = 1536
SINGLE_COMPACT_MAX_TOKENS = 512
BATCH_TRUNCATION_FAIL_LIMIT = 2
SINGLE_TRUNCATION_FAIL_LIMIT = 3
SILICONFLOW_BATCH_MAX_REFERENCE_IMAGES = 4

# 4. 断点续跑配置
RESUME_ENABLED = True
FORCE_RERUN = False
CACHE_FILE_NAME = "_image_scores_cache.json"

# 5. 失败重试配置（避免图片因瞬时错误被跳过）
RETRY_FOREVER = True
MAX_RETRIES = None  # 仅在 RETRY_FOREVER=False 时生效，None 表示不限制
RETRY_SHORT_DELAY = 2.0
RETRY_SHORT_ATTEMPTS = 10  # 前10次失败每次等待2秒
RETRY_LONG_DELAY = 5.0

# 5.1 并发配置（按“每个API key最多5并发请求”估算）
PER_KEY_MAX_CONCURRENCY = 5
MAX_CONCURRENCY = max(
    1,
    sum(max(1, int(cfg.get("max_concurrency", PER_KEY_MAX_CONCURRENCY))) for cfg in MODEL_FALLBACKS),
)

# 6. 运行前模型可用性检查
PRECHECK_MODELS_BEFORE_RUN = True
PRECHECK_TIMEOUT = 30
PRECHECK_IMAGE_URLS = [
    "https://img.iplaysoft.com/wp-content/uploads/2019/free-images/free_stock_photo.jpg",
    "https://images.unsplash.com/photo-1472214103451-9374bd1c798e?auto=format&fit=crop&w=640&q=60",
]
PRECHECK_IMAGE_DATA_URL = "data:image/png;base64,iVBORw0KGgoAAAANSUhEUgAAAAEAAAABCAQAAAC1HAwCAAAAC0lEQVR42mP8/x8AAwMCAO7+X9kAAAAASUVORK5CYII="

# 7. 人工评审校准评测配置
HUMAN_EVAL_ROOT_DEFAULT = Path("analysis_summary") / "人工评审"
HUMAN_EVAL_TRAIN_DIR_NAME = "Train"
HUMAN_EVAL_TEST_DIR_NAME = "Test"
HUMAN_EVAL_OUTPUT_NAME = "human_eval_comparison.json"
HUMAN_EVAL_CACHE_NAME = "_human_eval_test_scores_cache.json"
HUMAN_EVAL_USE_CACHE = True
HUMAN_EVAL_STABILITY_RUNS_DEFAULT = 3
HUMAN_EVAL_TEST_RUNS_PER_IMAGE_DEFAULT = 3
HUMAN_EVAL_BATCH_QUERY_SIZE_DEFAULT = 5
HUMAN_EVAL_MAX_IMAGES_PER_REQUEST = 5
HUMAN_EVAL_CALIBRATION_MODE_DEFAULT = "hybrid"
HUMAN_EVAL_CALIBRATION_METHOD_DEFAULT = "auto"
HUMAN_EVAL_CLIP_NEGATIVE_HUMAN_TO_ZERO_DEFAULT = True
HUMAN_EVAL_BATCH_MAX_RETRIES = None  # None表示无限重试
HUMAN_EVAL_SINGLE_MAX_RETRIES = None  # None表示无限重试
HUMAN_EVAL_STABILITY_STD_THRESHOLD_DEFAULT = 2.0
HUMAN_EVAL_STABILITY_RANGE_THRESHOLD_DEFAULT = 5
HUMAN_EVAL_STABILITY_MAX_EXTRA_RUNS_DEFAULT = 3
HUMAN_EVAL_SCORE_AGGREGATION_DEFAULT = "mean"
HUMAN_EVAL_CLEAN_KEEP_THRESHOLD_DEFAULT = 75
HUMAN_EVAL_CLEAN_DROP_THRESHOLD_DEFAULT = 60
HUMAN_EVAL_CLEAN_UNCERTAIN_ACTION_DEFAULT = "drop"
HUMAN_EVAL_CLEAN_USE_AUTO_THRESHOLD_DEFAULT = True
HUMAN_EVAL_CLEAN_TARGET_KEEP_HUMAN_THRESHOLD_DEFAULT = 75
HUMAN_EVAL_CLEAN_TARGET_DROP_HUMAN_THRESHOLD_DEFAULT = 60
HUMAN_EVAL_CLEAN_TARGET_PRECISION_DEFAULT = 0.9
HUMAN_EVAL_CLEAN_MIN_SUPPORT_DEFAULT = 2
HUMAN_EVAL_MIN_PEARSON_THRESHOLD_DEFAULT = 0.8
HUMAN_EVAL_STRICT_PEARSON_CHECK_DEFAULT = False

# 8. 数据筛选模式下的类型化提示样本配置（迭代优化）
PROJECT_ROOT = Path(__file__).resolve().parent.parent
TYPE_REFERENCE_ITERATION_DIR = PROJECT_ROOT / "Sample" / "iteration_1"
TYPE_REFERENCE_TARGET_COUNT = 6
TYPE_REFERENCE_TARGET_COUNT_SILICONFLOW = 4
TYPE_REFERENCE_ITERATION_DIRS = {
    "iteration_1": PROJECT_ROOT / "Sample" / "iteration_1",
    "iteration_2": PROJECT_ROOT / "Sample" / "iteration_2",
    "iteration_3": PROJECT_ROOT / "Sample" / "iteration_3",
}
TYPE_REFERENCE_POOLS = {
    # anomaly/classification当前没有新增人工好样本，保留现有池；每次从池内随机采样至固定6张。
    "anomaly_detection": [
        {"iteration": "iteration_2", "indexes": [1, 2, 4, 5, 6, 7, 8, 9]},
    ],
    "classification": [
        {"iteration": "iteration_1", "indexes": [3]},
    ],
    # prediction池已合并第三/四轮人工好样本，与既有池混合后随机替换，保持总数固定。
    "prediction": [
        {"iteration": "iteration_2", "indexes": [2, 3, 5, 6, 10]},
        {"iteration": "iteration_3", "indexes": [2, 8, 9]},
        {"iteration": "iteration_4", "indexes": [4, 6, 7, 10]},
    ],
}
TYPE_REFERENCE_LABELS = {
    "anomaly_detection": "异常检测",
    "classification": "分类",
    "prediction": "预测",
}
TYPE_REFERENCE_LIMITS = {
    "anomaly_detection": TYPE_REFERENCE_TARGET_COUNT,
    "classification": TYPE_REFERENCE_TARGET_COUNT,
    "prediction": TYPE_REFERENCE_TARGET_COUNT,
}
TYPE_ALIASES = {
    "anomaly": "anomaly_detection",
    "anomaly_detection": "anomaly_detection",
    "classification": "classification",
    "classify": "classification",
    "prediction": "prediction",
    "forecast": "prediction",
}
# 可选保留的既有人工提示目录（存在则混合，不存在则跳过）
LEGACY_REFERENCE_DIR_BY_TYPE = {
    "classification": [
        PROJECT_ROOT
        / "Dataset"
        / "UEA&UCR_Multivariate_Time_Series_Classification_Archive"
        / "analysis_summary"
        / "Manual_Review"
        / "Train"
    ],
}
LEGACY_REFERENCE_MAX_PER_DIR = 6

# 人工高/低分提示样本（来自 Manual_Review）
MANUAL_REVIEW_ROOT = (
    PROJECT_ROOT
    / "Dataset"
    / "UEA&UCR_Multivariate_Time_Series_Classification_Archive"
    / "analysis_summary"
    / "Manual_Review"
)
MANUAL_REVIEW_POSITIVE_DIR = MANUAL_REVIEW_ROOT / "Positive"
MANUAL_REVIEW_NEGATIVE_DIR = MANUAL_REVIEW_ROOT / "Negative"
MANUAL_REVIEW_TRAIN_SHEET_PATH = MANUAL_REVIEW_ROOT / "Train" / "Manual_Review_Scoring_Sheet_Train.xlsx"

# 提示样本阈值与配比：默认 3高+3低；硅基流动 2高+2低。
REFERENCE_SAMPLE_ACCEPT_THRESHOLD = 75
REFERENCE_SAMPLE_HIGH_SCORE_THRESHOLD = 88
REFERENCE_SAMPLE_LOW_SCORE_THRESHOLD = 60
REFERENCE_SAMPLE_DEFAULT_HIGH_COUNT = 3
REFERENCE_SAMPLE_DEFAULT_LOW_COUNT = 3
REFERENCE_SAMPLE_SILICONFLOW_HIGH_COUNT = 2
REFERENCE_SAMPLE_SILICONFLOW_LOW_COUNT = 2

REFERENCE_UPDATE_MIN_SCORE = REFERENCE_SAMPLE_HIGH_SCORE_THRESHOLD
DOUBAO_REFERENCE_UPDATE_MIN_SCORE = REFERENCE_SAMPLE_HIGH_SCORE_THRESHOLD
DOUBAO_PRO_REFERENCE_UPDATE_MIN_SCORE = REFERENCE_SAMPLE_HIGH_SCORE_THRESHOLD
DOUBAO_MINI_REFERENCE_UPDATE_MIN_SCORE = REFERENCE_SAMPLE_HIGH_SCORE_THRESHOLD
REFERENCE_POOL_MAX_CANDIDATES = 24
REFERENCE_DIVERSITY_MAX_PER_DATASET = 2
_TYPE_REFERENCE_CACHE: dict[str, list[dict]] = {}
_TYPE_REFERENCE_LOCK = threading.Lock()
_MANUAL_REVIEW_REFERENCE_CACHE: list[dict] | None = None
_MANUAL_REVIEW_SCORE_CACHE: dict[str, int] | None = None

# 模型评分阈值建议（keep按包含边界执行，基于最新 human_eval 实测结果）。
DOUBAO_RECOMMENDED_SCORE_THRESHOLD = 75
DOUBAO_PRO_RECOMMENDED_SCORE_THRESHOLD = 75
DOUBAO_MINI_RECOMMENDED_SCORE_THRESHOLD = 75
DOUBAO_RECOMMENDED_DROP_THRESHOLD = 62
DOUBAO_PRO_RECOMMENDED_DROP_THRESHOLD = 62
DOUBAO_MINI_RECOMMENDED_DROP_THRESHOLD = 56
MODEL_SCORE_THRESHOLD_OVERRIDES = {
    "doubao_seed_20_lite_260215": DOUBAO_RECOMMENDED_SCORE_THRESHOLD,
    "doubao::doubao-seed-2-0-lite-260215::doubao_seed_20_lite_260215": DOUBAO_RECOMMENDED_SCORE_THRESHOLD,
    "doubao_seed_20_pro_260215": DOUBAO_PRO_RECOMMENDED_SCORE_THRESHOLD,
    "doubao::doubao-seed-2-0-pro-260215::doubao_seed_20_pro_260215": DOUBAO_PRO_RECOMMENDED_SCORE_THRESHOLD,
    "doubao_seed_20_mini_260215": DOUBAO_MINI_RECOMMENDED_SCORE_THRESHOLD,
    "doubao::doubao-seed-2-0-mini-260215::doubao_seed_20_mini_260215": DOUBAO_MINI_RECOMMENDED_SCORE_THRESHOLD,
    "doubao_seed_18_251228": DOUBAO_RECOMMENDED_SCORE_THRESHOLD,
    "doubao::doubao-seed-1-8-251228::doubao_seed_18_251228": DOUBAO_RECOMMENDED_SCORE_THRESHOLD,
    "doubao_seed_16_250615": DOUBAO_RECOMMENDED_SCORE_THRESHOLD,
    "doubao::doubao-seed-1.6-250615::doubao_seed_16_250615": DOUBAO_RECOMMENDED_SCORE_THRESHOLD,
    "doubao_seed_16_flash_250615": DOUBAO_RECOMMENDED_SCORE_THRESHOLD,
    "doubao::doubao-seed-1-6-flash-250615::doubao_seed_16_flash_250615": DOUBAO_RECOMMENDED_SCORE_THRESHOLD,
    "doubao_seed_16_vision_250815": DOUBAO_RECOMMENDED_SCORE_THRESHOLD,
    "doubao::doubao-seed-1-6-vision-250815::doubao_seed_16_vision_250815": DOUBAO_RECOMMENDED_SCORE_THRESHOLD,
}
MODEL_REFERENCE_UPDATE_MIN_SCORE_OVERRIDES = {
    "doubao_seed_20_lite_260215": DOUBAO_REFERENCE_UPDATE_MIN_SCORE,
    "doubao::doubao-seed-2-0-lite-260215::doubao_seed_20_lite_260215": DOUBAO_REFERENCE_UPDATE_MIN_SCORE,
    "doubao_seed_20_pro_260215": DOUBAO_PRO_REFERENCE_UPDATE_MIN_SCORE,
    "doubao::doubao-seed-2-0-pro-260215::doubao_seed_20_pro_260215": DOUBAO_PRO_REFERENCE_UPDATE_MIN_SCORE,
    "doubao_seed_20_mini_260215": DOUBAO_MINI_REFERENCE_UPDATE_MIN_SCORE,
    "doubao::doubao-seed-2-0-mini-260215::doubao_seed_20_mini_260215": DOUBAO_MINI_REFERENCE_UPDATE_MIN_SCORE,
    "doubao_seed_18_251228": DOUBAO_REFERENCE_UPDATE_MIN_SCORE,
    "doubao::doubao-seed-1-8-251228::doubao_seed_18_251228": DOUBAO_REFERENCE_UPDATE_MIN_SCORE,
    "doubao_seed_16_250615": DOUBAO_REFERENCE_UPDATE_MIN_SCORE,
    "doubao::doubao-seed-1.6-250615::doubao_seed_16_250615": DOUBAO_REFERENCE_UPDATE_MIN_SCORE,
    "doubao_seed_16_flash_250615": DOUBAO_REFERENCE_UPDATE_MIN_SCORE,
    "doubao::doubao-seed-1-6-flash-250615::doubao_seed_16_flash_250615": DOUBAO_REFERENCE_UPDATE_MIN_SCORE,
    "doubao_seed_16_vision_250815": DOUBAO_REFERENCE_UPDATE_MIN_SCORE,
    "doubao::doubao-seed-1-6-vision-250815::doubao_seed_16_vision_250815": DOUBAO_REFERENCE_UPDATE_MIN_SCORE,
}


def resolve_score_threshold_for_model(base_threshold: int, model_meta: dict | None = None) -> int:
    threshold = int(base_threshold)
    if not isinstance(model_meta, dict):
        return threshold

    model_id = str(model_meta.get("model_id") or "").strip()
    cfg_id = str(model_meta.get("id") or "").strip()

    if model_id in MODEL_SCORE_THRESHOLD_OVERRIDES:
        return int(MODEL_SCORE_THRESHOLD_OVERRIDES[model_id])
    if cfg_id in MODEL_SCORE_THRESHOLD_OVERRIDES:
        return int(MODEL_SCORE_THRESHOLD_OVERRIDES[cfg_id])
    return threshold


def resolve_reference_update_min_score_for_model(base_min_score: int, model_meta: dict | None = None) -> int:
    min_score = int(base_min_score)
    if not isinstance(model_meta, dict):
        return min_score

    model_id = str(model_meta.get("model_id") or "").strip()
    cfg_id = str(model_meta.get("id") or "").strip()

    if model_id in MODEL_REFERENCE_UPDATE_MIN_SCORE_OVERRIDES:
        return int(MODEL_REFERENCE_UPDATE_MIN_SCORE_OVERRIDES[model_id])
    if cfg_id in MODEL_REFERENCE_UPDATE_MIN_SCORE_OVERRIDES:
        return int(MODEL_REFERENCE_UPDATE_MIN_SCORE_OVERRIDES[cfg_id])
    return min_score


class SimpleRateLimiter:
    def __init__(self, rpm_limit: int, tpm_limit: int):
        self.rpm_limit = max(1, int(rpm_limit))
        self.tpm_limit = max(1, int(tpm_limit))
        self.req_times = deque()
        self.token_times = deque()
        self.lock = threading.Lock()

    def _cleanup(self, now: float):
        while self.req_times and now - self.req_times[0] >= 60.0:
            self.req_times.popleft()
        while self.token_times and now - self.token_times[0][0] >= 60.0:
            self.token_times.popleft()

    def _current_tokens(self):
        return int(sum(t for _, t in self.token_times))

    def wait(self, token_estimate: int):
        token_estimate = max(1, int(token_estimate))
        while True:
            with self.lock:
                now = time.time()
                self._cleanup(now)
                rpm_ok = len(self.req_times) < self.rpm_limit
                tpm_ok = (self._current_tokens() + token_estimate) <= self.tpm_limit
                if rpm_ok and tpm_ok:
                    self.req_times.append(now)
                    self.token_times.append((now, token_estimate))
                    return

                wait_rpm = 0.0
                wait_tpm = 0.0
                if not rpm_ok and self.req_times:
                    wait_rpm = max(0.0, 60.0 - (now - self.req_times[0]))
                if not tpm_ok and self.token_times:
                    wait_tpm = max(0.0, 60.0 - (now - self.token_times[0][0]))
            time.sleep(max(0.05, min(max(wait_rpm, wait_tpm), 3.0)))


def _model_id(model_cfg: dict) -> str:
    provider = model_cfg.get("provider", "unknown")
    name = model_cfg.get("name", "unknown")
    cfg_id = str(model_cfg.get("id") or "default")
    return f"{provider}::{name}::{cfg_id}"


def _request_model_name(model_cfg: dict) -> str:
    token = str(model_cfg.get("model") or "").strip()
    if token:
        return token
    return str(model_cfg.get("name") or "")


def _get_cached_request_model_name(model_cfg: dict) -> str:
    model_key = _model_id(model_cfg)
    with _MODEL_REQUEST_NAME_LOCK:
        cached = str(_MODEL_REQUEST_NAME_OVERRIDES.get(model_key) or "").strip()
    return cached


def _candidate_request_model_names(model_cfg: dict) -> list[str]:
    primary = str(_request_model_name(model_cfg) or "").strip()
    if not primary:
        return []

    out: list[str] = []

    def _push(name: str) -> None:
        token = str(name or "").strip()
        if not token:
            return
        if token not in out:
            out.append(token)

    cached = _get_cached_request_model_name(model_cfg)
    if cached:
        _push(cached)
    _push(primary)

    aliases = model_cfg.get("model_aliases")
    if isinstance(aliases, (list, tuple)):
        for item in aliases:
            _push(str(item))

    if _provider_name(model_cfg) == "doubao":
        replacement_pairs = [
            ("1-6", "1.6"),
            ("1.6", "1-6"),
            ("1-5", "1.5"),
            ("1.5", "1-5"),
        ]
        for src, dst in replacement_pairs:
            for base in list(out):
                if src in base:
                    _push(base.replace(src, dst))

    return out


def _remember_successful_request_model_name(model_cfg: dict, resolved_model_name: str) -> None:
    model_key = _model_id(model_cfg)
    primary = str(_request_model_name(model_cfg) or "").strip()
    resolved = str(resolved_model_name or "").strip()
    if not resolved:
        return

    with _MODEL_REQUEST_NAME_LOCK:
        if resolved == primary:
            _MODEL_REQUEST_NAME_OVERRIDES.pop(model_key, None)
        else:
            _MODEL_REQUEST_NAME_OVERRIDES[model_key] = resolved


def _should_retry_with_next_model_name_alias(model_cfg: dict, exc: Exception) -> bool:
    if _provider_name(model_cfg) != "doubao":
        return False
    status_code, detail = _extract_error_status_and_text(exc)
    decision = _classify_model_failure_for_circuit(status_code, detail)
    return bool(decision and decision[0] == "not_found_or_no_access")


def _iter_doubao_model_cfgs(model_chain: list[dict] | None = None) -> list[dict]:
    chain = model_chain if isinstance(model_chain, list) else MODEL_FALLBACKS
    return [cfg for cfg in chain if str(cfg.get("provider") or "").strip().lower() == "doubao"]


def _new_doubao_quota_state() -> dict:
    return {
        "exhausted": False,
        "reason": "",
        "next_recheck_ts": 0.0,
        "recheck_in_progress": False,
    }


def _new_model_circuit_state() -> dict:
    return {
        "open": False,
        "kind": "",
        "reason": "",
        "next_retry_ts": 0.0,
    }


def _sync_doubao_quota_states(reset: bool = False) -> None:
    doubao_keys = {_model_id(cfg) for cfg in _iter_doubao_model_cfgs()}
    with _DOUBAO_QUOTA_LOCK:
        stale_keys = [k for k in list(_DOUBAO_QUOTA_STATES.keys()) if k not in doubao_keys]
        for k in stale_keys:
            _DOUBAO_QUOTA_STATES.pop(k, None)

        for k in doubao_keys:
            if reset or k not in _DOUBAO_QUOTA_STATES:
                _DOUBAO_QUOTA_STATES[k] = _new_doubao_quota_state()


def _sync_model_circuit_states(reset: bool = False) -> None:
    model_keys = {_model_id(cfg) for cfg in MODEL_FALLBACKS}
    with _MODEL_CIRCUIT_LOCK:
        stale_keys = [k for k in list(_MODEL_CIRCUIT_STATES.keys()) if k not in model_keys]
        for k in stale_keys:
            _MODEL_CIRCUIT_STATES.pop(k, None)
            _MODEL_CIRCUIT_LAST_SKIP_LOG_TS.pop(k, None)

        for k in model_keys:
            if reset or k not in _MODEL_CIRCUIT_STATES:
                _MODEL_CIRCUIT_STATES[k] = _new_model_circuit_state()

        if reset:
            _MODEL_CIRCUIT_LAST_SKIP_LOG_TS.clear()


def _is_doubao_model_quota_exhausted(model_cfg: dict) -> bool:
    if not DOUBAO_QUOTA_FALLBACK_TO_ZHIPU:
        return False
    if str(model_cfg.get("provider") or "").strip().lower() != "doubao":
        return False

    with _DOUBAO_QUOTA_LOCK:
        state = _DOUBAO_QUOTA_STATES.setdefault(_model_id(model_cfg), _new_doubao_quota_state())
        return bool(state.get("exhausted"))


def _is_all_doubao_quota_exhausted() -> bool:
    if not DOUBAO_QUOTA_FALLBACK_TO_ZHIPU:
        return False

    doubao_cfgs = _iter_doubao_model_cfgs()
    if not doubao_cfgs:
        return False

    with _DOUBAO_QUOTA_LOCK:
        exhausted_flags = []
        for cfg in doubao_cfgs:
            state = _DOUBAO_QUOTA_STATES.setdefault(_model_id(cfg), _new_doubao_quota_state())
            exhausted_flags.append(bool(state.get("exhausted")))
    return bool(exhausted_flags) and all(exhausted_flags)


def _has_doubao_model_configured() -> bool:
    return bool(_iter_doubao_model_cfgs())


def _sum_configured_concurrency_slots(include_doubao: bool = True) -> int:
    slots = 0
    for cfg in MODEL_FALLBACKS:
        provider = str(cfg.get("provider") or "").strip().lower()
        if (not include_doubao) and provider == "doubao":
            continue
        slots += max(1, int(cfg.get("max_concurrency", PER_KEY_MAX_CONCURRENCY)))
    return max(1, int(slots))


def _refresh_max_concurrency_hint() -> None:
    """Keep upstream worker hint in sync with doubao availability state."""
    global MAX_CONCURRENCY
    configured_slots = _sum_configured_concurrency_slots(include_doubao=True)
    configured_slots_without_doubao = _sum_configured_concurrency_slots(include_doubao=False)

    if not _has_doubao_model_configured():
        MAX_CONCURRENCY = configured_slots
        return

    if not DOUBAO_QUOTA_FALLBACK_TO_ZHIPU:
        MAX_CONCURRENCY = configured_slots
        return

    _sync_doubao_quota_states(reset=False)
    with _DOUBAO_QUOTA_LOCK:
        exhausted_flags = [bool(v.get("exhausted")) for v in _DOUBAO_QUOTA_STATES.values()]
    doubao_exhausted = bool(exhausted_flags) and all(exhausted_flags)

    target_hint = configured_slots_without_doubao if doubao_exhausted else configured_slots
    MAX_CONCURRENCY = max(1, target_hint)


_MODEL_RUNTIME_LOCK = threading.Lock()
_MODEL_CHAIN_LOCK = threading.Lock()
_MODEL_CHAIN_CURSOR = 0
_MODEL_REQUEST_NAME_LOCK = threading.Lock()
_MODEL_REQUEST_NAME_OVERRIDES: dict[str, str] = {}

_DOUBAO_QUOTA_LOCK = threading.Lock()
_DOUBAO_QUOTA_STATES: dict[str, dict] = {}

_MODEL_CIRCUIT_LOCK = threading.Lock()
_MODEL_CIRCUIT_STATES: dict[str, dict] = {}
_MODEL_CIRCUIT_LAST_SKIP_LOG_TS: dict[str, float] = {}
_ACTIVE_MODEL_FORCE_FALLBACK_LOCK = threading.Lock()
_ACTIVE_MODEL_FORCE_FALLBACK_LAST_LOG_TS = 0.0

_SILICONFLOW_DEGRADE_LOCK = threading.Lock()
_SILICONFLOW_SINGLE_TASK_TOKENS: set[str] = set()

_SILICONFLOW_400_LOG_LOCK = threading.Lock()
_SILICONFLOW_400_LOGGED_SIGNATURES: set[str] = set()

_BASE_MODEL_FALLBACKS = [dict(cfg) for cfg in MODEL_FALLBACKS]


def _clone_model_cfgs(model_cfgs: list[dict]) -> list[dict]:
    return [dict(cfg) for cfg in model_cfgs]


def _rebuild_runtime_model_states(reset_doubao_quota: bool = False):
    global MODEL_LIMITERS, MODEL_SEMAPHORES, _MODEL_INFLIGHT_COUNTS, _MODEL_CHAIN_CURSOR
    global _SILICONFLOW_SINGLE_TASK_TOKENS
    global _MODEL_REQUEST_NAME_OVERRIDES

    MODEL_LIMITERS = {
        _model_id(cfg): SimpleRateLimiter(cfg.get("rpm", RPM_LIMIT), cfg.get("tpm", TPM_LIMIT))
        for cfg in MODEL_FALLBACKS
    }
    MODEL_SEMAPHORES = {
        _model_id(cfg): threading.BoundedSemaphore(max(1, int(cfg.get("max_concurrency", PER_KEY_MAX_CONCURRENCY))))
        for cfg in MODEL_FALLBACKS
    }

    with _MODEL_RUNTIME_LOCK:
        _MODEL_INFLIGHT_COUNTS = {_model_id(cfg): 0 for cfg in MODEL_FALLBACKS}

    with _MODEL_CHAIN_LOCK:
        _MODEL_CHAIN_CURSOR = 0

    with _MODEL_REQUEST_NAME_LOCK:
        _MODEL_REQUEST_NAME_OVERRIDES = {}

    _sync_doubao_quota_states(reset=bool(reset_doubao_quota))
    _sync_model_circuit_states(reset=bool(reset_doubao_quota))

    if reset_doubao_quota:
        with _SILICONFLOW_DEGRADE_LOCK:
            _SILICONFLOW_SINGLE_TASK_TOKENS = set()

    _refresh_max_concurrency_hint()


MODEL_LIMITERS = {}
MODEL_SEMAPHORES = {}
_MODEL_INFLIGHT_COUNTS = {}
_rebuild_runtime_model_states(reset_doubao_quota=True)


def _is_siliconflow_model_cfg(model_cfg: dict) -> bool:
    return str(model_cfg.get("provider") or "").strip().lower() == "siliconflow"


def _is_siliconflow_degraded_to_single() -> bool:
    with _SILICONFLOW_DEGRADE_LOCK:
        return bool(_SILICONFLOW_SINGLE_TASK_TOKENS)


def _enable_siliconflow_single_for_task(task_token: str, reason: str, request_tag: str) -> bool:
    token = str(task_token or "").strip()
    if not token:
        return False

    notify = False
    with _SILICONFLOW_DEGRADE_LOCK:
        if token in _SILICONFLOW_SINGLE_TASK_TOKENS:
            return False
        if not _SILICONFLOW_SINGLE_TASK_TOKENS:
            notify = True
        _SILICONFLOW_SINGLE_TASK_TOKENS.add(token)

    if notify:
        reason_text = str(reason or "").strip()[:240]
        print(
            "提示: "
            f"{request_tag} 检测到硅基流动连续重试压力过高，已将硅基并发临时降为1（仅当前任务生效）。"
            f"触发原因: {reason_text}"
        )
    return True


def _disable_siliconflow_single_for_task(task_token: str) -> None:
    token = str(task_token or "").strip()
    if not token:
        return

    notify = False
    with _SILICONFLOW_DEGRADE_LOCK:
        if token in _SILICONFLOW_SINGLE_TASK_TOKENS:
            _SILICONFLOW_SINGLE_TASK_TOKENS.discard(token)
            if not _SILICONFLOW_SINGLE_TASK_TOKENS:
                notify = True

    if notify:
        print("提示: 当前触发降并发的任务已结束，硅基并发恢复到配置值。")


def _get_effective_model_max_concurrency(model_cfg: dict) -> int:
    base_slots = max(1, int(model_cfg.get("max_concurrency", PER_KEY_MAX_CONCURRENCY)))
    if (not _is_siliconflow_model_cfg(model_cfg)) or (not SILICONFLOW_DYNAMIC_DEGRADE_ENABLED):
        return base_slots
    if _is_siliconflow_degraded_to_single():
        return min(base_slots, 1)
    return base_slots


@contextmanager
def _acquire_model_slot(model_cfg: dict):
    model_key = _model_id(model_cfg)
    sem = MODEL_SEMAPHORES.get(model_key)
    if sem is None:
        yield
        return
    sem.acquire()
    acquired_runtime_slot = False
    while True:
        with _MODEL_RUNTIME_LOCK:
            current_inflight = int(_MODEL_INFLIGHT_COUNTS.get(model_key, 0))
            allowed_slots = _get_effective_model_max_concurrency(model_cfg)
            if current_inflight < allowed_slots:
                _MODEL_INFLIGHT_COUNTS[model_key] = current_inflight + 1
                acquired_runtime_slot = True
                break
        time.sleep(0.01)
    try:
        yield
    finally:
        if acquired_runtime_slot:
            with _MODEL_RUNTIME_LOCK:
                _MODEL_INFLIGHT_COUNTS[model_key] = max(0, int(_MODEL_INFLIGHT_COUNTS.get(model_key, 0)) - 1)
        sem.release()


def _get_model_chain_round_robin(model_chain: list[dict] | None = None) -> list[dict]:
    """Return a rotated model list so concurrent requests are balanced across API keys."""
    global _MODEL_CHAIN_CURSOR
    models = list(model_chain) if isinstance(model_chain, list) else list(MODEL_FALLBACKS)
    if not models:
        return []

    with _MODEL_CHAIN_LOCK:
        start = _MODEL_CHAIN_CURSOR % len(models)
        _MODEL_CHAIN_CURSOR = (_MODEL_CHAIN_CURSOR + 1) % len(models)

    if start <= 0:
        return models
    return models[start:] + models[:start]


def _provider_name(model_cfg: dict) -> str:
    return str(model_cfg.get("provider") or "").strip().lower()


def _masked_key_tail(api_key: str | None) -> str:
    token = str(api_key or "").strip()
    if not token:
        return "<EMPTY>"
    if len(token) <= 6:
        return token
    return f"***{token[-6:]}"


def _build_request_kwargs(model_cfg: dict, timeout_sec: int) -> dict:
    _ = model_cfg
    return {"timeout": max(5, int(timeout_sec))}


def _extract_http_error_context(exc: Exception) -> tuple[int | None, str]:
    response = getattr(exc, "response", None)
    if response is None:
        return None, ""

    status_code: int | None = None
    try:
        status_code = int(response.status_code)
    except Exception:
        status_code = None

    detail = ""
    try:
        payload = response.json()
        detail = json.dumps(payload, ensure_ascii=False)
    except Exception:
        try:
            detail = str(response.text)
        except Exception:
            detail = ""

    detail = re.sub(r"\s+", " ", str(detail)).strip()
    if len(detail) > 1200:
        detail = detail[:1200] + " ...(truncated)"

    return status_code, detail


def _extract_error_status_and_text(exc: Exception) -> tuple[int | None, str]:
    status_code, detail = _extract_http_error_context(exc)
    parts = [str(exc)]
    if detail:
        parts.append(detail)
    merged = " | ".join([p for p in parts if p]).strip()
    return status_code, merged


def _log_siliconflow_400_detail_once(detail: str, request_tag: str) -> None:
    signature = str(detail or "").strip()[:600]
    if not signature:
        return

    should_log = False
    with _SILICONFLOW_400_LOG_LOCK:
        if signature not in _SILICONFLOW_400_LOGGED_SIGNATURES:
            _SILICONFLOW_400_LOGGED_SIGNATURES.add(signature)
            should_log = True
            if len(_SILICONFLOW_400_LOGGED_SIGNATURES) > 128:
                _SILICONFLOW_400_LOGGED_SIGNATURES.clear()

    if should_log:
        print(
            "[WARN] siliconflow 400 response body "
            f"request={request_tag}: {detail}"
        )


def _is_doubao_quota_exhausted_error(exc: Exception) -> bool:
    status_code, detail = _extract_error_status_and_text(exc)
    detail_lower = detail.lower()

    if any(str(kw).strip().lower() in detail_lower for kw in DOUBAO_QUOTA_ERROR_KEYWORDS):
        return True

    if status_code in DOUBAO_QUOTA_HTTP_STATUS:
        quota_hint_keywords = ["quota", "balance", "credit", "余额", "额度", "配额"]
        if any(kw in detail_lower for kw in quota_hint_keywords):
            return True

    return False


def _find_doubao_models() -> list[dict]:
    return _iter_doubao_model_cfgs(MODEL_FALLBACKS)


def _format_model_brief(model_cfg: dict) -> str:
    provider = _provider_name(model_cfg)
    name = str(model_cfg.get("name") or "unknown")
    mid = str(model_cfg.get("id") or "default")
    return f"{provider}/{name}/{mid}"


def _find_first_siliconflow_model() -> dict | None:
    for cfg in MODEL_FALLBACKS:
        if _provider_name(cfg) == "siliconflow":
            return cfg
    return None


def _classify_model_failure_for_circuit(status_code: int | None, message: str) -> tuple[str, float] | None:
    if not MODEL_CIRCUIT_BREAKER_ENABLED:
        return None

    msg = str(message or "").lower()
    if not msg:
        return None

    inferred_status = status_code
    if inferred_status is None:
        for code in (401, 403, 404, 429):
            if re.search(rf"\b{code}\b", msg):
                inferred_status = code
                break

    if any(kw in msg for kw in MODEL_CIRCUIT_NOT_FOUND_KEYWORDS):
        return "not_found_or_no_access", float(max(5, MODEL_CIRCUIT_NOT_FOUND_RECHECK_INTERVAL_SEC))

    if inferred_status == 404:
        if any(kw in msg for kw in ["not found", "does not exist", "unknown", "endpoint", "model"]):
            return "not_found_or_no_access", float(max(5, MODEL_CIRCUIT_NOT_FOUND_RECHECK_INTERVAL_SEC))

    if inferred_status in {401, 403}:
        return "auth_or_permission", float(max(5, MODEL_CIRCUIT_AUTH_RECHECK_INTERVAL_SEC))

    if any(kw in msg for kw in MODEL_CIRCUIT_AUTH_KEYWORDS) and inferred_status in {None, 400, 401, 403}:
        return "auth_or_permission", float(max(5, MODEL_CIRCUIT_AUTH_RECHECK_INTERVAL_SEC))

    if inferred_status == 429:
        if any(kw in msg for kw in MODEL_CIRCUIT_QUOTA_KEYWORDS):
            return "quota_or_credit", float(max(5, MODEL_CIRCUIT_QUOTA_RECHECK_INTERVAL_SEC))
        return "rate_limited", float(max(5, MODEL_CIRCUIT_RATE_LIMIT_RECHECK_INTERVAL_SEC))

    return None


def _mark_model_temporarily_unavailable(
    model_cfg: dict,
    reason: str,
    request_tag: str,
    kind: str,
    cooldown_sec: float,
) -> None:
    if not MODEL_CIRCUIT_BREAKER_ENABLED:
        return

    model_key = _model_id(model_cfg)
    model_brief = _format_model_brief(model_cfg)
    reason_text = str(reason or "").strip()[:240]
    notify = False

    with _MODEL_CIRCUIT_LOCK:
        now = time.time()
        state = _MODEL_CIRCUIT_STATES.setdefault(model_key, _new_model_circuit_state())
        prev_open = bool(state.get("open"))
        prev_next_retry_ts = float(state.get("next_retry_ts") or 0.0)
        next_retry_ts = now + float(max(5.0, cooldown_sec))

        state["open"] = True
        state["kind"] = str(kind)
        state["reason"] = reason_text
        state["next_retry_ts"] = next_retry_ts
        _MODEL_CIRCUIT_LAST_SKIP_LOG_TS[model_key] = now

        if (not prev_open) or (next_retry_ts > prev_next_retry_ts + 1.0):
            notify = True

    if notify:
        print(
            "提示: "
            f"{request_tag} 检测到模型失败，已临时下线 {model_brief}。"
            f"kind={kind}，将在 {int(max(1, cooldown_sec))} 秒后自动重试。"
            f"触发原因: {reason_text}"
        )


def _is_model_temporarily_unavailable(model_cfg: dict) -> bool:
    if not MODEL_CIRCUIT_BREAKER_ENABLED:
        return False
    provider = str(model_cfg.get("provider") or "").strip().lower()
    if provider not in MODEL_CIRCUIT_MANAGED_PROVIDERS:
        return False

    model_key = _model_id(model_cfg)
    model_brief = _format_model_brief(model_cfg)
    should_log_skip = False
    cooldown_remaining_sec = 0
    kind = ""

    with _MODEL_CIRCUIT_LOCK:
        state = _MODEL_CIRCUIT_STATES.setdefault(model_key, _new_model_circuit_state())
        if not bool(state.get("open")):
            return False

        now = time.time()
        next_retry_ts = float(state.get("next_retry_ts") or 0.0)

        if now >= next_retry_ts:
            state["open"] = False
            state["kind"] = ""
            state["reason"] = ""
            state["next_retry_ts"] = 0.0
            _MODEL_CIRCUIT_LAST_SKIP_LOG_TS.pop(model_key, None)
            print(f"提示: 模型冷却窗口结束，恢复尝试: {model_brief}")
            return False

        cooldown_remaining_sec = int(max(1, next_retry_ts - now))
        kind = str(state.get("kind") or "unknown")

        last_log_ts = float(_MODEL_CIRCUIT_LAST_SKIP_LOG_TS.get(model_key, 0.0))
        if now - last_log_ts >= float(max(5, MODEL_CIRCUIT_SKIP_LOG_INTERVAL_SEC)):
            _MODEL_CIRCUIT_LAST_SKIP_LOG_TS[model_key] = now
            should_log_skip = True

    if should_log_skip:
        print(
            "[INFO] 模型临时下线中，跳过路由: "
            f"{model_brief} | kind={kind} | 剩余={cooldown_remaining_sec}s"
        )
    return True


def _maybe_open_model_circuit(model_cfg: dict, exc: Exception, request_tag: str) -> None:
    if not MODEL_CIRCUIT_BREAKER_ENABLED:
        return
    provider = str(model_cfg.get("provider") or "").strip().lower()
    if provider not in MODEL_CIRCUIT_MANAGED_PROVIDERS:
        return

    status_code, detail = _extract_error_status_and_text(exc)
    decision = _classify_model_failure_for_circuit(status_code, detail)
    if decision is None:
        return

    kind, cooldown_sec = decision
    _mark_model_temporarily_unavailable(
        model_cfg=model_cfg,
        reason=detail,
        request_tag=request_tag,
        kind=kind,
        cooldown_sec=cooldown_sec,
    )


def _is_doubao_quota_limited_message(message: str) -> bool:
    msg = str(message or "").lower()
    if any(str(kw).strip().lower() in msg for kw in DOUBAO_QUOTA_ERROR_KEYWORDS):
        return True
    if any(f"{status}" in msg for status in DOUBAO_QUOTA_HTTP_STATUS):
        quota_hint_keywords = ["quota", "balance", "credit", "余额", "额度", "配额"]
        if any(kw in msg for kw in quota_hint_keywords):
            return True
    return False


def _mark_doubao_quota_exhausted(model_cfg: dict, reason: str, request_tag: str):
    if not DOUBAO_QUOTA_FALLBACK_TO_ZHIPU:
        return
    if _provider_name(model_cfg) != "doubao":
        return

    model_key = _model_id(model_cfg)
    model_brief = _format_model_brief(model_cfg)

    notify = False
    exhausted_all = False
    with _DOUBAO_QUOTA_LOCK:
        state = _DOUBAO_QUOTA_STATES.setdefault(model_key, _new_doubao_quota_state())
        if not bool(state.get("exhausted")):
            notify = True
        state["exhausted"] = True
        state["reason"] = str(reason or "").strip()[:240]
        state["next_recheck_ts"] = time.time() + float(DOUBAO_QUOTA_RECHECK_INTERVAL_SEC)
        state["recheck_in_progress"] = False

        exhausted_flags = []
        for cfg in _find_doubao_models():
            s = _DOUBAO_QUOTA_STATES.setdefault(_model_id(cfg), _new_doubao_quota_state())
            exhausted_flags.append(bool(s.get("exhausted")))
        exhausted_all = bool(exhausted_flags) and all(exhausted_flags)

    _refresh_max_concurrency_hint()

    if notify:
        reason_text = str(reason or "").strip()[:240]
        concurrency_without_doubao = _sum_configured_concurrency_slots(include_doubao=False)
        availability_text = (
            f"当前豆包模型均受限，上游并发提示已切换到 {int(concurrency_without_doubao)}。"
            if exhausted_all
            else "其余豆包模型继续可用。"
        )
        print(
            "提示: "
            f"{request_tag} 检测到豆包额度不足/耗尽，已临时下线模型 {model_brief}。"
            f"{availability_text}"
            f"将每 {int(DOUBAO_QUOTA_RECHECK_INTERVAL_SEC // 60)} 分钟自动重试该模型可用性。"
            f"触发原因: {reason_text}"
        )


def _probe_doubao_availability(model_cfg: dict, timeout_sec: int, request_tag: str) -> tuple[bool, bool, str]:
    if _provider_name(model_cfg) != "doubao":
        return False, False, "目标模型不是豆包"

    ok, _elapsed, err_msg = _send_one_probe_request(
        model_cfg,
        timeout_sec=max(5, int(timeout_sec)),
        probe_batch_images=1,
    )
    if ok:
        return True, False, ""

    detail = str(err_msg or "").strip()
    is_quota_limited = _is_doubao_quota_limited_message(detail)
    if is_quota_limited and detail:
        _mark_doubao_quota_exhausted(model_cfg=model_cfg, reason=detail, request_tag=request_tag)
    return False, is_quota_limited, detail


def _try_reactivate_doubao_if_due():
    if not DOUBAO_QUOTA_FALLBACK_TO_ZHIPU:
        return

    doubao_cfgs = _find_doubao_models()
    if not doubao_cfgs:
        return

    now = time.time()
    due_cfgs: list[dict] = []
    with _DOUBAO_QUOTA_LOCK:
        for cfg in doubao_cfgs:
            state = _DOUBAO_QUOTA_STATES.setdefault(_model_id(cfg), _new_doubao_quota_state())
            if not bool(state.get("exhausted")):
                continue
            if bool(state.get("recheck_in_progress")):
                continue
            if now < float(state.get("next_recheck_ts") or 0.0):
                continue
            state["recheck_in_progress"] = True
            state["next_recheck_ts"] = now + float(DOUBAO_QUOTA_RECHECK_INTERVAL_SEC)
            due_cfgs.append(cfg)

    if not due_cfgs:
        return

    for cfg in due_cfgs:
        model_key = _model_id(cfg)
        model_brief = _format_model_brief(cfg)
        try:
            ok, quota_limited, detail = _probe_doubao_availability(
                model_cfg=cfg,
                timeout_sec=int(DOUBAO_QUOTA_RECHECK_TIMEOUT_SEC),
                request_tag=f"doubao_hourly_recheck:{cfg.get('id', 'default')}",
            )

            if ok:
                with _DOUBAO_QUOTA_LOCK:
                    state = _DOUBAO_QUOTA_STATES.setdefault(model_key, _new_doubao_quota_state())
                    state["exhausted"] = False
                    state["reason"] = ""
                    state["next_recheck_ts"] = 0.0
                print(f"提示: 豆包模型每小时重试成功，已恢复加入动态调度: {model_brief}")
            elif quota_limited:
                with _DOUBAO_QUOTA_LOCK:
                    state = _DOUBAO_QUOTA_STATES.setdefault(model_key, _new_doubao_quota_state())
                    state["exhausted"] = True
                    if detail:
                        state["reason"] = detail[:240]
                print(f"提示: 豆包模型额度仍受限，继续下线并1小时后重试: {model_brief}")
            else:
                with _DOUBAO_QUOTA_LOCK:
                    state = _DOUBAO_QUOTA_STATES.setdefault(model_key, _new_doubao_quota_state())
                    state["exhausted"] = True
                    if detail:
                        state["reason"] = detail[:240]
                print(f"提示: 豆包模型每小时重试失败（非额度错误），本轮继续回退其它API: {model_brief}")
        finally:
            with _DOUBAO_QUOTA_LOCK:
                state = _DOUBAO_QUOTA_STATES.setdefault(model_key, _new_doubao_quota_state())
                state["recheck_in_progress"] = False

    _refresh_max_concurrency_hint()


def _maybe_degrade_siliconflow_by_retry_pressure(
    attempt: int,
    error_text: str,
    request_tag: str,
    task_token: str | None = None,
) -> bool:
    if not SILICONFLOW_DYNAMIC_DEGRADE_ENABLED:
        return False
    if int(attempt) < int(SILICONFLOW_RETRY_DEGRADE_THRESHOLD):
        return False

    msg = str(error_text or "").lower()
    if "siliconflow" not in msg:
        return False

    token = str(task_token or "").strip()
    if not token:
        return False

    return _enable_siliconflow_single_for_task(
        task_token=token,
        reason=str(error_text or ""),
        request_tag=request_tag,
    )


def _startup_probe_doubao_quota():
    if not DOUBAO_QUOTA_FALLBACK_TO_ZHIPU:
        return

    doubao_cfgs = _find_doubao_models()
    if not doubao_cfgs:
        return

    print("启动探测: 尝试调用豆包API以确认当前额度状态（逐模型）。")

    ok_count = 0
    quota_limited_count = 0
    transient_fail_count = 0
    for cfg in doubao_cfgs:
        model_key = _model_id(cfg)
        model_brief = _format_model_brief(cfg)
        ok, quota_limited, detail = _probe_doubao_availability(
            model_cfg=cfg,
            timeout_sec=int(DOUBAO_QUOTA_RECHECK_TIMEOUT_SEC),
            request_tag=f"doubao_startup_probe:{cfg.get('id', 'default')}",
        )

        if ok:
            ok_count += 1
            with _DOUBAO_QUOTA_LOCK:
                state = _DOUBAO_QUOTA_STATES.setdefault(model_key, _new_doubao_quota_state())
                state["exhausted"] = False
                state["reason"] = ""
                state["next_recheck_ts"] = 0.0
                state["recheck_in_progress"] = False
            print(f"启动探测结果: 豆包模型可用，纳入动态调度 -> {model_brief}")
            continue

        if quota_limited:
            quota_limited_count += 1
            print(f"启动探测结果: 豆包模型额度受限，暂时下线 -> {model_brief}")
            continue

        transient_fail_count += 1
        print(f"启动探测结果: 豆包模型当前不可用（非额度限制） -> {model_brief} | detail={detail}")
        decision = _classify_model_failure_for_circuit(None, detail)
        if decision is not None:
            kind, cooldown_sec = decision
            _mark_model_temporarily_unavailable(
                model_cfg=cfg,
                reason=detail,
                request_tag=f"doubao_startup_probe:{cfg.get('id', 'default')}",
                kind=kind,
                cooldown_sec=cooldown_sec,
            )

    _refresh_max_concurrency_hint()
    print(
        "启动探测汇总: "
        f"可用={ok_count}, 额度受限={quota_limited_count}, 临时失败={transient_fail_count}, "
        f"当前上游并发提示={MAX_CONCURRENCY}"
    )


def _active_models_for_request() -> list[dict]:
    global _ACTIVE_MODEL_FORCE_FALLBACK_LAST_LOG_TS

    _try_reactivate_doubao_if_due()
    _refresh_max_concurrency_hint()

    active_models = []
    non_doubao_non_quota_models = []
    for cfg in MODEL_FALLBACKS:
        provider = _provider_name(cfg)
        circuit_blocked = _is_model_temporarily_unavailable(cfg)
        quota_blocked = provider == "doubao" and _is_doubao_model_quota_exhausted(cfg)

        if provider != "doubao" and not quota_blocked:
            non_doubao_non_quota_models.append(cfg)

        if circuit_blocked:
            continue
        if quota_blocked:
            continue
        active_models.append(cfg)

    if active_models:
        return active_models

    # 兜底：若因熔断导致当前窗口无可用模型，优先强制放行非豆包主模型，避免任务整体中断。
    if non_doubao_non_quota_models:
        now = time.time()
        should_log = False
        with _ACTIVE_MODEL_FORCE_FALLBACK_LOCK:
            if now - float(_ACTIVE_MODEL_FORCE_FALLBACK_LAST_LOG_TS) >= float(max(5, ACTIVE_MODEL_FORCE_FALLBACK_LOG_INTERVAL_SEC)):
                _ACTIVE_MODEL_FORCE_FALLBACK_LAST_LOG_TS = now
                should_log = True
        if should_log:
            ids = ", ".join([str(cfg.get("id") or "unknown") for cfg in non_doubao_non_quota_models])
            print(
                "提示: 当前可用模型为空，已临时放行非豆包模型（忽略其熔断）以继续评审。"
                f"放行模型: {ids}"
            )
        return non_doubao_non_quota_models

    return active_models


def _get_model_slot_snapshot(model_cfg: dict) -> tuple[int, int, int]:
    model_key = _model_id(model_cfg)
    max_slots = _get_effective_model_max_concurrency(model_cfg)
    with _MODEL_RUNTIME_LOCK:
        inflight = int(_MODEL_INFLIGHT_COUNTS.get(model_key, 0))
    available = max(0, max_slots - inflight)
    return available, inflight, max_slots


def _order_models_by_idle(model_chain: list[dict]) -> list[dict]:
    rotated = _get_model_chain_round_robin(model_chain)

    def _sort_key(cfg: dict):
        available, inflight, max_slots = _get_model_slot_snapshot(cfg)
        return (-available, inflight, -max_slots)

    return sorted(rotated, key=_sort_key)


def _order_models_for_dispatch(model_chain: list[dict]) -> list[dict]:
    """
    优先填满智谱与硅基流动；当两者无空闲槽位时，溢出到豆包。
    若主模型仍有空闲，豆包只作为后备，不抢占主模型任务。
    """
    if not model_chain:
        return []

    primary_models = [cfg for cfg in model_chain if _provider_name(cfg) in {"zhipu", "siliconflow"}]
    doubao_models = [cfg for cfg in model_chain if _provider_name(cfg) == "doubao"]
    other_models = [cfg for cfg in model_chain if cfg not in primary_models and cfg not in doubao_models]

    primary_free = [cfg for cfg in primary_models if _get_model_slot_snapshot(cfg)[0] > 0]
    doubao_free = [cfg for cfg in doubao_models if _get_model_slot_snapshot(cfg)[0] > 0]
    primary_busy = [cfg for cfg in primary_models if cfg not in primary_free]
    doubao_busy = [cfg for cfg in doubao_models if cfg not in doubao_free]

    if primary_free:
        # 主模型存在空闲：优先使用主模型，豆包仅在主模型调用失败时参与回退。
        ordered = (
            _order_models_by_idle(primary_free)
            + _order_models_by_idle(doubao_free)
            + _order_models_by_idle(primary_busy)
            + _order_models_by_idle(doubao_busy)
        )
    elif doubao_free:
        # 主模型已满：将新增任务溢出给豆包。
        ordered = (
            _order_models_by_idle(doubao_free)
            + _order_models_by_idle(primary_busy)
            + _order_models_by_idle(doubao_busy)
        )
    else:
        # 全部繁忙时维持稳定回退顺序。
        ordered = _order_models_by_idle(primary_models + doubao_models)

    if other_models:
        ordered += _order_models_by_idle(other_models)
    return ordered


def _new_failover_state() -> dict:
    return {"pinned_model_id": ""}


def _call_with_tiered_failover(call_fn, failover_state: dict, request_tag: str):
    """按模型空闲槽位动态择优调用；支持按模型配置禁用回退并锁定重试。"""
    if not isinstance(failover_state, dict):
        failover_state = {}

    last_model_errors = []

    active_models = _active_models_for_request()
    pinned_model_id = str(failover_state.get("pinned_model_id") or "").strip()
    if pinned_model_id:
        pinned_models = [cfg for cfg in active_models if _model_id(cfg) == pinned_model_id]
        if pinned_models:
            active_models = pinned_models
        else:
            failover_state["pinned_model_id"] = ""

    if not active_models:
        if _is_all_doubao_quota_exhausted():
            raise RuntimeError("豆包模型额度已耗尽，且当前未配置其它可用模型。")
        raise RuntimeError("未配置任何可用评审模型。")

    dispatch_order = _order_models_for_dispatch(active_models)

    for mcfg in dispatch_order:
        provider = _provider_name(mcfg)
        mname = mcfg.get("name", "unknown")
        mid = mcfg.get("id", "default")
        try:
            if str(request_tag).startswith("batch["):
                print(
                    "[INFO] 批量评分路由: "
                    f"model={provider}/{mname}/{mid} | request={request_tag}"
                )
            result = call_fn(mcfg)
            return result, mcfg
        except Exception as model_e:
            status_code, error_detail = _extract_http_error_context(model_e)
            err_msg = f"{provider}/{mname}/{mid}: {model_e}"

            if provider == "siliconflow" and status_code == 400 and error_detail:
                _log_siliconflow_400_detail_once(error_detail, request_tag=request_tag)
                err_msg = f"{err_msg} | response={error_detail}"

            last_model_errors.append(err_msg)

            is_doubao_quota_error = provider == "doubao" and _is_doubao_quota_exhausted_error(model_e)
            if is_doubao_quota_error:
                _mark_doubao_quota_exhausted(model_cfg=mcfg, reason=err_msg, request_tag=request_tag)
            else:
                _maybe_open_model_circuit(model_cfg=mcfg, exc=model_e, request_tag=request_tag)

            if bool(mcfg.get("disable_failover_on_error", False)):
                failover_state["pinned_model_id"] = _model_id(mcfg)
                raise RuntimeError(
                    "多路API调用失败(已禁用回退): "
                    f"{err_msg} | request={request_tag}"
                ) from model_e

    if last_model_errors:
        raise RuntimeError("多路API调用失败: " + " | ".join(last_model_errors[-6:]))
    raise RuntimeError("多路API调用失败: 未找到可执行模型。")

# ==============================================================================
#                                 核心功能函数
# ==============================================================================

def _estimate_tokens(prompt: str) -> int:
    # 粗略估计：中英文混排按每4字符约1 token
    return max(64, len(prompt) // 4 + 128)


def _read_image_as_data_url(image_path: Path) -> str:
    mime_type = "image/jpeg" if image_path.suffix.lower() in [".jpg", ".jpeg"] else "image/png"
    image_b64 = base64.b64encode(image_path.read_bytes()).decode("utf-8")
    return f"data:{mime_type};base64,{image_b64}"


def _normalize_sample_type(sample_type: str | None) -> str | None:
    if sample_type is None:
        return None
    key = str(sample_type).strip().lower()
    if not key:
        return None
    return TYPE_ALIASES.get(key)


def _extract_json_string_field(line_text: str, field_name: str) -> str | None:
    m = re.search(rf'"{re.escape(field_name)}"\s*:\s*"([^"]+)"', str(line_text))
    if not m:
        return None
    return str(m.group(1))


def _extract_json_int_field(line_text: str, field_name: str) -> int | None:
    m = re.search(rf'"{re.escape(field_name)}"\s*:\s*(-?[0-9]+)', str(line_text))
    if not m:
        return None
    try:
        return int(m.group(1))
    except Exception:
        return None


def _normalize_reference_items(reference_items: list[dict] | None) -> list[dict]:
    out = []
    if not isinstance(reference_items, list):
        return out

    for i, item in enumerate(reference_items, start=1):
        if not isinstance(item, dict):
            continue
        image_data_url = item.get("image_data_url")
        if not isinstance(image_data_url, str) or not image_data_url.strip():
            continue
        label = item.get("label")
        if not isinstance(label, str) or not label.strip():
            label = f"参考样例{i}"
        image_name = item.get("image_name")
        if not isinstance(image_name, str):
            image_name = ""
        out_item = {
            "image_name": image_name,
            "label": label.strip(),
            "image_data_url": image_data_url.strip(),
        }

        dataset = item.get("dataset")
        if isinstance(dataset, str) and dataset.strip():
            out_item["dataset"] = dataset.strip()

        source_id = item.get("source_id")
        if isinstance(source_id, str) and source_id.strip():
            out_item["source_id"] = source_id.strip()

        score = item.get("score")
        try:
            score_int = int(score)
            if 0 <= score_int <= 100:
                out_item["score"] = score_int
        except Exception:
            pass

        quality = str(item.get("quality") or "").strip().lower()
        if quality in {"high", "low", "neutral"}:
            out_item["quality"] = quality

        priority = item.get("priority")
        try:
            out_item["priority"] = int(priority)
        except Exception:
            pass

        iteration = item.get("iteration")
        if isinstance(iteration, str) and iteration.strip():
            out_item["iteration"] = iteration.strip()

        out.append(out_item)
    return out


def _load_reference_items_from_iteration(sample_type: str) -> list[dict]:
    pool_cfg = TYPE_REFERENCE_POOLS.get(sample_type, [])
    if not pool_cfg:
        return []

    type_label = TYPE_REFERENCE_LABELS.get(sample_type, sample_type)
    items = []
    for source in pool_cfg:
        iteration_name = str(source.get("iteration") or "iteration_1")
        indexes = source.get("indexes") or []
        if not indexes:
            continue

        iteration_dir = TYPE_REFERENCE_ITERATION_DIRS.get(iteration_name, TYPE_REFERENCE_ITERATION_DIR)
        jsonl_path = iteration_dir / sample_type / "samples_filtered.jsonl"
        if not jsonl_path.exists():
            print(f"提示: 类型参考样本文件不存在，跳过: {jsonl_path}")
            continue

        try:
            lines = jsonl_path.read_text(encoding="utf-8", errors="ignore").splitlines()
        except Exception as exc:
            print(f"警告: 读取类型参考样本失败: {jsonl_path} | {exc}")
            continue

        for idx in indexes:
            if idx <= 0 or idx > len(lines):
                print(
                    "警告: 类型参考样本索引越界: "
                    f"type={sample_type}, iteration={iteration_name}, idx={idx}, total={len(lines)}"
                )
                continue

            line = str(lines[idx - 1]).strip()
            if not line:
                continue

            image_rel = _extract_json_string_field(line, "image")
            if not image_rel:
                print(
                    "警告: 类型参考样本缺少 image 字段: "
                    f"type={sample_type}, iteration={iteration_name}, idx={idx}"
                )
                continue

            image_rel = image_rel.replace("\\", "/")
            while "//" in image_rel:
                image_rel = image_rel.replace("//", "/")

            image_path = PROJECT_ROOT / image_rel.lstrip("/")
            if not image_path.exists():
                rel_obj = Path(image_rel)
                # 兼容旧数据中未包含 iteration_x 的相对路径。
                if len(rel_obj.parts) >= 2 and rel_obj.parts[0].lower() == "sample":
                    tail = Path(*rel_obj.parts[1:])
                    alt = PROJECT_ROOT / "Sample" / iteration_name / tail
                    if alt.exists():
                        image_path = alt
                if not image_path.exists():
                    alt2 = iteration_dir / sample_type / "image" / rel_obj.name
                    if alt2.exists():
                        image_path = alt2
            if not image_path.exists():
                print(f"警告: 类型参考图片不存在，跳过: {image_path}")
                continue

            score = _extract_json_int_field(line, "score")
            dataset = _extract_json_string_field(line, "dataset") or "unknown"
            source_id = _extract_json_string_field(line, "source_id") or image_path.stem
            score_text = "NA" if score is None else str(score)

            label = (
                f"{type_label}优质样本({iteration_name}): idx={idx}, dataset={dataset}, "
                f"score={score_text}, source={source_id}"
            )

            quality = "high"
            priority = 2 if (isinstance(score, int) and int(score) >= int(REFERENCE_SAMPLE_HIGH_SCORE_THRESHOLD)) else 1

            try:
                items.append(
                    {
                        "image_name": image_path.name,
                        "label": label,
                        "image_data_url": _read_image_as_data_url(image_path),
                        "dataset": dataset,
                        "source_id": source_id,
                        "score": (int(score) if isinstance(score, int) else None),
                        "quality": quality,
                        "priority": int(priority),
                        "iteration": iteration_name,
                    }
                )
            except Exception as exc:
                print(f"警告: 编码类型参考图片失败，跳过: {image_path} | {exc}")
                continue

    return items


def _load_legacy_reference_items(sample_type: str) -> list[dict]:
    dirs = LEGACY_REFERENCE_DIR_BY_TYPE.get(sample_type, [])
    if not dirs:
        return []

    items = []
    for ref_dir in dirs:
        if not ref_dir.exists():
            continue
        images = sorted(
            list(ref_dir.glob("*.png"))
            + list(ref_dir.glob("*.jpg"))
            + list(ref_dir.glob("*.jpeg")),
            key=lambda p: p.name,
        )
        for img in images[: max(1, int(LEGACY_REFERENCE_MAX_PER_DIR))]:
            try:
                items.append(
                    {
                        "image_name": img.name,
                        "label": f"历史人工参考样本: {img.stem}",
                        "image_data_url": _read_image_as_data_url(img),
                        "dataset": "legacy",
                        "source_id": img.stem,
                        "iteration": "legacy",
                    }
                )
            except Exception:
                continue
    return items


def _extract_dataset_from_image_name(image_name: str) -> str:
    stem = Path(str(image_name or "")).stem
    parts = [p for p in stem.split("_") if p]
    if len(parts) >= 3:
        return str(parts[2]).strip() or "manual_review"
    return "manual_review"


def _load_manual_review_score_map() -> dict[str, int]:
    global _MANUAL_REVIEW_SCORE_CACHE
    if isinstance(_MANUAL_REVIEW_SCORE_CACHE, dict):
        return dict(_MANUAL_REVIEW_SCORE_CACHE)

    score_map: dict[str, int] = {}
    xlsx_path = Path(MANUAL_REVIEW_TRAIN_SHEET_PATH)
    if xlsx_path.exists() and xlsx_path.is_file():
        try:
            sheet = _load_human_review_sheet(xlsx_path)
            records = list(sheet.get("records") or [])
            for rec in records:
                if not isinstance(rec, dict):
                    continue
                image_name = str(rec.get("image_name") or "").strip()
                if not image_name:
                    continue
                hv = _safe_float(rec.get("human_avg"))
                if hv is None:
                    continue
                score_map[image_name.lower()] = int(max(0, min(100, round(hv))))
        except Exception as exc:
            print(f"警告: 读取 Manual_Review 评分表失败: {xlsx_path} | {exc}")

    _MANUAL_REVIEW_SCORE_CACHE = dict(score_map)
    return dict(score_map)


def _load_manual_review_reference_items() -> list[dict]:
    global _MANUAL_REVIEW_REFERENCE_CACHE
    if isinstance(_MANUAL_REVIEW_REFERENCE_CACHE, list):
        return list(_MANUAL_REVIEW_REFERENCE_CACHE)

    score_map = _load_manual_review_score_map()
    items: list[dict] = []

    def _collect_from_dir(folder: Path, quality: str) -> None:
        if not folder.exists() or (not folder.is_dir()):
            return

        images = sorted(
            list(folder.glob("*.png")) + list(folder.glob("*.jpg")) + list(folder.glob("*.jpeg")),
            key=lambda p: p.name,
        )
        for img in images:
            try:
                score = score_map.get(img.name.lower())
                dataset = _extract_dataset_from_image_name(img.name)
                source_id = img.stem

                if quality == "high":
                    meets_threshold = isinstance(score, int) and score >= int(REFERENCE_SAMPLE_HIGH_SCORE_THRESHOLD)
                    quality_label = "人工高分正例"
                else:
                    meets_threshold = isinstance(score, int) and score <= int(REFERENCE_SAMPLE_LOW_SCORE_THRESHOLD)
                    quality_label = "人工低分反例"

                score_text = "NA" if score is None else str(int(score))
                priority = 3 if meets_threshold else (2 if isinstance(score, int) else 1)
                extra = ""
                if (not meets_threshold) and isinstance(score, int):
                    if quality == "high":
                        extra = f"(低于{REFERENCE_SAMPLE_HIGH_SCORE_THRESHOLD}补位)"
                    else:
                        extra = f"(高于{REFERENCE_SAMPLE_LOW_SCORE_THRESHOLD}补位)"

                items.append(
                    {
                        "image_name": img.name,
                        "label": (
                            f"{quality_label}: dataset={dataset}, score={score_text}, source={source_id}{extra}"
                        ),
                        "image_data_url": _read_image_as_data_url(img),
                        "dataset": dataset,
                        "source_id": source_id,
                        "score": (int(score) if isinstance(score, int) else None),
                        "quality": quality,
                        "priority": int(priority),
                        "iteration": "manual_review",
                    }
                )
            except Exception as exc:
                print(f"警告: 编码 Manual_Review 参考图失败，跳过: {img} | {exc}")
                continue

    _collect_from_dir(Path(MANUAL_REVIEW_POSITIVE_DIR), quality="high")
    _collect_from_dir(Path(MANUAL_REVIEW_NEGATIVE_DIR), quality="low")

    _MANUAL_REVIEW_REFERENCE_CACHE = _normalize_reference_items(items)
    return list(_MANUAL_REVIEW_REFERENCE_CACHE)


def _merge_reference_items(existing_items: list[dict], extra_items: list[dict], limit: int) -> list[dict]:
    merged = []
    seen = set()

    for item in list(existing_items) + list(extra_items):
        if not isinstance(item, dict):
            continue
        key = str(item.get("image_name") or "").strip().lower()
        if not key:
            key = str(item.get("image_data_url") or "").strip()
        if not key or key in seen:
            continue
        seen.add(key)
        merged.append(item)
        if len(merged) >= max(1, int(limit)):
            break
    return merged


def _extract_cfg_id_from_model_id(model_id: str | None) -> str:
    token = str(model_id or "").strip().lower()
    if not token:
        return ""
    if token.startswith("group::"):
        return token[len("group::") :]
    if "::" in token:
        return token.split("::")[-1]
    return token


def _resolve_reference_target_count_for_model(sample_type: str, model_id: str | None) -> int:
    base_limit = int(TYPE_REFERENCE_LIMITS.get(sample_type, TYPE_REFERENCE_TARGET_COUNT))
    cfg_id = _extract_cfg_id_from_model_id(model_id)
    if cfg_id == "siliconflow_glm_41v_9b_thinking":
        return max(1, min(base_limit, int(TYPE_REFERENCE_TARGET_COUNT_SILICONFLOW)))
    return max(1, int(base_limit))


def _resolve_reference_mix_counts_for_model(model_id: str | None, target_count: int) -> tuple[int, int]:
    target = max(1, int(target_count))
    cfg_id = _extract_cfg_id_from_model_id(model_id)

    if cfg_id == "siliconflow_glm_41v_9b_thinking":
        pref_high = int(REFERENCE_SAMPLE_SILICONFLOW_HIGH_COUNT)
        pref_low = int(REFERENCE_SAMPLE_SILICONFLOW_LOW_COUNT)
    else:
        pref_high = int(REFERENCE_SAMPLE_DEFAULT_HIGH_COUNT)
        pref_low = int(REFERENCE_SAMPLE_DEFAULT_LOW_COUNT)

    pref_sum = max(1, pref_high + pref_low)
    if target >= pref_sum:
        return max(0, pref_high), max(0, pref_low)

    half_high = target // 2
    half_low = target - half_high
    if target % 2 == 1:
        half_high = half_high + 1
        half_low = target - half_high
    return max(0, half_high), max(0, half_low)


def _reference_item_quality(item: dict) -> str:
    quality = str(item.get("quality") or "").strip().lower()
    if quality in {"high", "low", "neutral"}:
        return quality

    score = None
    try:
        score = int(item.get("score"))
    except Exception:
        score = None

    if isinstance(score, int):
        if score >= int(REFERENCE_SAMPLE_HIGH_SCORE_THRESHOLD):
            return "high"
        if score <= int(REFERENCE_SAMPLE_LOW_SCORE_THRESHOLD):
            return "low"

    label = str(item.get("label") or "").strip().lower()
    if any(k in label for k in ["低分", "反例", "negative", "bad"]):
        return "low"
    if any(k in label for k in ["高分", "优质", "正例", "positive", "good"]):
        return "high"
    return "neutral"


def _normalize_reference_model_key(model_id: str | None) -> str:
    token = str(model_id or "").strip().lower()
    if token.startswith("group::"):
        return token
    cfg_id = token.split("::")[-1] if token else ""
    shared_group = REFERENCE_POOL_SHARED_MODEL_GROUPS.get(cfg_id)
    if shared_group:
        return f"group::{shared_group}"
    return token if token else "__shared__"


def _reference_pool_key_for_model_cfg(model_cfg: dict) -> str:
    cfg_id = str(model_cfg.get("id") or "").strip().lower()
    shared_group = REFERENCE_POOL_SHARED_MODEL_GROUPS.get(cfg_id)
    if shared_group:
        return f"group::{shared_group}"
    return _model_id(model_cfg)


def _make_type_reference_cache_key(norm_type: str, model_id: str | None) -> str:
    return f"{str(norm_type).strip().lower()}::{_normalize_reference_model_key(model_id)}"


def _ensure_type_reference_cache(norm_type: str, model_id: str | None = None) -> list[dict]:
    cache_key = _make_type_reference_cache_key(norm_type, model_id)
    cached = _TYPE_REFERENCE_CACHE.get(cache_key)
    if cached is not None:
        return cached

    good_items = _load_reference_items_from_iteration(norm_type)
    legacy_items = _load_legacy_reference_items(norm_type)
    manual_items = _load_manual_review_reference_items()
    cached = _normalize_reference_items(good_items + legacy_items + manual_items)
    _TYPE_REFERENCE_CACHE[cache_key] = cached
    return cached


def _prune_reference_pool(items: list[dict], max_candidates: int) -> list[dict]:
    normalized = _normalize_reference_items(items)
    if not normalized:
        return []

    def _image_key(item: dict) -> str:
        key = str(item.get("image_name") or "").strip().lower()
        if not key:
            key = str(item.get("image_data_url") or "").strip()
        return key

    def _item_score(item: dict) -> int:
        try:
            return int(item.get("score"))
        except Exception:
            return -1

    def _item_priority(item: dict) -> int:
        try:
            return int(item.get("priority"))
        except Exception:
            return 0

    def _quality_rank(item: dict) -> int:
        quality = _reference_item_quality(item)
        if quality == "high":
            return 3
        if quality == "low":
            return 2
        return 1

    def _quality_score(item: dict) -> int:
        quality = _reference_item_quality(item)
        score = _item_score(item)
        if quality == "low" and score >= 0:
            return -int(score)
        return int(score)

    ranked = sorted(
        normalized,
        key=lambda d: (
            _quality_rank(d),
            _item_priority(d),
            _quality_score(d),
            str(d.get("label") or ""),
        ),
        reverse=True,
    )

    kept = []
    seen_images = set()
    for it in ranked:
        key = _image_key(it)
        if not key or key in seen_images:
            continue
        seen_images.add(key)
        kept.append(it)
        if len(kept) >= max(1, int(max_candidates)):
            break
    return kept


def register_high_score_reference(
    sample_type: str | None,
    image_path: str | Path,
    score: int | float,
    dataset: str | None = None,
    source_id: str | None = None,
    model_id: str | None = None,
    min_score: int = REFERENCE_UPDATE_MIN_SCORE,
) -> bool:
    """Register high-score sample as online reference candidate for the given type/model."""
    norm_type = _normalize_sample_type(sample_type)
    if norm_type is None:
        return False

    try:
        score_int = int(score)
    except Exception:
        return False
    if score_int < max(0, int(min_score)):
        return False

    path_obj = Path(image_path)
    if not path_obj.exists() or (not path_obj.is_file()):
        return False

    ds = str(dataset or "unknown").strip() or "unknown"
    src = str(source_id or path_obj.stem).strip() or path_obj.stem

    try:
        item = {
            "image_name": path_obj.name,
            "label": f"在线高分样本: dataset={ds}, score={score_int}, source={src}",
            "image_data_url": _read_image_as_data_url(path_obj),
            "dataset": ds,
            "source_id": src,
            "score": int(score_int),
            "quality": "high",
            "priority": 3,
            "iteration": "online",
        }
    except Exception as exc:
        print(f"警告: 在线提示样本编码失败，跳过: {path_obj} | {exc}")
        return False

    with _TYPE_REFERENCE_LOCK:
        cache_key = _make_type_reference_cache_key(norm_type, model_id)
        pool = list(_ensure_type_reference_cache(norm_type, model_id=model_id))
        pool.append(item)
        pool = _prune_reference_pool(pool, max_candidates=REFERENCE_POOL_MAX_CANDIDATES)
        _TYPE_REFERENCE_CACHE[cache_key] = pool
    return True


def register_high_score_reference_with_model_meta(
    sample_type: str | None,
    image_path: str | Path,
    score: int | float,
    dataset: str | None = None,
    source_id: str | None = None,
    model_meta: dict | None = None,
    min_score: int = REFERENCE_UPDATE_MIN_SCORE,
) -> bool:
    resolved_min_score = int(min_score)
    model_key = ""
    if isinstance(model_meta, dict):
        model_key = str(model_meta.get("model_id") or model_meta.get("id") or "").strip()
        resolved_min_score = resolve_reference_update_min_score_for_model(
            base_min_score=int(min_score),
            model_meta=model_meta,
        )

    return register_high_score_reference(
        sample_type=sample_type,
        image_path=image_path,
        score=score,
        dataset=dataset,
        source_id=source_id,
        model_id=(model_key or None),
        min_score=resolved_min_score,
    )


def _pick_reference_items_with_fixed_count(
    existing_items: list[dict],
    pool_items: list[dict],
    target_count: int,
    model_id: str | None = None,
) -> list[dict]:
    """Keep references at a fixed count and enforce high/low balance when possible."""
    target = max(1, int(target_count))
    target_high, target_low = _resolve_reference_mix_counts_for_model(model_id=model_id, target_count=target)
    chosen = _merge_reference_items(existing_items, [], limit=target)
    if len(chosen) >= target:
        return chosen[:target]

    def _image_key(item: dict) -> str:
        key = str(item.get("image_name") or "").strip().lower()
        if not key:
            key = str(item.get("image_data_url") or "").strip()
        return key

    def _dataset_key(item: dict) -> str:
        ds = str(item.get("dataset") or "").strip().lower()
        return ds if ds else "__unknown__"

    def _source_key(item: dict) -> str:
        src = str(item.get("source_id") or "").strip().lower()
        return src

    seen = set()
    seen_source = set()
    dataset_counts: dict[str, int] = {}
    for it in chosen:
        key = _image_key(it)
        if key:
            seen.add(key)
        src = _source_key(it)
        if src:
            seen_source.add(src)
        ds = _dataset_key(it)
        dataset_counts[ds] = dataset_counts.get(ds, 0) + 1

    candidates = []
    for it in pool_items:
        if not isinstance(it, dict):
            continue
        key = _image_key(it)
        if not key or key in seen:
            continue
        seen.add(key)
        candidates.append(it)

    def _score_key(item: dict, quality: str) -> tuple[int, int, str]:
        try:
            score = int(item.get("score"))
        except Exception:
            score = -1
        try:
            priority = int(item.get("priority"))
        except Exception:
            priority = 0

        if quality == "low" and score >= 0:
            score = -score
        return (priority, score, str(item.get("label") or ""))

    high_candidates = []
    low_candidates = []
    neutral_candidates = []
    for it in candidates:
        quality = _reference_item_quality(it)
        if quality == "high":
            high_candidates.append(it)
        elif quality == "low":
            low_candidates.append(it)
        else:
            neutral_candidates.append(it)

    high_candidates = sorted(high_candidates, key=lambda d: _score_key(d, "high"), reverse=True)
    low_candidates = sorted(low_candidates, key=lambda d: _score_key(d, "low"), reverse=True)
    random.shuffle(neutral_candidates)

    def _append_candidates(
        candidate_list: list[dict],
        relax_dataset: bool,
        relax_source: bool,
        max_take: int | None = None,
    ) -> tuple[list[dict], int]:
        remain = []
        took = 0
        max_take_val = None if max_take is None else max(0, int(max_take))

        for it in candidate_list:
            if len(chosen) >= target:
                remain.append(it)
                continue
            if max_take_val is not None and took >= max_take_val:
                remain.append(it)
                continue

            src = _source_key(it)
            if src and (not relax_source) and src in seen_source:
                remain.append(it)
                continue

            ds = _dataset_key(it)
            ds_cnt = dataset_counts.get(ds, 0)
            if (
                (not relax_dataset)
                and ds != "__unknown__"
                and ds_cnt >= max(1, int(REFERENCE_DIVERSITY_MAX_PER_DATASET))
            ):
                remain.append(it)
                continue

            chosen.append(it)
            if src:
                seen_source.add(src)
            dataset_counts[ds] = ds_cnt + 1
            took += 1

        return remain, took

    def _chosen_quality_count(tag: str) -> int:
        return sum(1 for it in chosen if _reference_item_quality(it) == tag)

    need_high = max(0, int(target_high) - _chosen_quality_count("high"))
    for relax_dataset, relax_source in [(False, False), (True, False), (True, True)]:
        if need_high <= 0:
            break
        high_candidates, took = _append_candidates(
            high_candidates,
            relax_dataset=relax_dataset,
            relax_source=relax_source,
            max_take=need_high,
        )
        need_high -= int(took)

    need_low = max(0, int(target_low) - _chosen_quality_count("low"))
    for relax_dataset, relax_source in [(False, False), (True, False), (True, True)]:
        if need_low <= 0:
            break
        low_candidates, took = _append_candidates(
            low_candidates,
            relax_dataset=relax_dataset,
            relax_source=relax_source,
            max_take=need_low,
        )
        need_low -= int(took)

    fallback_candidates = list(neutral_candidates) + list(high_candidates) + list(low_candidates)
    random.shuffle(fallback_candidates)
    for relax_dataset, relax_source in [(False, False), (True, False), (True, True)]:
        if len(chosen) >= target:
            break
        fallback_candidates, _ = _append_candidates(
            fallback_candidates,
            relax_dataset=relax_dataset,
            relax_source=relax_source,
            max_take=None,
        )

    # 如果依旧不足，则允许复用已选项，确保提示样本数量不下降。
    while len(chosen) < target and chosen:
        base = random.choice(chosen)
        clone = dict(base)
        clone["label"] = f"{str(base.get('label') or '参考样本')}（复用）"
        chosen.append(clone)

    high_selected = [it for it in chosen if _reference_item_quality(it) == "high"]
    low_selected = [it for it in chosen if _reference_item_quality(it) == "low"]
    other_selected = [it for it in chosen if _reference_item_quality(it) not in {"high", "low"}]

    ordered = []
    while (high_selected or low_selected) and len(ordered) < target:
        if high_selected and len(ordered) < target:
            ordered.append(high_selected.pop(0))
        if low_selected and len(ordered) < target:
            ordered.append(low_selected.pop(0))
    ordered.extend(high_selected)
    ordered.extend(low_selected)
    ordered.extend(other_selected)

    return ordered[:target]


def get_type_reference_items(
    sample_type: str | None,
    reference_items: list[dict] | None = None,
    model_id: str | None = None,
) -> list[dict] | None:
    """Build effective reference items by merging existing hints with type-specific good samples."""
    normalized_existing = _normalize_reference_items(reference_items)

    norm_type = _normalize_sample_type(sample_type)
    if norm_type is None:
        return normalized_existing or None

    with _TYPE_REFERENCE_LOCK:
        cached = list(_ensure_type_reference_cache(norm_type, model_id=model_id))

    base_limit = _resolve_reference_target_count_for_model(sample_type=norm_type, model_id=model_id)
    merged = _pick_reference_items_with_fixed_count(
        existing_items=normalized_existing,
        pool_items=cached,
        target_count=base_limit,
        model_id=model_id,
    )
    return merged or None


def _extract_json_text(raw_text: str) -> str:
    txt = str(raw_text).strip()
    txt = re.sub(r'```json\s*|\s*```', '', txt)
    return txt


def _extract_finish_reason(response) -> str:
    if response is None:
        return ""

    try:
        obj = response.json()
        choices = obj.get("choices") if isinstance(obj, dict) else None
        if isinstance(choices, list) and choices:
            fr = choices[0].get("finish_reason")
            if fr is not None:
                return str(fr).strip().lower()
    except Exception:
        pass

    try:
        raw = str(response.text)
        m = re.search(r'"finish_reason"\s*:\s*"([^"]+)"', raw, flags=re.IGNORECASE)
        if m:
            return str(m.group(1)).strip().lower()
    except Exception:
        pass

    return ""


def _is_length_truncation_response(response) -> bool:
    return _extract_finish_reason(response) == "length"


def _parse_score_from_text(model_text: str) -> int:
    txt = _extract_json_text(model_text)

    # 0) 纯数字输出
    pure_num = re.fullmatch(r'\s*([0-9]{1,3})\s*', txt)
    if pure_num:
        score = int(pure_num.group(1))
        if 0 <= score <= 100:
            return score

    # 1) 直接是 JSON
    try:
        obj = json.loads(txt)
        score = int(obj["score"])
        if 0 <= score <= 100:
            return score
    except Exception:
        pass

    # 2) 提取可能包在 <answer> 或混合文本中的 JSON 块
    candidates = re.findall(r"\{[\s\S]*?\}", txt)
    for cand in candidates:
        try:
            obj = json.loads(cand)
            score = int(obj["score"])
            if 0 <= score <= 100:
                return score
        except Exception:
            continue

    # 3) 最后兜底：直接抓取 score 数字
    m = re.search(r'"?score"?\s*[:=]\s*([0-9]{1,3})', txt, flags=re.IGNORECASE)
    if m:
        score = int(m.group(1))
        if 0 <= score <= 100:
            return score

    # 4) 常见文本分数字段：SCORE: 85 / 最终分数: 85 / 分数为85
    patterns = [
        r'\bscore\b\s*[:：=]\s*([0-9]{1,3})',
        r'最终分数\s*[:：=]\s*([0-9]{1,3})',
        r'分数\s*[:：=]\s*([0-9]{1,3})',
        r'分数为\s*([0-9]{1,3})',
        r'评分为\s*([0-9]{1,3})',
        r'<answer>\s*([0-9]{1,3})\s*</answer>',
        r'<\|begin_of_box\|>\s*([0-9]{1,3})\s*<\|end_of_box\|>',
    ]
    for pat in patterns:
        m = re.search(pat, txt, flags=re.IGNORECASE)
        if not m:
            continue
        score = int(m.group(1))
        if 0 <= score <= 100:
            return score

    # 5) 截断兜底：若最终分数缺失，尝试从 S1/S2/S3/S4/P 组件计算
    def _pick_component_score(label: str, text: str):
        comp_patterns = [
            rf'\b{label}\b[^\n\r]{{0,20}}[:：=]\s*([0-9]{{1,3}})',
            rf'\b{label}\b[^\n\r]{{0,20}}得分[^\n\r]{{0,10}}([0-9]{{1,3}})',
            rf'\b{label}\b\s*\(\s*([0-9]{{1,3}})\s*/',
            rf'\b{label}\b[\s\S]{{0,120}}?([0-9]{{1,3}})\s*分',
        ]
        for p in comp_patterns:
            mm = re.search(p, text, flags=re.IGNORECASE)
            if mm:
                return int(mm.group(1))
        return None

    s1 = _pick_component_score("S1", txt)
    s2 = _pick_component_score("S2", txt)
    s3 = _pick_component_score("S3", txt)
    s4 = _pick_component_score("S4", txt)
    p_score = _pick_component_score("P", txt)

    if None not in (s1, s2, s3, s4, p_score):
        merged = int(s1 + s2 + s3 + s4 - p_score)
        if 0 <= merged <= 100:
            return merged

    raise ValueError("无法从模型输出中解析 score")


def _call_one_model(
    image_data_url: str,
    model_cfg: dict,
    extra_prompt_text: str | None = None,
    reference_items: list[dict] | None = None,
    prompt_text: str | None = None,
    max_tokens_override: int | None = None,
):
    model_candidates = _candidate_request_model_names(model_cfg)
    if not model_candidates:
        raise RuntimeError(f"模型配置缺失可用 model/name: {_model_id(model_cfg)}")
    model_name = model_candidates[0]

    active_prompt = PROMPT
    if isinstance(prompt_text, str) and prompt_text.strip():
        active_prompt = prompt_text

    content_items = [{"type": "text", "text": active_prompt}]
    if extra_prompt_text:
        content_items.append({"type": "text", "text": extra_prompt_text})

    if reference_items:
        content_items.append(
            {
                "type": "text",
                "text": "以下是人工校准参考样例图（可能包含高分正例与低分反例），请先学习评分尺度。",
            }
        )
        for i, item in enumerate(reference_items, start=1):
            label = item.get("label") or f"参考样例{i}"
            ref_url = item.get("image_data_url")
            if not ref_url:
                continue
            content_items.append({"type": "text", "text": label})
            content_items.append({"type": "image_url", "image_url": {"url": ref_url}})

    content_items.append({"type": "text", "text": "以下是待评图片，请基于上述标准与参考样例给出分数。"})
    content_items.append({"type": "image_url", "image_url": {"url": image_data_url}})

    payload = {
        "model": model_name,
        "temperature": generation_config["temperature"],
        "top_p": generation_config["top_p"],
        "max_tokens": int(max_tokens_override) if max_tokens_override is not None else generation_config["max_tokens"],
        "messages": [
            {
                "role": "user",
                "content": content_items,
            }
        ],
    }

    extra_prompt = extra_prompt_text or ""
    ref_count = len(reference_items) if reference_items else 0
    # 多图场景下适度放大 token 估计，避免限流侧低估。
    token_est = _estimate_tokens(active_prompt + extra_prompt) + ref_count * 512 + 256
    limiter_key = _model_id(model_cfg)
    limiter = MODEL_LIMITERS.get(limiter_key, SimpleRateLimiter(RPM_LIMIT, TPM_LIMIT))
    limiter.wait(token_est)

    base_url = model_cfg.get("base_url")
    api_key = model_cfg.get("api_key")
    if not base_url or not api_key:
        raise RuntimeError(f"模型配置缺失 base_url 或 api_key: {limiter_key}")

    request_timeout = max(5, int(model_cfg.get("timeout_sec_single", 120)))
    request_kwargs = _build_request_kwargs(model_cfg, request_timeout)

    with _acquire_model_slot(model_cfg):
        response = None
        last_http_exc = None
        for idx, candidate_model_name in enumerate(model_candidates):
            payload["model"] = candidate_model_name
            try:
                response = requests.post(
                    base_url,
                    headers={
                        "Authorization": f"Bearer {api_key}",
                        "Content-Type": "application/json",
                    },
                    json=payload,
                    **request_kwargs,
                )
                response.raise_for_status()
                _remember_successful_request_model_name(model_cfg, candidate_model_name)
                if idx > 0:
                    print(
                        "提示: "
                        f"{_format_model_brief(model_cfg)} 自动切换可用请求model为 {candidate_model_name}"
                    )
                break
            except requests.HTTPError as http_exc:
                last_http_exc = http_exc
                if idx < len(model_candidates) - 1 and _should_retry_with_next_model_name_alias(model_cfg, http_exc):
                    continue
                raise

        if response is None and last_http_exc is not None:
            raise last_http_exc

    resp_obj = response.json()
    model_text = resp_obj["choices"][0]["message"]["content"]
    return _extract_json_text(model_text), response


def _call_one_model_batch(
    image_items: list[dict],
    model_cfg: dict,
    extra_prompt_text: str | None = None,
    reference_items: list[dict] | None = None,
    max_tokens_override: int | None = None,
):
    if not image_items:
        raise ValueError("批量评分输入为空")

    model_candidates = _candidate_request_model_names(model_cfg)
    if not model_candidates:
        raise RuntimeError(f"模型配置缺失可用 model/name: {_model_id(model_cfg)}")
    model_name = model_candidates[0]
    provider = _provider_name(model_cfg)

    effective_reference_items = reference_items
    if provider == "siliconflow" and isinstance(reference_items, list):
        ref_cap = max(0, int(SILICONFLOW_BATCH_MAX_REFERENCE_IMAGES))
        if len(reference_items) > ref_cap:
            effective_reference_items = list(reference_items[:ref_cap])
    content_items = [{"type": "text", "text": BATCH_PROMPT}]
    if extra_prompt_text:
        content_items.append({"type": "text", "text": extra_prompt_text})

    if effective_reference_items:
        content_items.append(
            {
                "type": "text",
                "text": "以下是人工校准参考样例图（可能包含高分正例与低分反例），请先学习评分尺度。",
            }
        )
        for i, item in enumerate(effective_reference_items, start=1):
            label = item.get("label") or f"参考样例{i}"
            ref_url = item.get("image_data_url")
            if not ref_url:
                continue
            content_items.append({"type": "text", "text": label})
            content_items.append({"type": "image_url", "image_url": {"url": ref_url}})

    content_items.append(
        {
            "type": "text",
            "text": (
                "下面是多张待评图片。请逐图打分，并严格按JSON输出："
                "{\"scores\":[{\"id\":\"img1\",\"score\":85},{\"id\":\"img2\",\"score\":73}]}. "
                "必须保留每张图片对应的id，score必须是0-100整数。"
            ),
        }
    )

    for item in image_items:
        item_id = str(item.get("id") or "").strip()
        item_url = item.get("image_data_url")
        if not item_id or not item_url:
            continue
        item_name = str(item.get("name") or "")
        content_items.append({"type": "text", "text": f"待评图片 id={item_id}, name={item_name}"})
        content_items.append({"type": "image_url", "image_url": {"url": item_url}})

    batch_temperature = 0.0
    if provider == "siliconflow":
        # SiliconFlow 某些视觉模型会拒绝 temperature=0.0（返回 400/code=20015）。
        try:
            batch_temperature = max(0.1, float(generation_config.get("temperature", 0.1)))
        except Exception:
            batch_temperature = 0.1

    payload = {
        "model": model_name,
        "temperature": batch_temperature,
        "top_p": 1.0,
        "max_tokens": (
            int(max_tokens_override)
            if max_tokens_override is not None
            else min(int(generation_config["max_tokens"]), BATCH_MAX_TOKENS)
        ),
        "messages": [
            {
                "role": "user",
                "content": content_items,
            }
        ],
    }

    extra_prompt = extra_prompt_text or ""
    ref_count = len(effective_reference_items) if effective_reference_items else 0
    item_count = len(image_items)
    token_est = _estimate_tokens(BATCH_PROMPT + extra_prompt) + (ref_count + item_count) * 512 + 512

    limiter_key = _model_id(model_cfg)
    limiter = MODEL_LIMITERS.get(limiter_key, SimpleRateLimiter(RPM_LIMIT, TPM_LIMIT))
    limiter.wait(token_est)

    base_url = model_cfg.get("base_url")
    api_key = model_cfg.get("api_key")
    if not base_url or not api_key:
        raise RuntimeError(f"模型配置缺失 base_url 或 api_key: {limiter_key}")

    request_timeout = max(5, int(model_cfg.get("timeout_sec_batch", 180)))
    request_kwargs = _build_request_kwargs(model_cfg, request_timeout)

    with _acquire_model_slot(model_cfg):
        response = None
        last_http_exc = None
        for idx, candidate_model_name in enumerate(model_candidates):
            payload["model"] = candidate_model_name
            try:
                response = requests.post(
                    base_url,
                    headers={
                        "Authorization": f"Bearer {api_key}",
                        "Content-Type": "application/json",
                    },
                    json=payload,
                    **request_kwargs,
                )
                response.raise_for_status()
                _remember_successful_request_model_name(model_cfg, candidate_model_name)
                if idx > 0:
                    print(
                        "提示: "
                        f"{_format_model_brief(model_cfg)} 自动切换可用请求model为 {candidate_model_name}"
                    )
                break
            except requests.HTTPError as http_exc:
                last_http_exc = http_exc
                if idx < len(model_candidates) - 1 and _should_retry_with_next_model_name_alias(model_cfg, http_exc):
                    continue
                raise

        if response is None and last_http_exc is not None:
            raise last_http_exc

    resp_obj = response.json()
    model_text = resp_obj["choices"][0]["message"]["content"]
    return _extract_json_text(model_text), response


def _parse_batch_scores_from_text(model_text: str, item_ids: list[str]) -> dict[str, int]:
    txt = _extract_json_text(model_text)
    expected = {str(i).strip() for i in item_ids if str(i).strip()}
    out = {}

    def _save_one(item_id: str, score_val):
        item_id = str(item_id).strip()
        if item_id not in expected:
            return
        try:
            score_int = int(score_val)
        except Exception:
            return
        if 0 <= score_int <= 100:
            out[item_id] = score_int

    # 1) 直接JSON解析
    try:
        obj = json.loads(txt)
        if isinstance(obj, dict):
            for k, v in obj.items():
                _save_one(k, v)
            if isinstance(obj.get("scores"), list):
                for it in obj["scores"]:
                    if isinstance(it, dict):
                        _save_one(it.get("id"), it.get("score"))
            if isinstance(obj.get("score_map"), dict):
                for k, v in obj["score_map"].items():
                    _save_one(k, v)
        elif isinstance(obj, list):
            for it in obj:
                if isinstance(it, dict):
                    _save_one(it.get("id"), it.get("score"))
    except Exception:
        pass

    # 2) 提取嵌套JSON块中的 id/score 对
    candidates = re.findall(r"\{[\s\S]*?\}", txt)
    for cand in candidates:
        try:
            obj = json.loads(cand)
            if isinstance(obj, dict):
                if "id" in obj and "score" in obj:
                    _save_one(obj.get("id"), obj.get("score"))
                if isinstance(obj.get("scores"), list):
                    for it in obj["scores"]:
                        if isinstance(it, dict):
                            _save_one(it.get("id"), it.get("score"))
        except Exception:
            continue

    # 3) 文本兜底：按 id 抓取分数
    for item_id in expected:
        if item_id in out:
            continue
        pats = [
            rf'"id"\s*[:=]\s*"?{re.escape(item_id)}"?[\s\S]{{0,120}}?"score"\s*[:=]\s*([0-9]{{1,3}})',
            rf'"{re.escape(item_id)}"\s*[:=]\s*([0-9]{{1,3}})',
            rf'\b{re.escape(item_id)}\b\s*[:：=]\s*([0-9]{{1,3}})',
            rf'{re.escape(item_id)}[^\n\r]{{0,40}}(?:score|分数)?\s*[:：=]\s*([0-9]{{1,3}})',
        ]
        for pat in pats:
            m = re.search(pat, txt, flags=re.IGNORECASE)
            if not m:
                continue
            _save_one(item_id, m.group(1))
            if item_id in out:
                break

    return out


def get_batch_image_scores(
    image_paths: list[Path],
    extra_prompt_text: str | None = None,
    reference_items: list[dict] | None = None,
    sample_type: str | None = None,
    retries: int | None = None,
    return_meta: bool = False,
) -> dict[str, int] | tuple[dict[str, int], dict | None]:
    """批量评分，返回 {image_name: score}。失败项可由上层再单图兜底。"""
    def _pack(scores: dict[str, int], meta: dict | None):
        return (scores, meta) if return_meta else scores

    if not image_paths:
        return _pack({}, None)

    image_items = []
    for idx, img_path in enumerate(image_paths, start=1):
        image_items.append(
            {
                "id": f"img{idx}",
                "name": img_path.name,
                "image_data_url": _read_image_as_data_url(img_path),
            }
        )

    id_to_name = {it["id"]: it["name"] for it in image_items}
    max_retries = MAX_RETRIES if retries is None else retries
    infinite_retry = RETRY_FOREVER and max_retries is None
    length_failures = 0
    adaptive_batch_max_tokens = min(int(generation_config["max_tokens"]), BATCH_MAX_TOKENS)
    failover_state = _new_failover_state()
    siliconflow_ref_cap = max(1, int(TYPE_REFERENCE_TARGET_COUNT_SILICONFLOW))
    request_task_token = f"batch::{threading.get_ident()}::{time.time_ns()}::{len(image_items)}"
    siliconflow_task_degraded = False

    try:
        attempt = 0
        while True:
            attempt += 1
            response = None
            request_tag = f"batch[{len(image_items)}图]"
            print(f"批量评分请求: 第 {attempt} 次尝试 | 待评图片 {len(image_items)} 张 | model=auto_route")
            try:
                def _batch_call_one(mcfg: dict):
                    effective_reference_items = get_type_reference_items(
                        sample_type=sample_type,
                        reference_items=reference_items,
                        model_id=_model_id(mcfg),
                    )
                    if _provider_name(mcfg) == "siliconflow" and isinstance(effective_reference_items, list):
                        old_cnt = len(effective_reference_items)
                        if old_cnt > siliconflow_ref_cap:
                            effective_reference_items = effective_reference_items[:siliconflow_ref_cap]
                    return _call_one_model_batch(
                        image_items,
                        mcfg,
                        extra_prompt_text=extra_prompt_text,
                        reference_items=effective_reference_items,
                        max_tokens_override=adaptive_batch_max_tokens,
                    )

                (cleaned_response, response), used_model_cfg = _call_with_tiered_failover(
                    _batch_call_one,
                    failover_state,
                    request_tag=request_tag,
                )

                parsed = _parse_batch_scores_from_text(cleaned_response, list(id_to_name.keys()))
                if not parsed:
                    if _is_length_truncation_response(response):
                        length_failures += 1
                        if length_failures >= BATCH_TRUNCATION_FAIL_LIMIT:
                            print("警告: 批量响应连续被 length 截断，当前批次改为单图兜底。")
                            return _pack({}, None)
                        adaptive_batch_max_tokens = min(3072, max(adaptive_batch_max_tokens, BATCH_MAX_TOKENS * 2))
                        print(
                            "警告: 批量响应被 length 截断且未解析到分数，"
                            f"下次尝试提高 max_tokens 到 {adaptive_batch_max_tokens}。"
                        )
                    raise ValueError("批量输出未解析到任何 score")

                out = {}
                for item_id, score in parsed.items():
                    img_name = id_to_name.get(item_id)
                    if img_name is not None:
                        out[img_name] = int(score)
                model_meta = {
                    "model_id": _model_id(used_model_cfg),
                    "provider": used_model_cfg.get("provider"),
                    "name": used_model_cfg.get("name"),
                    "id": used_model_cfg.get("id", "default"),
                }
                print(
                    f"批量评分成功: 第 {attempt} 次尝试 | 命中 {len(out)}/{len(image_items)} 张 "
                    f"| model={used_model_cfg.get('provider')}/{used_model_cfg.get('name')}/{used_model_cfg.get('id', 'default')}"
                )
                return _pack(out, model_meta)

            except (json.JSONDecodeError, KeyError, ValueError) as e:
                raw = ""
                try:
                    raw = response.text[:600]
                except Exception:
                    raw = "<no response text>"
                print(f"警告: 批量解析失败: {e}。响应片段: {raw}")
            except KeyboardInterrupt:
                raise
            except Exception as e:
                print(f"错误: 批量评分调用失败: {e}")
                degraded_now = _maybe_degrade_siliconflow_by_retry_pressure(
                    attempt=attempt,
                    error_text=str(e),
                    request_tag=request_tag,
                    task_token=request_task_token,
                )
                siliconflow_task_degraded = siliconflow_task_degraded or bool(degraded_now)
                err_lower = str(e).lower()
                if (
                    "siliconflow" in err_lower
                    and "max_prompt_tokens" in err_lower
                    and siliconflow_ref_cap > 3
                ):
                    old_cap = siliconflow_ref_cap
                    siliconflow_ref_cap = max(3, siliconflow_ref_cap - 1)
                    print(
                        "[INFO] siliconflow prompt overflow detected, "
                        f"reduce batch ref cap {old_cap}->{siliconflow_ref_cap} and retry"
                    )

            if not infinite_retry and isinstance(max_retries, int) and max_retries > 0 and attempt >= max_retries:
                print(f"失败: 批量评分达到最大重试次数({max_retries})，返回空结果。")
                return _pack({}, None)

            wait_seconds = _compute_retry_delay(attempt, None, response=response)
            print(f"批量评分将在 {wait_seconds:.1f} 秒后重试（第 {attempt} 次失败）...")
            time.sleep(wait_seconds)
    finally:
        if siliconflow_task_degraded:
            _disable_siliconflow_single_for_task(request_task_token)


def _check_one_model_available(model_cfg: dict) -> tuple[bool, str]:
    base_url = model_cfg.get("base_url")
    api_key = model_cfg.get("api_key")
    provider = model_cfg.get("provider", "unknown")
    model_name = model_cfg.get("name", "unknown")
    model_candidates = _candidate_request_model_names(model_cfg)
    if not model_candidates:
        return False, f"{provider}/{model_name} 配置缺失可用 model/name"
    if not base_url or not api_key:
        return False, f"{provider}/{model_name} 配置缺失 base_url 或 api_key"

    candidate_image_urls = list(PRECHECK_IMAGE_URLS)
    if isinstance(PRECHECK_IMAGE_DATA_URL, str) and PRECHECK_IMAGE_DATA_URL.strip():
        candidate_image_urls.append(PRECHECK_IMAGE_DATA_URL.strip())

    last_err = None
    for img_url in candidate_image_urls:
        payload = {
            "model": model_candidates[0],
            "temperature": 0.1,
            "top_p": 1.0,
            "max_tokens": 32,
            "messages": [
                {
                    "role": "user",
                    "content": [
                        {"type": "text", "text": "请仅回复OK"},
                        {"type": "image_url", "image_url": {"url": img_url}},
                    ],
                }
            ],
        }

        try:
            request_kwargs = _build_request_kwargs(
                model_cfg,
                int(model_cfg.get("timeout_sec_single", PRECHECK_TIMEOUT)),
            )
            last_http_exc = None
            for idx, candidate_model_name in enumerate(model_candidates):
                payload["model"] = candidate_model_name
                try:
                    response = requests.post(
                        base_url,
                        headers={
                            "Authorization": f"Bearer {api_key}",
                            "Content-Type": "application/json",
                        },
                        json=payload,
                        **request_kwargs,
                    )
                    response.raise_for_status()
                    data = response.json()
                    content = data["choices"][0]["message"]["content"]
                    if content is None:
                        last_err = f"{provider}/{model_name} 返回内容为空"
                        break
                    _remember_successful_request_model_name(model_cfg, candidate_model_name)
                    if idx > 0:
                        return True, f"{provider}/{model_name} 可用（自动切换请求model={candidate_model_name}）"
                    return True, f"{provider}/{model_name} 可用"
                except requests.HTTPError as http_exc:
                    last_http_exc = http_exc
                    if idx < len(model_candidates) - 1 and _should_retry_with_next_model_name_alias(model_cfg, http_exc):
                        continue
                    raise

            if last_http_exc is not None:
                raise last_http_exc
        except requests.HTTPError as e:
            body = ""
            try:
                body = e.response.text[:300] if e.response is not None else ""
            except Exception:
                body = ""
            detail = f"{e}; body={body}" if body else str(e)
            last_err = f"{provider}/{model_name} 不可用: {detail}"
        except Exception as e:
            last_err = f"{provider}/{model_name} 不可用: {e}"

    return False, (last_err or f"{provider}/{model_name} 不可用: 未知错误")


def _is_likely_server_side_issue(msg: str) -> bool:
    m = str(msg).lower()
    keywords = [
        "500",
        "502",
        "503",
        "504",
        "timeout",
        "timed out",
        "connection reset",
        "temporarily",
        "rate limit",
        "429",
        "server",
    ]
    return any(k in m for k in keywords)


def _is_likely_client_side_issue(msg: str) -> bool:
    m = str(msg).lower()
    keywords = [
        "400",
        "401",
        "403",
        "parameter is invalid",
        "invalid",
        "配置缺失",
        "api_key",
    ]
    return any(k in m for k in keywords)


def run_model_precheck_or_raise():
    print("\n--- 运行前模型检查开始 ---")
    ok_count = 0
    failed_server = []
    failed_client = []
    failed_other = []

    for cfg in MODEL_FALLBACKS:
        ok, msg = _check_one_model_available(cfg)
        print(msg)
        if ok:
            ok_count += 1
            continue

        if _is_likely_server_side_issue(msg):
            failed_server.append(msg)
        elif _is_likely_client_side_issue(msg):
            failed_client.append(msg)
        else:
            failed_other.append(msg)

    if failed_server:
        print("提示: 以下模型预检查失败更像服务端/网络波动，运行中仍会保留并按回退顺序尝试。")
        for item in failed_server:
            print(f"  - {item}")

    if failed_client:
        print("提示: 以下模型预检查失败更像请求参数/鉴权问题，建议后续排查。")
        for item in failed_client:
            print(f"  - {item}")

    if failed_other:
        print("提示: 以下模型预检查失败原因未明确。")
        for item in failed_other:
            print(f"  - {item}")

    if ok_count <= 0:
        all_failed = failed_server + failed_client + failed_other
        raise RuntimeError("模型可用性检查失败：没有任何可用模型，已中止运行。失败项: " + " | ".join(all_failed))

    print(f"--- 运行前模型检查结束：可用模型 {ok_count}/{len(MODEL_FALLBACKS)}，继续执行 ---\n")


def _parse_int_list(text: str | None) -> list[int]:
    if not isinstance(text, str) or not text.strip():
        return []
    out: list[int] = []
    for part in text.split(","):
        token = str(part).strip()
        if not token:
            continue
        try:
            v = int(token)
        except Exception:
            continue
        if v > 0:
            out.append(v)
    return out


def _resolve_runtime_models(runtime_model_id: str | None) -> list[dict]:
    token = str(runtime_model_id or "all").strip().lower()
    if token in {"", "all", "*"}:
        return _clone_model_cfgs(_BASE_MODEL_FALLBACKS)

    wanted = {s.strip().lower() for s in str(runtime_model_id).split(",") if s.strip()}
    selected = []
    for cfg in _BASE_MODEL_FALLBACKS:
        cfg_id = str(cfg.get("id", "")).strip().lower()
        cfg_name = str(cfg.get("name", "")).strip().lower()
        if cfg_id in wanted or cfg_name in wanted:
            selected.append(dict(cfg))
    return selected


def _apply_runtime_model_filter(runtime_model_id: str | None):
    global MODEL_FALLBACKS
    selected = _resolve_runtime_models(runtime_model_id)
    if not selected:
        raise ValueError(f"未匹配到运行时模型: {runtime_model_id}")

    MODEL_FALLBACKS = _clone_model_cfgs(selected)
    _rebuild_runtime_model_states(reset_doubao_quota=True)


def _resolve_probe_models(probe_model_id: str | None) -> list[dict]:
    token = str(probe_model_id or "all").strip().lower()
    if token in {"", "all", "*"}:
        return list(MODEL_FALLBACKS)

    wanted = {s.strip().lower() for s in str(probe_model_id).split(",") if s.strip()}
    selected = []
    for cfg in MODEL_FALLBACKS:
        cfg_id = str(cfg.get("id", "")).strip().lower()
        cfg_name = str(cfg.get("name", "")).strip().lower()
        if cfg_id in wanted or cfg_name in wanted:
            selected.append(cfg)
    return selected


def _send_one_probe_request(
    model_cfg: dict,
    timeout_sec: int,
    probe_batch_images: int,
    probe_semaphore: threading.BoundedSemaphore | None = None,
) -> tuple[bool, float, str]:
    """Send a lightweight request for availability/concurrency probing."""
    t0 = time.time()
    limiter_key = _model_id(model_cfg)

    try:
        base_url = model_cfg.get("base_url")
        api_key = model_cfg.get("api_key")
        if not base_url or not api_key:
            raise RuntimeError(f"模型配置缺失 base_url 或 api_key: {limiter_key}")
        model_candidates = _candidate_request_model_names(model_cfg)
        if not model_candidates:
            raise RuntimeError(f"模型配置缺失可用 model/name: {limiter_key}")

        probe_batch_images = max(1, int(probe_batch_images))

        limiter = MODEL_LIMITERS.get(limiter_key, SimpleRateLimiter(RPM_LIMIT, TPM_LIMIT))
        limiter.wait(_estimate_tokens("请仅回复OK") + 96)
        probe_candidates = list(PRECHECK_IMAGE_URLS)
        if isinstance(PRECHECK_IMAGE_DATA_URL, str) and PRECHECK_IMAGE_DATA_URL.strip():
            probe_candidates.append(PRECHECK_IMAGE_DATA_URL.strip())
        last_err = ""
        for probe_image_url in probe_candidates:
            content_items = [{"type": "text", "text": "请仅回复OK"}]
            for _ in range(probe_batch_images):
                content_items.append({"type": "image_url", "image_url": {"url": probe_image_url}})

            payload = {
                "model": model_candidates[0],
                "temperature": 0.1,
                "top_p": 1.0,
                "max_tokens": 32,
                "messages": [
                    {
                        "role": "user",
                        "content": content_items,
                    }
                ],
            }

            try:
                probe_request_kwargs = _build_request_kwargs(model_cfg, max(5, int(timeout_sec)))
                if probe_semaphore is not None:
                    with probe_semaphore:
                        last_http_exc = None
                        for idx, candidate_model_name in enumerate(model_candidates):
                            payload["model"] = candidate_model_name
                            try:
                                response = requests.post(
                                    base_url,
                                    headers={
                                        "Authorization": f"Bearer {api_key}",
                                        "Content-Type": "application/json",
                                    },
                                    json=payload,
                                    **probe_request_kwargs,
                                )
                                response.raise_for_status()
                                data = response.json()
                                _ = data["choices"][0]["message"]["content"]
                                _remember_successful_request_model_name(model_cfg, candidate_model_name)
                                if idx > 0:
                                    print(
                                        "提示: "
                                        f"{_format_model_brief(model_cfg)} 探测自动切换请求model为 {candidate_model_name}"
                                    )
                                return True, time.time() - t0, ""
                            except requests.HTTPError as http_exc:
                                last_http_exc = http_exc
                                if idx < len(model_candidates) - 1 and _should_retry_with_next_model_name_alias(model_cfg, http_exc):
                                    continue
                                raise
                        if last_http_exc is not None:
                            raise last_http_exc
                else:
                    with _acquire_model_slot(model_cfg):
                        last_http_exc = None
                        for idx, candidate_model_name in enumerate(model_candidates):
                            payload["model"] = candidate_model_name
                            try:
                                response = requests.post(
                                    base_url,
                                    headers={
                                        "Authorization": f"Bearer {api_key}",
                                        "Content-Type": "application/json",
                                    },
                                    json=payload,
                                    **probe_request_kwargs,
                                )
                                response.raise_for_status()
                                data = response.json()
                                _ = data["choices"][0]["message"]["content"]
                                _remember_successful_request_model_name(model_cfg, candidate_model_name)
                                if idx > 0:
                                    print(
                                        "提示: "
                                        f"{_format_model_brief(model_cfg)} 探测自动切换请求model为 {candidate_model_name}"
                                    )
                                return True, time.time() - t0, ""
                            except requests.HTTPError as http_exc:
                                last_http_exc = http_exc
                                if idx < len(model_candidates) - 1 and _should_retry_with_next_model_name_alias(model_cfg, http_exc):
                                    continue
                                raise
                        if last_http_exc is not None:
                            raise last_http_exc
            except requests.HTTPError as probe_http_exc:
                body = ""
                try:
                    body = probe_http_exc.response.text[:300] if probe_http_exc.response is not None else ""
                except Exception:
                    body = ""
                if body:
                    last_err = f"{probe_http_exc}; body={body}"
                else:
                    last_err = str(probe_http_exc)
            except Exception as probe_exc:
                last_err = str(probe_exc)

        raise RuntimeError(last_err or "probe request failed")
    except Exception as exc:
        return False, time.time() - t0, str(exc)


def run_api_concurrency_probe(
    total_requests: int,
    workers: int,
    timeout_sec: int,
    probe_model_id: str | None = "all",
    probe_batch_images: int = 1,
    probe_override_model_max_concurrency: int = 0,
) -> dict:
    """Probe API availability under concurrent load and report practical limits."""
    total_requests = max(1, int(total_requests))
    workers = max(1, int(workers))
    timeout_sec = max(5, int(timeout_sec))
    probe_batch_images = max(1, int(probe_batch_images))
    probe_override_model_max_concurrency = max(0, int(probe_override_model_max_concurrency))

    selected_models = _resolve_probe_models(probe_model_id)
    if not selected_models:
        raise ValueError(f"未匹配到待探测模型: {probe_model_id}")

    selected_ids = [_model_id(cfg) for cfg in selected_models]
    model_id_to_cfg = {_model_id(cfg): cfg for cfg in selected_models}
    probe_semaphore_map: dict[str, threading.BoundedSemaphore] = {}
    if probe_override_model_max_concurrency > 0:
        probe_semaphore_map = {
            mid: threading.BoundedSemaphore(probe_override_model_max_concurrency) for mid in selected_ids
        }

    print("\n--- API并发探测开始 ---")
    print(
        "探测参数: "
        f"requests={total_requests}, workers={workers}, timeout={timeout_sec}s, "
        f"model_keys={len(selected_models)}, probe_batch_images={probe_batch_images}, "
        f"override_model_max_concurrency={probe_override_model_max_concurrency if probe_override_model_max_concurrency > 0 else 'off'}"
    )
    print(f"探测模型: {', '.join(selected_ids)}")

    latencies = []
    ok_count = 0
    fail_count = 0
    model_ok_counts: dict[str, int] = {}
    model_fail_counts: dict[str, int] = {}
    fail_examples: list[str] = []

    def _task(_: int) -> dict:
        chain = _get_model_chain_round_robin(selected_models)
        if not chain:
            return {"ok": False, "model_id": "none", "latency": 0.0, "err": "selected_models为空"}

        cfg = chain[0]
        model_id = _model_id(cfg)
        probe_semaphore = probe_semaphore_map.get(model_id)
        ok, elapsed, err = _send_one_probe_request(
            cfg,
            timeout_sec=timeout_sec,
            probe_batch_images=probe_batch_images,
            probe_semaphore=probe_semaphore,
        )
        return {
            "ok": bool(ok),
            "model_id": model_id,
            "latency": float(elapsed),
            "err": str(err or ""),
        }

    with ThreadPoolExecutor(max_workers=workers) as executor:
        futures = [executor.submit(_task, i) for i in range(total_requests)]
        for fut in tqdm(as_completed(futures), total=len(futures), desc="API并发探测"):
            try:
                result = fut.result()
            except Exception as exc:
                fail_count += 1
                if len(fail_examples) < 5:
                    fail_examples.append(f"future异常: {exc}")
                continue

            model_id = result.get("model_id", "unknown")
            latencies.append(float(result.get("latency", 0.0)))

            if bool(result.get("ok", False)):
                ok_count += 1
                model_ok_counts[model_id] = int(model_ok_counts.get(model_id, 0)) + 1
            else:
                fail_count += 1
                model_fail_counts[model_id] = int(model_fail_counts.get(model_id, 0)) + 1
                err_msg = str(result.get("err", ""))
                if err_msg and len(fail_examples) < 5:
                    fail_examples.append(f"{model_id}: {err_msg}")

    elapsed = max(1e-6, float(sum(latencies)))
    avg_latency = float(np.mean(latencies)) if latencies else 0.0
    p95_latency = float(np.percentile(np.array(latencies, dtype=float), 95)) if latencies else 0.0
    success_ratio = float(ok_count) / float(total_requests)

    active_model_ids = [mid for mid, cnt in model_ok_counts.items() if int(cnt) > 0]
    active_models = len(active_model_ids)
    est_parallel_requests = int(
        sum(
            (
                int(probe_override_model_max_concurrency)
                if probe_override_model_max_concurrency > 0
                else int(model_id_to_cfg[mid].get("max_concurrency", PER_KEY_MAX_CONCURRENCY))
            )
            for mid in active_model_ids
            if mid in model_id_to_cfg
        )
    )
    est_parallel_images_current_probe = int(est_parallel_requests * probe_batch_images)
    est_parallel_images_batch5 = int(est_parallel_requests * HUMAN_EVAL_MAX_IMAGES_PER_REQUEST)

    summary = {
        "total_requests": int(total_requests),
        "workers": int(workers),
        "probe_batch_images": int(probe_batch_images),
        "probe_override_model_max_concurrency": int(probe_override_model_max_concurrency),
        "selected_model_ids": selected_ids,
        "ok_count": int(ok_count),
        "fail_count": int(fail_count),
        "success_ratio": round(success_ratio, 4),
        "avg_latency_sec": round(avg_latency, 4),
        "p95_latency_sec": round(p95_latency, 4),
        "active_models": int(active_models),
        "active_model_ids": active_model_ids,
        "model_ok_counts": model_ok_counts,
        "model_fail_counts": model_fail_counts,
        "estimated_parallel_requests": int(est_parallel_requests),
        "estimated_parallel_images_current_probe": int(est_parallel_images_current_probe),
        "estimated_parallel_images_batch5": int(est_parallel_images_batch5),
        "fail_examples": fail_examples,
    }

    print("并发探测结果:")
    print(json.dumps(summary, ensure_ascii=False, indent=2))
    print("--- API并发探测结束 ---\n")
    return summary


def _compute_retry_delay(attempt_index: int, base_delay: float | None = None, response=None) -> float:
    # 两阶段固定等待：前 N 次失败短等待，之后长等待
    if attempt_index <= RETRY_SHORT_ATTEMPTS:
        return float(RETRY_SHORT_DELAY)
    return float(RETRY_LONG_DELAY)


def get_image_score(
    image_path: Path,
    retries: int | None = None,
    delay: float | None = None,
    extra_prompt_text: str | None = None,
    reference_items: list[dict] | None = None,
    sample_type: str | None = None,
    return_meta: bool = False,
) -> int | None | tuple[int | None, dict | None]:
    """
    为单个图像评分。
    默认在失败时持续重试，前10次失败每次等待2秒，之后每次等待5秒。

    Args:
        image_path: 图像文件路径。
        retries: 失败后的最大重试次数。None 表示按全局配置控制。
        delay: 兼容保留参数，当前不使用。
        extra_prompt_text: 附加提示文本（用于人工评审校准信息）。
        reference_items: 参考样例图列表，每项包含 image_data_url/label。

    Returns:
        返回一个0-100的整数分数。若配置为有限重试且超过上限，返回None。
    """
    def _pack(score_val: int | None, meta: dict | None):
        return (score_val, meta) if return_meta else score_val

    image_data_url = _read_image_as_data_url(image_path)

    max_retries = MAX_RETRIES if retries is None else retries
    infinite_retry = RETRY_FOREVER and max_retries is None
    compact_mode = False
    drop_reference_mode = False
    truncation_failures = 0
    failover_state = _new_failover_state()
    siliconflow_ref_cap = max(1, int(TYPE_REFERENCE_TARGET_COUNT_SILICONFLOW))
    request_task_token = f"single::{threading.get_ident()}::{time.time_ns()}::{image_path.name}"
    siliconflow_task_degraded = False

    try:
        attempt = 0
        while True:
            attempt += 1
            response = None
            request_tag = f"single[{image_path.name}]"
            active_refs: list[dict] | None = None
            try:
                active_prompt = SINGLE_COMPACT_PROMPT if compact_mode else PROMPT
                active_max_tokens = SINGLE_COMPACT_MAX_TOKENS if compact_mode else int(generation_config["max_tokens"])

                def _single_call_one(mcfg: dict):
                    nonlocal active_refs
                    if drop_reference_mode:
                        active_refs = None
                    else:
                        active_refs = get_type_reference_items(
                            sample_type=sample_type,
                            reference_items=reference_items,
                            model_id=_model_id(mcfg),
                        )
                    if _provider_name(mcfg) == "siliconflow" and isinstance(active_refs, list):
                        if len(active_refs) > siliconflow_ref_cap:
                            active_refs = active_refs[:siliconflow_ref_cap]
                    return _call_one_model(
                        image_data_url,
                        mcfg,
                        extra_prompt_text=extra_prompt_text,
                        reference_items=active_refs,
                        prompt_text=active_prompt,
                        max_tokens_override=active_max_tokens,
                    )

                (cleaned_response, response), used_model_cfg = _call_with_tiered_failover(
                    _single_call_one,
                    failover_state,
                    request_tag=request_tag,
                )

                score = _parse_score_from_text(cleaned_response)
                if 0 <= score <= 100:
                    model_meta = {
                        "model_id": _model_id(used_model_cfg),
                        "provider": used_model_cfg.get("provider"),
                        "name": used_model_cfg.get("name"),
                        "id": used_model_cfg.get("id", "default"),
                    }
                    print(
                        "单图评分命中模型: "
                        f"{used_model_cfg.get('provider')}/{used_model_cfg.get('name')}/{used_model_cfg.get('id', 'default')}"
                    )
                    return _pack(score, model_meta)
                else:
                    print(f"警告: 模型返回了无效分数 {score} for {image_path.name}。将返回None。")
                    return _pack(None, None)

            except (json.JSONDecodeError, KeyError, ValueError) as e:
                raw = ""
                try:
                    raw = response.text
                except Exception:
                    raw = "<no response text>"
                print(f"警告: 解析模型对 {image_path.name} 的返回时出错: {e}。返回内容: '{raw}'")

                if _is_length_truncation_response(response):
                    truncation_failures += 1
                    if not compact_mode:
                        compact_mode = True
                        print("提示: 检测到 length 截断，单图评分切换到紧凑提示模式。")
                    elif not drop_reference_mode and active_refs:
                        drop_reference_mode = True
                        print("提示: 单图仍被截断，临时移除参考图后重试。")
                    elif truncation_failures >= SINGLE_TRUNCATION_FAIL_LIMIT:
                        print(
                            f"失败: {image_path.name} 连续 {truncation_failures} 次 length 截断，"
                            "已停止单图重试并返回None。"
                        )
                        return _pack(None, None)
            except KeyboardInterrupt:
                raise
            except Exception as e:
                print(f"错误: 调用API对 {image_path.name} 评分时出错: {e}")
                degraded_now = _maybe_degrade_siliconflow_by_retry_pressure(
                    attempt=attempt,
                    error_text=str(e),
                    request_tag=request_tag,
                    task_token=request_task_token,
                )
                siliconflow_task_degraded = siliconflow_task_degraded or bool(degraded_now)

            if not infinite_retry and isinstance(max_retries, int) and max_retries > 0 and attempt >= max_retries:
                print(f"失败: 已达到最大重试次数({max_retries})，仍无法获取 {image_path.name} 的分数。")
                return _pack(None, None)

            wait_seconds = _compute_retry_delay(attempt, None, response=response)
            print(f"将在 {wait_seconds:.1f} 秒后重试（第 {attempt} 次失败）...")
            time.sleep(wait_seconds)
    finally:
        if siliconflow_task_degraded:
            _disable_siliconflow_single_for_task(request_task_token)


def _safe_float(value) -> float | None:
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, (int, float, np.number)):
        return float(value)
    if isinstance(value, str):
        txt = value.strip().replace(",", "")
        if not txt or txt.startswith("="):
            return None
        try:
            return float(txt)
        except ValueError:
            return None
    return None


def _find_single_xlsx_file(folder: Path) -> Path:
    xlsx_files = sorted([p for p in folder.glob("*.xlsx") if not p.name.startswith("~$")])
    if not xlsx_files:
        raise FileNotFoundError(f"未在 {folder} 找到 xlsx 文件")
    if len(xlsx_files) > 1:
        print(f"警告: 在 {folder} 找到多个 xlsx，默认使用 {xlsx_files[0].name}")
    return xlsx_files[0]


def _extract_row_total(row_data: dict, prefix: str) -> float | None:
    total = _safe_float(row_data.get(f"{prefix}_Total"))
    if total is not None:
        return total

    s1 = _safe_float(row_data.get(f"{prefix}_S1"))
    s2 = _safe_float(row_data.get(f"{prefix}_S2"))
    s3 = _safe_float(row_data.get(f"{prefix}_S3"))
    s4 = _safe_float(row_data.get(f"{prefix}_S4"))
    p_score = _safe_float(row_data.get(f"{prefix}_P"))
    if None in (s1, s2, s3, s4, p_score):
        return None
    return float(s1 + s2 + s3 + s4 - p_score)


def _load_human_review_sheet(xlsx_path: Path) -> dict:
    if openpyxl is None:
        raise RuntimeError("未安装 openpyxl，无法读取人工评审 xlsx。请先执行: pip install openpyxl")

    wb = openpyxl.load_workbook(xlsx_path, data_only=False)
    ws = wb.active

    headers = []
    for c in range(1, ws.max_column + 1):
        val = ws.cell(row=1, column=c).value
        headers.append(str(val).strip() if val is not None else "")

    if "sample_id" not in headers or "image_name" not in headers:
        raise ValueError(f"{xlsx_path.name} 缺少必要列 sample_id 或 image_name")

    final_level_key = next((h for h in headers if h.startswith("Final_Level")), "")
    records = []
    scoring_standard = ""

    for r in range(2, ws.max_row + 1):
        row_data = {}
        for c, header in enumerate(headers, start=1):
            if not header:
                continue
            row_data[header] = ws.cell(row=r, column=c).value

        sample_id = str(row_data.get("sample_id") or "").strip()
        image_name = str(row_data.get("image_name") or "").strip()
        if not sample_id and not image_name:
            continue

        std_text = row_data.get("评分标准")
        if not scoring_standard and isinstance(std_text, str) and std_text.strip():
            scoring_standard = std_text.strip()

        r1_total = _extract_row_total(row_data, "R1")
        r2_total = _extract_row_total(row_data, "R2")
        r3_total = _extract_row_total(row_data, "R3")

        avg_total = _safe_float(row_data.get("Avg_Total"))
        totals = [t for t in [r1_total, r2_total, r3_total] if t is not None]
        if avg_total is None and totals:
            avg_total = float(np.mean(totals))
        
        # 将负分映射为0分
        if avg_total is not None and avg_total < 0:
            avg_total = max(0.0, avg_total)

        records.append(
            {
                "sample_id": sample_id or f"ROW{r}",
                "relation": str(row_data.get("relation") or "").strip(),
                "dataset": str(row_data.get("dataset") or "").strip(),
                "image_name": image_name,
                "reason": str(row_data.get("Reason") or "").strip(),
                "final_level": str(row_data.get(final_level_key) or "").strip() if final_level_key else "",
                "r1_total": r1_total,
                "r2_total": r2_total,
                "r3_total": r3_total,
                "human_avg": avg_total,
            }
        )

    if not records:
        raise ValueError(f"{xlsx_path.name} 中未读取到任何有效样本记录")

    return {
        "xlsx_path": str(xlsx_path),
        "scoring_standard": scoring_standard,
        "records": records,
    }


def _build_train_calibration_text(
    scoring_standard: str,
    train_records: list[dict],
    include_reasons: bool = False,
) -> str:
    lines = [
        "以下是人工评审校准信息，请严格对齐这套评分尺度。",
        "目标: 对时间序列图文一致性打分时，尽量贴近人工评审员给分风格。",
    ]

    if scoring_standard:
        lines.append("【人工评分标准（来自Train表）】")
        lines.append(scoring_standard)

    lines.append("【人工评审样例（Train）】")
    for i, rec in enumerate(train_records, start=1):
        r1 = "NA" if rec.get("r1_total") is None else f"{rec['r1_total']:.2f}"
        r2 = "NA" if rec.get("r2_total") is None else f"{rec['r2_total']:.2f}"
        r3 = "NA" if rec.get("r3_total") is None else f"{rec['r3_total']:.2f}"
        avg = "NA" if rec.get("human_avg") is None else f"{rec['human_avg']:.2f}"
        level = rec.get("final_level") or "Unknown"
        lines.append(
            f"样例{i}: sample_id={rec.get('sample_id')} | dataset={rec.get('dataset')} | relation={rec.get('relation')}"
        )
        lines.append(
            f"三人总分: R1={r1}, R2={r2}, R3={r3}; 人工均分={avg}; 分段={level}; image_name={rec.get('image_name')}"
        )
        reason = rec.get("reason")
        if include_reasons and reason:
            lines.append(f"人工打分理由: {reason}")

    lines.append("请对待评图片输出0-100整数分，且第一行先给分数。")
    return "\n".join(lines)


def _build_train_reference_items(train_dir: Path, train_records: list[dict]) -> list[dict]:
    items = []
    for rec in train_records:
        image_name = rec.get("image_name")
        if not image_name:
            continue
        image_path = train_dir / image_name
        if not image_path.exists():
            print(f"警告: Train参考图片不存在，已跳过: {image_path}")
            continue

        avg = rec.get("human_avg")
        avg_str = "NA" if avg is None else f"{avg:.2f}"
        label = (
            f"参考图 sample_id={rec.get('sample_id')}, dataset={rec.get('dataset')}, "
            f"human_avg={avg_str}, level={rec.get('final_level') or 'Unknown'}"
        )
        items.append(
            {
                "image_name": image_name,
                "label": label,
                "image_data_url": _read_image_as_data_url(image_path),
            }
        )
    return items


def _coerce_score_runs(value) -> list[int]:
    runs = []
    if isinstance(value, bool) or value is None:
        return runs

    if isinstance(value, (int, float, np.number)):
        score = int(round(float(value)))
        if 0 <= score <= 100:
            return [score]
        return runs

    if isinstance(value, dict):
        return _coerce_score_runs(value.get("runs"))

    if isinstance(value, list):
        for it in value:
            if isinstance(it, bool):
                continue
            if isinstance(it, (int, float, np.number)):
                sc = int(round(float(it)))
                if 0 <= sc <= 100:
                    runs.append(sc)
    return runs


def _normalize_runs_container(container) -> dict[str, list[int]]:
    out = {}
    if not isinstance(container, dict):
        return out

    for k, v in container.items():
        key = str(k).strip()
        if not key:
            continue
        runs = _coerce_score_runs(v)
        if runs:
            out[key] = runs
    return out


def _load_human_eval_cache(cache_path: Path) -> dict:
    base = {"train_runs": {}, "test_runs": {}}
    if not cache_path.exists():
        return base
    try:
        raw = json.loads(cache_path.read_text(encoding="utf-8"))
    except Exception:
        return base

    if not isinstance(raw, dict):
        return base

    # 新结构
    if "train_runs" in raw or "test_runs" in raw:
        return {
            "train_runs": _normalize_runs_container(raw.get("train_runs")),
            "test_runs": _normalize_runs_container(raw.get("test_runs")),
        }

    # 兼容旧结构（image_name -> score），默认视为 test_runs
    return {
        "train_runs": {},
        "test_runs": _normalize_runs_container(raw),
    }


def _save_human_eval_cache(cache_path: Path, cache_obj: dict):
    payload = {
        "train_runs": _normalize_runs_container(cache_obj.get("train_runs")),
        "test_runs": _normalize_runs_container(cache_obj.get("test_runs")),
    }
    cache_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def _normalize_human_for_calibration(human_value, clip_negative_to_zero: bool = True) -> float | None:
    hv = _safe_float(human_value)
    if hv is None:
        return None
    if clip_negative_to_zero and -10.0 <= hv < 0.0:
        return 0.0
    if 0.0 <= hv <= 100.0:
        return float(hv)
    return None


def _fit_ab_with_method(x: np.ndarray, y: np.ndarray, method: str) -> tuple[float, float, str]:
    m = str(method).strip().lower()
    if m == "identity":
        return 1.0, 0.0, "identity"

    if m == "bias_only":
        bias = float(np.mean(y - x))
        return 1.0, bias, "bias_only"

    if np.std(x) < 1e-6:
        bias = float(np.mean(y - x))
        return 1.0, bias, "bias_only_fallback"

    if m == "linear":
        a, _ = np.polyfit(x, y, 1)
        a = float(max(0.2, min(1.8, a)))
        b = float(np.mean(y) - a * np.mean(x))
        return a, b, "linear"

    if m == "theil_sen":
        try:
            slope, _, _, _ = stats.theilslopes(y, x)
            a = float(max(0.2, min(1.8, slope)))
            b = float(np.mean(y) - a * np.mean(x))
            return a, b, "theil_sen"
        except Exception:
            a, _ = np.polyfit(x, y, 1)
            a = float(max(0.2, min(1.8, a)))
            b = float(np.mean(y) - a * np.mean(x))
            return a, b, "linear_fallback_for_theil"

    raise ValueError(f"不支持的校准方法: {method}")


def _fit_train_linear_calibration(
    train_pairs: list[dict],
    method: str = HUMAN_EVAL_CALIBRATION_METHOD_DEFAULT,
    clip_negative_to_zero: bool = HUMAN_EVAL_CLIP_NEGATIVE_HUMAN_TO_ZERO_DEFAULT,
) -> dict:
    """
    在 Train 上拟合 human ~= a * model + b。
    method=auto 时，在 identity/bias_only/linear/theil_sen 中自动选择。
    """
    valid = []
    for p in train_pairs:
        human = _normalize_human_for_calibration(p.get("human_avg"), clip_negative_to_zero=clip_negative_to_zero)
        model = _safe_float(p.get("model_raw"))
        if human is None or model is None:
            continue
        if not (0.0 <= model <= 100.0):
            continue
        valid.append((float(model), float(human)))

    if len(valid) < 2:
        return {
            "method": "identity",
            "requested_method": method,
            "a": 1.0,
            "b": 0.0,
            "train_count": len(valid),
            "train_raw_r": None,
            "train_mae": None,
            "train_loo_mae": None,
            "clip_negative_to_zero": bool(clip_negative_to_zero),
            "candidates": [],
            "note": "有效Train样本不足，未进行校准",
        }

    x = np.array([v[0] for v in valid], dtype=float)
    y = np.array([v[1] for v in valid], dtype=float)

    corr = None
    if np.std(x) > 1e-8 and np.std(y) > 1e-8:
        corr = float(np.corrcoef(x, y)[0, 1])

    method = str(method).strip().lower()
    candidate_methods = ["identity", "bias_only"]
    if np.std(x) >= 1e-6:
        candidate_methods.extend(["linear", "theil_sen"])

    def _eval_candidate(m: str) -> dict:
        a, b, used = _fit_ab_with_method(x, y, m)
        preds = np.array([_apply_calibration(v, {"a": a, "b": b}) for v in x], dtype=float)
        train_mae = float(np.mean(np.abs(preds - y)))
        train_bias = float(np.mean(preds - y))

        loo_mae = None
        if len(x) >= 4:
            errs = []
            for i in range(len(x)):
                x_sub = np.delete(x, i)
                y_sub = np.delete(y, i)
                a_i, b_i, _ = _fit_ab_with_method(x_sub, y_sub, m)
                pred_i = float(_apply_calibration(x[i], {"a": a_i, "b": b_i}))
                errs.append(abs(pred_i - y[i]))
            loo_mae = float(np.mean(errs))

        return {
            "requested_method": m,
            "method": used,
            "a": float(a),
            "b": float(b),
            "train_mae": round(train_mae, 4),
            "train_bias": round(train_bias, 4),
            "train_loo_mae": None if loo_mae is None else round(float(loo_mae), 4),
        }

    candidates = []
    for m in candidate_methods:
        try:
            candidates.append(_eval_candidate(m))
        except Exception:
            continue

    if not candidates:
        return {
            "method": "identity",
            "requested_method": method,
            "a": 1.0,
            "b": 0.0,
            "train_count": len(valid),
            "train_raw_r": None if corr is None else round(corr, 4),
            "train_mae": None,
            "train_loo_mae": None,
            "clip_negative_to_zero": bool(clip_negative_to_zero),
            "candidates": [],
            "note": "校准候选方法均失败，回退 identity",
        }

    chosen = None
    if method == "auto":
        chosen = min(
            candidates,
            key=lambda c: (
                c["train_loo_mae"] if c["train_loo_mae"] is not None else c["train_mae"],
                abs(c["train_bias"]),
            ),
        )
        note = "auto 模式按 Train 留一误差优先选择校准方法"
    else:
        chosen = next((c for c in candidates if c["requested_method"] == method), None)
        if chosen is None:
            chosen = min(
                candidates,
                key=lambda c: c["train_loo_mae"] if c["train_loo_mae"] is not None else c["train_mae"],
            )
            note = f"指定方法 {method} 不可用，已自动回退 {chosen.get('requested_method')}"
        else:
            note = f"按指定方法 {method} 校准"

    return {
        "method": chosen["method"],
        "requested_method": method,
        "a": float(chosen["a"]),
        "b": float(chosen["b"]),
        "train_count": len(valid),
        "train_raw_r": None if corr is None else round(corr, 4),
        "train_mae": chosen.get("train_mae"),
        "train_loo_mae": chosen.get("train_loo_mae"),
        "clip_negative_to_zero": bool(clip_negative_to_zero),
        "candidates": candidates,
        "note": note,
    }


def _apply_calibration(raw_score: float | int, calibration: dict) -> int:
    y = _apply_calibration_float(raw_score, calibration)
    return int(round(y))


def _apply_calibration_float(raw_score: float | int, calibration: dict) -> float:
    a = _safe_float(calibration.get("a"))
    b = _safe_float(calibration.get("b"))
    x = _safe_float(raw_score)
    if x is None:
        raise ValueError("raw_score 不是有效数值")
    if a is None:
        a = 1.0
    if b is None:
        b = 0.0
    y = float(a) * float(x) + float(b)
    y = max(0.0, min(100.0, y))
    return float(y)


def _compute_test_metrics(result_rows: list[dict]) -> dict:
    valid_rows = [r for r in result_rows if r.get("human_avg") is not None and r.get("model_score") is not None]
    if not valid_rows:
        return {
            "count": 0,
            "mae": None,
            "rmse": None,
            "mean_bias": None,
            "within_5": None,
            "within_10": None,
            "spearman": None,
            "pearson": None,
        }

    human = np.array([r["human_avg"] for r in valid_rows], dtype=float)
    model_value = np.array([r["model_score"] for r in valid_rows], dtype=float)
    model_rank = np.array(
        [
            _safe_float(r.get("model_rank_score"))
            if _safe_float(r.get("model_rank_score")) is not None
            else _safe_float(r.get("model_score"))
            for r in valid_rows
        ],
        dtype=float,
    )
    diff = model_value - human
    abs_diff = np.abs(diff)

    mae = float(np.mean(abs_diff))
    rmse = float(np.sqrt(np.mean(np.square(diff))))
    mean_bias = float(np.mean(diff))
    within_5 = float(np.mean(abs_diff <= 5.0))
    within_10 = float(np.mean(abs_diff <= 10.0))

    spearman = 0.0
    if len(valid_rows) >= 2:
        rho, _ = stats.spearmanr(human, model_rank)
        if not np.isnan(rho):
            spearman = float(rho)

    pearson = 0.0
    if len(valid_rows) >= 2 and np.std(human) > 1e-8 and np.std(model_rank) > 1e-8:
        pearson = float(np.corrcoef(human, model_rank)[0, 1])

    return {
        "count": len(valid_rows),
        "mae": round(mae, 4),
        "rmse": round(rmse, 4),
        "mean_bias": round(mean_bias, 4),
        "within_5": round(within_5, 4),
        "within_10": round(within_10, 4),
        "spearman": round(spearman, 4),
        "pearson": round(pearson, 4),
    }


def _scan_threshold_with_precision(
    train_rows: list[dict],
    mode: str,
    human_threshold: int,
    target_precision: float,
    min_support: int,
) -> dict:
    mode = str(mode).strip().lower()
    if mode not in {"keep", "drop"}:
        raise ValueError(f"不支持的阈值扫描模式: {mode}")

    if not train_rows:
        return {
            "mode": mode,
            "found": False,
            "threshold": None,
            "support": 0,
            "correct": 0,
            "precision": None,
            "recall": None,
            "base_count": 0,
            "target_precision": round(float(target_precision), 4),
            "min_support": int(min_support),
            "top_candidates": [],
            "note": "无可用训练样本",
        }

    if mode == "keep":
        base_count = sum(1 for r in train_rows if float(r["human"]) >= float(human_threshold))
    else:
        base_count = sum(1 for r in train_rows if float(r["human"]) <= float(human_threshold))

    best = None
    scanned = []
    for threshold in range(0, 101):
        if mode == "keep":
            pred_rows = [r for r in train_rows if int(r["model"]) >= threshold]
            correct = sum(1 for r in pred_rows if float(r["human"]) >= float(human_threshold))
        else:
            pred_rows = [r for r in train_rows if int(r["model"]) <= threshold]
            correct = sum(1 for r in pred_rows if float(r["human"]) <= float(human_threshold))

        support = len(pred_rows)
        if support <= 0:
            continue

        precision = float(correct / support)
        recall = None if base_count <= 0 else float(correct / base_count)

        item = {
            "threshold": int(threshold),
            "support": int(support),
            "correct": int(correct),
            "precision": round(precision, 4),
            "recall": None if recall is None else round(recall, 4),
        }
        scanned.append(item)

        if support < min_support or precision < target_precision:
            continue

        if best is None:
            best = item
            continue

        best_recall = -1.0 if best["recall"] is None else float(best["recall"])
        cur_recall = -1.0 if recall is None else float(recall)
        if cur_recall > best_recall + 1e-12:
            best = item
            continue
        if abs(cur_recall - best_recall) > 1e-12:
            continue

        if support > int(best["support"]):
            best = item
            continue
        if support < int(best["support"]):
            continue

        if mode == "keep" and threshold < int(best["threshold"]):
            best = item
        elif mode == "drop" and threshold > int(best["threshold"]):
            best = item

    top_candidates = sorted(
        scanned,
        key=lambda x: (float(x["precision"]), int(x["support"]), -1.0 if x["recall"] is None else float(x["recall"])),
        reverse=True,
    )[:5]

    if best is None:
        return {
            "mode": mode,
            "found": False,
            "threshold": None,
            "support": 0,
            "correct": 0,
            "precision": None,
            "recall": None,
            "base_count": int(base_count),
            "target_precision": round(float(target_precision), 4),
            "min_support": int(min_support),
            "top_candidates": top_candidates,
            "note": "未找到满足精度目标的阈值",
        }

    return {
        "mode": mode,
        "found": True,
        "threshold": int(best["threshold"]),
        "support": int(best["support"]),
        "correct": int(best["correct"]),
        "precision": round(float(best["precision"]), 4),
        "recall": None if best["recall"] is None else round(float(best["recall"]), 4),
        "base_count": int(base_count),
        "target_precision": round(float(target_precision), 4),
        "min_support": int(min_support),
        "top_candidates": top_candidates,
        "note": "已找到满足目标精度的阈值",
    }


def _build_auto_cleaning_threshold_policy(
    train_pairs: list[dict],
    calibration: dict,
    use_posthoc_calibration: bool,
    use_auto_threshold: bool,
    default_keep_threshold: int,
    default_drop_threshold: int,
    target_keep_human_threshold: int,
    target_drop_human_threshold: int,
    target_precision: float,
    min_support: int,
    clip_negative_human_to_zero: bool,
) -> dict:
    default_keep_threshold = int(max(0, min(100, default_keep_threshold)))
    default_drop_threshold = int(max(0, min(100, default_drop_threshold)))
    if default_keep_threshold <= default_drop_threshold:
        default_keep_threshold = min(100, default_drop_threshold + 1)

    train_rows = []
    for p in train_pairs:
        raw_score = _safe_float(p.get("model_raw"))
        human = _normalize_human_for_calibration(
            p.get("human_avg"),
            clip_negative_to_zero=bool(clip_negative_human_to_zero),
        )
        if raw_score is None or human is None:
            continue

        model_score = int(round(raw_score))
        if use_posthoc_calibration:
            model_score = _apply_calibration(raw_score, calibration)

        train_rows.append(
            {
                "sample_id": p.get("sample_id"),
                "image_name": p.get("image_name"),
                "model": int(max(0, min(100, model_score))),
                "human": float(human),
            }
        )

    policy = {
        "source": "manual",
        "use_auto_threshold": bool(use_auto_threshold),
        "default_keep_threshold": int(default_keep_threshold),
        "default_drop_threshold": int(default_drop_threshold),
        "keep_threshold": int(default_keep_threshold),
        "drop_threshold": int(default_drop_threshold),
        "target_keep_human_threshold": int(target_keep_human_threshold),
        "target_drop_human_threshold": int(target_drop_human_threshold),
        "target_precision": round(float(target_precision), 4),
        "min_support": int(min_support),
        "train_rows_used": int(len(train_rows)),
        "keep_scan": None,
        "drop_scan": None,
        "auto_keep_applied": False,
        "auto_drop_applied": False,
        "note": "",
    }

    if not use_auto_threshold:
        policy["note"] = "已禁用自动阈值拟合，使用手动阈值"
        return policy

    if len(train_rows) < max(2, int(min_support)):
        policy["note"] = "训练样本不足，自动阈值拟合未启用"
        return policy

    keep_scan = _scan_threshold_with_precision(
        train_rows,
        mode="keep",
        human_threshold=int(target_keep_human_threshold),
        target_precision=float(target_precision),
        min_support=int(min_support),
    )
    drop_scan = _scan_threshold_with_precision(
        train_rows,
        mode="drop",
        human_threshold=int(target_drop_human_threshold),
        target_precision=float(target_precision),
        min_support=int(min_support),
    )

    policy["keep_scan"] = keep_scan
    policy["drop_scan"] = drop_scan

    if keep_scan.get("found"):
        policy["keep_threshold"] = int(keep_scan["threshold"])
        policy["auto_keep_applied"] = True
    if drop_scan.get("found"):
        policy["drop_threshold"] = int(drop_scan["threshold"])
        policy["auto_drop_applied"] = True

    if policy["keep_threshold"] <= policy["drop_threshold"]:
        policy["keep_threshold"] = int(default_keep_threshold)
        policy["drop_threshold"] = int(default_drop_threshold)
        policy["source"] = "manual"
        policy["auto_keep_applied"] = False
        policy["auto_drop_applied"] = False
        policy["note"] = "自动阈值发生冲突，已回退到手动阈值"
        return policy

    if policy["auto_keep_applied"] and policy["auto_drop_applied"]:
        policy["source"] = "auto"
        policy["note"] = "keep/drop 阈值均由 Train 自动拟合"
    elif policy["auto_keep_applied"] or policy["auto_drop_applied"]:
        policy["source"] = "auto_partial"
        policy["note"] = "部分阈值由 Train 自动拟合，部分使用手动默认值"
    else:
        policy["source"] = "manual"
        policy["note"] = "未找到满足目标精度的自动阈值，沿用手动阈值"

    return policy


def _decide_cleaning_action(
    score: int,
    raw_std: float,
    raw_range: int,
    keep_threshold: int,
    drop_threshold: int,
    stability_std_threshold: float,
    stability_range_threshold: int,
    uncertain_action: str,
) -> dict:
    score_val = int(max(0, min(100, int(score))))
    std_val = max(0.0, float(raw_std))
    range_val = max(0, int(raw_range))

    std_stable = True if stability_std_threshold <= 0 else (std_val <= float(stability_std_threshold))
    range_stable = True if stability_range_threshold <= 0 else (range_val <= int(stability_range_threshold))
    stable = bool(std_stable and range_stable)

    if score_val >= int(keep_threshold):
        score_band = "high"
    elif score_val <= int(drop_threshold):
        score_band = "low"
    else:
        score_band = "gray"

    uncertain_before_policy = False
    if score_band == "high" and stable:
        decision = "keep"
        reason = "high_score_and_stable"
    elif score_band == "low" and stable:
        decision = "drop"
        reason = "low_score_and_stable"
    else:
        uncertain_before_policy = True
        action = str(uncertain_action).strip().lower()
        if action == "keep":
            decision = "keep"
        elif action == "midpoint":
            mid = (float(keep_threshold) + float(drop_threshold)) / 2.0
            decision = "keep" if float(score_val) >= mid else "drop"
        else:
            action = "drop"
            decision = "drop"

        reason_parts = []
        if score_band == "gray":
            reason_parts.append("score_in_gray_zone")
        if not stable:
            reason_parts.append("unstable")
        if not reason_parts:
            reason_parts.append("policy_fallback")
        reason = "|".join(reason_parts) + f"|uncertain_action={action}"

    std_ratio = 0.0
    if stability_std_threshold > 0:
        std_ratio = min(1.0, std_val / float(stability_std_threshold))

    range_ratio = 0.0
    if stability_range_threshold > 0:
        range_ratio = min(1.0, float(range_val) / float(stability_range_threshold))

    stability_conf = 1.0 - 0.5 * std_ratio - 0.5 * range_ratio
    stability_conf = max(0.0, min(1.0, stability_conf))

    gap = max(1.0, float(keep_threshold - drop_threshold))
    if score_val >= keep_threshold:
        margin = float(score_val - keep_threshold)
    elif score_val <= drop_threshold:
        margin = float(drop_threshold - score_val)
    else:
        margin = float(min(score_val - drop_threshold, keep_threshold - score_val))
    margin_conf = min(1.0, max(0.0, margin / max(1.0, gap / 2.0)))

    confidence = 0.65 * stability_conf + 0.35 * margin_conf
    if uncertain_before_policy:
        confidence *= 0.85
    confidence = max(0.0, min(1.0, confidence))

    if confidence >= 0.8:
        confidence_level = "high"
    elif confidence >= 0.6:
        confidence_level = "medium"
    else:
        confidence_level = "low"

    return {
        "decision": decision,
        "decision_reason": reason,
        "score_band": score_band,
        "stable": stable,
        "uncertain_before_policy": uncertain_before_policy,
        "confidence": round(float(confidence), 4),
        "confidence_level": confidence_level,
    }


def _summarize_auto_cleaning(
    result_rows: list[dict],
    keep_human_threshold: int,
    drop_human_threshold: int,
) -> dict:
    total = len(result_rows)
    keep_rows = [r for r in result_rows if str(r.get("clean_decision")) == "keep"]
    drop_rows = [r for r in result_rows if str(r.get("clean_decision")) == "drop"]

    uncertain_before_policy = sum(1 for r in result_rows if bool(r.get("clean_uncertain_before_policy")))
    unstable_count = sum(1 for r in result_rows if not bool(r.get("clean_stable")))

    confidences = [
        float(r.get("clean_confidence"))
        for r in result_rows
        if isinstance(r.get("clean_confidence"), (int, float, np.number))
    ]

    decision_reason_counts = {}
    for r in result_rows:
        reason = str(r.get("clean_decision_reason") or "")
        if not reason:
            continue
        decision_reason_counts[reason] = decision_reason_counts.get(reason, 0) + 1

    summary = {
        "total": int(total),
        "keep_count": int(len(keep_rows)),
        "drop_count": int(len(drop_rows)),
        "keep_ratio": None if total <= 0 else round(float(len(keep_rows) / total), 4),
        "drop_ratio": None if total <= 0 else round(float(len(drop_rows) / total), 4),
        "uncertain_before_policy": int(uncertain_before_policy),
        "unstable_count": int(unstable_count),
        "mean_confidence": None if not confidences else round(float(np.mean(confidences)), 4),
        "high_confidence_count": int(sum(1 for c in confidences if c >= 0.8)),
        "low_confidence_count": int(sum(1 for c in confidences if c < 0.6)),
        "keep_human_threshold": int(keep_human_threshold),
        "drop_human_threshold": int(drop_human_threshold),
        "keep_precision_vs_human": None,
        "drop_precision_vs_human": None,
        "keep_recall_vs_human": None,
        "drop_recall_vs_human": None,
        "keep_support_with_human": 0,
        "drop_support_with_human": 0,
        "human_high_count": 0,
        "human_low_count": 0,
        "decision_reason_counts": dict(
            sorted(decision_reason_counts.items(), key=lambda kv: kv[1], reverse=True)
        ),
    }

    rows_with_human = [r for r in result_rows if r.get("human_avg") is not None]
    if not rows_with_human:
        return summary

    human_high = [r for r in rows_with_human if float(r.get("human_avg")) >= float(keep_human_threshold)]
    human_low = [r for r in rows_with_human if float(r.get("human_avg")) <= float(drop_human_threshold)]
    keep_with_human = [r for r in keep_rows if r.get("human_avg") is not None]
    drop_with_human = [r for r in drop_rows if r.get("human_avg") is not None]

    keep_correct = sum(1 for r in keep_with_human if float(r.get("human_avg")) >= float(keep_human_threshold))
    drop_correct = sum(1 for r in drop_with_human if float(r.get("human_avg")) <= float(drop_human_threshold))

    summary["keep_support_with_human"] = int(len(keep_with_human))
    summary["drop_support_with_human"] = int(len(drop_with_human))
    summary["human_high_count"] = int(len(human_high))
    summary["human_low_count"] = int(len(human_low))

    if keep_with_human:
        summary["keep_precision_vs_human"] = round(float(keep_correct / len(keep_with_human)), 4)
    if drop_with_human:
        summary["drop_precision_vs_human"] = round(float(drop_correct / len(drop_with_human)), 4)
    if human_high:
        summary["keep_recall_vs_human"] = round(float(keep_correct / len(human_high)), 4)
    if human_low:
        summary["drop_recall_vs_human"] = round(float(drop_correct / len(human_low)), 4)

    return summary


def _pick_stability_target(test_records: list[dict], target: str | None) -> dict:
    if target:
        tgt = target.strip()
        for rec in test_records:
            if rec.get("sample_id") == tgt or rec.get("image_name") == tgt:
                return rec
        raise ValueError(f"未在 Test 样本中找到 stability_target={target}")
    return test_records[0]


def run_human_eval(
    human_eval_root: Path,
    stability_target: str | None = None,
    stability_runs: int = HUMAN_EVAL_STABILITY_RUNS_DEFAULT,
    stability_std_threshold: float = HUMAN_EVAL_STABILITY_STD_THRESHOLD_DEFAULT,
    stability_range_threshold: int = HUMAN_EVAL_STABILITY_RANGE_THRESHOLD_DEFAULT,
    stability_max_extra_runs: int = HUMAN_EVAL_STABILITY_MAX_EXTRA_RUNS_DEFAULT,
    force_rerun: bool = False,
    test_runs_per_image: int = HUMAN_EVAL_TEST_RUNS_PER_IMAGE_DEFAULT,
    score_aggregation: str = HUMAN_EVAL_SCORE_AGGREGATION_DEFAULT,
    batch_query_size: int = HUMAN_EVAL_BATCH_QUERY_SIZE_DEFAULT,
    calibration_mode: str = HUMAN_EVAL_CALIBRATION_MODE_DEFAULT,
    calibration_method: str = HUMAN_EVAL_CALIBRATION_METHOD_DEFAULT,
    clip_negative_human_to_zero: bool = HUMAN_EVAL_CLIP_NEGATIVE_HUMAN_TO_ZERO_DEFAULT,
    clean_keep_threshold: int = HUMAN_EVAL_CLEAN_KEEP_THRESHOLD_DEFAULT,
    clean_drop_threshold: int = HUMAN_EVAL_CLEAN_DROP_THRESHOLD_DEFAULT,
    clean_uncertain_action: str = HUMAN_EVAL_CLEAN_UNCERTAIN_ACTION_DEFAULT,
    clean_use_auto_threshold: bool = HUMAN_EVAL_CLEAN_USE_AUTO_THRESHOLD_DEFAULT,
    clean_target_keep_human_threshold: int = HUMAN_EVAL_CLEAN_TARGET_KEEP_HUMAN_THRESHOLD_DEFAULT,
    clean_target_drop_human_threshold: int = HUMAN_EVAL_CLEAN_TARGET_DROP_HUMAN_THRESHOLD_DEFAULT,
    clean_target_precision: float = HUMAN_EVAL_CLEAN_TARGET_PRECISION_DEFAULT,
    clean_min_support: int = HUMAN_EVAL_CLEAN_MIN_SUPPORT_DEFAULT,
    min_pearson_threshold: float = HUMAN_EVAL_MIN_PEARSON_THRESHOLD_DEFAULT,
    strict_pearson_check: bool = HUMAN_EVAL_STRICT_PEARSON_CHECK_DEFAULT,
):
    train_dir = human_eval_root / HUMAN_EVAL_TRAIN_DIR_NAME
    test_dir = human_eval_root / HUMAN_EVAL_TEST_DIR_NAME
    output_path = human_eval_root / HUMAN_EVAL_OUTPUT_NAME
    cache_path = human_eval_root / HUMAN_EVAL_CACHE_NAME
    test_runs_per_image = max(1, int(test_runs_per_image))
    stability_runs = max(1, int(stability_runs))
    stability_std_threshold = max(0.0, float(stability_std_threshold))
    stability_range_threshold = max(0, int(stability_range_threshold))
    stability_max_extra_runs = max(0, int(stability_max_extra_runs))
    batch_size = max(1, min(HUMAN_EVAL_MAX_IMAGES_PER_REQUEST, int(batch_query_size)))
    score_aggregation = str(score_aggregation).strip().lower()
    clean_keep_threshold = int(max(0, min(100, int(clean_keep_threshold))))
    clean_drop_threshold = int(max(0, min(100, int(clean_drop_threshold))))
    clean_target_keep_human_threshold = int(max(0, min(100, int(clean_target_keep_human_threshold))))
    clean_target_drop_human_threshold = int(max(0, min(100, int(clean_target_drop_human_threshold))))
    clean_target_precision = float(max(0.0, min(1.0, float(clean_target_precision))))
    clean_min_support = max(1, int(clean_min_support))
    clean_use_auto_threshold = bool(clean_use_auto_threshold)

    clean_uncertain_action = str(clean_uncertain_action).strip().lower()
    if clean_uncertain_action not in {"drop", "keep", "midpoint"}:
        raise ValueError(f"不支持的 clean_uncertain_action: {clean_uncertain_action}")

    if clean_keep_threshold <= clean_drop_threshold:
        raise ValueError(
            "清洗阈值非法: clean_keep_threshold 必须大于 clean_drop_threshold，"
            f"当前 keep={clean_keep_threshold}, drop={clean_drop_threshold}"
        )

    if score_aggregation not in {"mean", "median", "last"}:
        raise ValueError(f"不支持的 score_aggregation: {score_aggregation}")

    if not train_dir.is_dir():
        raise FileNotFoundError(f"Train目录不存在: {train_dir}")
    if not test_dir.is_dir():
        raise FileNotFoundError(f"Test目录不存在: {test_dir}")

    train_xlsx = _find_single_xlsx_file(train_dir)
    test_xlsx = _find_single_xlsx_file(test_dir)
    print(f"Train评审表: {train_xlsx}")
    print(f"Test评审表: {test_xlsx}")

    train_sheet = _load_human_review_sheet(train_xlsx)
    test_sheet = _load_human_review_sheet(test_xlsx)
    train_records = train_sheet["records"]
    test_records = test_sheet["records"]

    calibration_mode = str(calibration_mode).strip().lower()
    if calibration_mode not in {"hybrid", "posthoc_only", "prompt_only"}:
        raise ValueError(f"不支持的 calibration_mode: {calibration_mode}")

    calibration_method = str(calibration_method).strip().lower()
    use_prompt_calibration = calibration_mode in {"hybrid", "prompt_only"}
    use_posthoc_calibration = calibration_mode in {"hybrid", "posthoc_only"}

    calibration_text = ""
    reference_items = []
    if use_prompt_calibration:
        calibration_text = _build_train_calibration_text(train_sheet.get("scoring_standard", ""), train_records)
        reference_items = _build_train_reference_items(train_dir, train_records)
        if not reference_items:
            raise RuntimeError("Train 参考图为空，无法执行提示词校准评测")

    print(
        "校准策略: "
        f"mode={calibration_mode} | "
        f"prompt={use_prompt_calibration} | "
        f"posthoc={use_posthoc_calibration} | "
        f"method={calibration_method} | "
        f"score_aggregation={score_aggregation} | "
        f"clip_negative_human_to_zero={bool(clip_negative_human_to_zero)}"
    )

    print(
        "稳定性策略: "
        f"初始复评={stability_runs} 次 | "
        f"阈值 std<={stability_std_threshold}, range<={stability_range_threshold} | "
        f"最多追加 {stability_max_extra_runs} 次"
    )

    print(
        "自动清洗策略: "
        f"manual_keep>={clean_keep_threshold}, manual_drop<={clean_drop_threshold} | "
        f"auto_threshold={clean_use_auto_threshold} | "
        f"target_precision={clean_target_precision:.2f}, min_support={clean_min_support} | "
        f"uncertain_action={clean_uncertain_action}"
    )
    print(
        "评审模型有效性检查: "
        f"min_pearson={min_pearson_threshold:.2f} | "
        f"strict_check={'是(不达标即中止)' if strict_pearson_check else '否(仅警告)'}"
    )

    print(f"校准样本: 文本样本 {len(train_records)} 条，参考图 {len(reference_items)} 张")
    print(f"Test待评样本: {len(test_records)} 张 | 每图重复评测 {test_runs_per_image} 次 | 每请求最多待评图 {batch_size} 张")

    cache_obj = {"train_runs": {}, "test_runs": {}}
    if HUMAN_EVAL_USE_CACHE:
        cache_obj = _load_human_eval_cache(cache_path)
    train_runs_map = {} if force_rerun else cache_obj.get("train_runs", {})
    test_runs_map = {} if force_rerun else cache_obj.get("test_runs", {})

    if use_posthoc_calibration:
        # 先在 Train 上获取模型原始分，用于拟合校准映射。
        # posthoc_only 模式默认不注入 Train 样例，避免泄漏并提升速度。
        train_paths_missing = []
        for rec in train_records:
            image_name = rec.get("image_name")
            image_path = train_dir / image_name
            if not image_name or not image_path.exists():
                continue
            runs = _coerce_score_runs(train_runs_map.get(image_name))
            if not runs:
                train_paths_missing.append(image_path)
            else:
                train_runs_map[image_name] = [runs[0]]

        if train_paths_missing:
            print(f"Train原始分补全: 需新增 {len(train_paths_missing)} 张")
            total_train_chunks = (len(train_paths_missing) + batch_size - 1) // batch_size
            for chunk_idx, i in enumerate(range(0, len(train_paths_missing), batch_size), start=1):
                chunk = train_paths_missing[i:i + batch_size]
                print(f"Train原始分补全批次 {chunk_idx}/{total_train_chunks}: 待评 {len(chunk)} 张")

                train_extra_prompt = None
                train_reference_items = None
                if calibration_mode == "hybrid":
                    excluded = {p.name for p in chunk}
                    train_reference_items = [it for it in reference_items if it.get("image_name") not in excluded]
                    train_extra_prompt = calibration_text

                t0 = time.time()
                batch_scores = get_batch_image_scores(
                    chunk,
                    extra_prompt_text=train_extra_prompt,
                    reference_items=train_reference_items,
                    retries=HUMAN_EVAL_BATCH_MAX_RETRIES,
                )
                t_cost = time.time() - t0
                print(
                    f"Train原始分补全批次 {chunk_idx}/{total_train_chunks}: "
                    f"批量返回 {len(batch_scores)}/{len(chunk)} 张, 用时 {t_cost:.1f} 秒"
                )

                for img_path in chunk:
                    score = batch_scores.get(img_path.name)
                    if score is None:
                        print(f"Train补全单图兜底: {img_path.name}")
                        single_ref_items = None
                        if calibration_mode == "hybrid":
                            single_ref_items = [it for it in reference_items if it.get("image_name") != img_path.name]
                        score = get_image_score(
                            img_path,
                            retries=HUMAN_EVAL_SINGLE_MAX_RETRIES,
                            extra_prompt_text=train_extra_prompt,
                            reference_items=single_ref_items,
                        )
                    if score is None:
                        print(f"警告: Train原始分获取失败: {img_path.name}")
                        continue
                    train_runs_map[img_path.name] = [int(score)]

                if HUMAN_EVAL_USE_CACHE:
                    cache_obj["train_runs"] = train_runs_map
                    cache_obj["test_runs"] = test_runs_map
                    _save_human_eval_cache(cache_path, cache_obj)

    train_pairs = []
    for rec in train_records:
        image_name = rec.get("image_name")
        runs = _coerce_score_runs(train_runs_map.get(image_name))
        raw = runs[0] if runs else None
        train_pairs.append(
            {
                "sample_id": rec.get("sample_id"),
                "image_name": image_name,
                "human_avg": rec.get("human_avg"),
                "model_raw": raw,
                "final_level": rec.get("final_level"),
            }
        )

    if use_posthoc_calibration:
        calibration = _fit_train_linear_calibration(
            train_pairs,
            method=calibration_method,
            clip_negative_to_zero=bool(clip_negative_human_to_zero),
        )
    else:
        calibration = {
            "method": "identity",
            "requested_method": "identity",
            "a": 1.0,
            "b": 0.0,
            "train_count": 0,
            "train_raw_r": None,
            "train_mae": None,
            "train_loo_mae": None,
            "clip_negative_to_zero": bool(clip_negative_human_to_zero),
            "candidates": [],
            "note": "prompt_only 模式不做后处理校准，直接使用模型分",
        }

    print(
        "Train校准映射: "
        f"method={calibration.get('method')} (req={calibration.get('requested_method')}), "
        f"a={calibration.get('a'):.4f}, b={calibration.get('b'):.4f}, "
        f"train_count={calibration.get('train_count')}, "
        f"train_pearson={calibration.get('train_raw_r')}, "
        f"train_mae={calibration.get('train_mae')}, "
        f"train_loo_mae={calibration.get('train_loo_mae')}"
    )

    clean_policy = _build_auto_cleaning_threshold_policy(
        train_pairs=train_pairs,
        calibration=calibration,
        use_posthoc_calibration=use_posthoc_calibration,
        use_auto_threshold=clean_use_auto_threshold,
        default_keep_threshold=clean_keep_threshold,
        default_drop_threshold=clean_drop_threshold,
        target_keep_human_threshold=clean_target_keep_human_threshold,
        target_drop_human_threshold=clean_target_drop_human_threshold,
        target_precision=clean_target_precision,
        min_support=clean_min_support,
        clip_negative_human_to_zero=bool(clip_negative_human_to_zero),
    )
    clean_keep_threshold_used = int(clean_policy.get("keep_threshold", clean_keep_threshold))
    clean_drop_threshold_used = int(clean_policy.get("drop_threshold", clean_drop_threshold))

    print(
        "自动清洗阈值: "
        f"keep>={clean_keep_threshold_used}, "
        f"drop<={clean_drop_threshold_used}, "
        f"source={clean_policy.get('source')}, "
        f"note={clean_policy.get('note')}"
    )

    # 在 Test 上进行多次评测；每轮按 batch_size 张图合并请求，提高吞吐。
    valid_test_records = []
    for rec in test_records:
        image_name = rec.get("image_name")
        image_path = test_dir / image_name
        if not image_name or not image_path.exists():
            print(f"警告: Test图片不存在，跳过: {image_name}")
            continue
        valid_test_records.append(rec)

    test_extra_prompt = calibration_text if use_prompt_calibration else None
    test_reference_items = reference_items if use_prompt_calibration else None

    for run_idx in range(test_runs_per_image):
        need_paths = []
        for rec in valid_test_records:
            image_name = rec.get("image_name")
            runs = _coerce_score_runs(test_runs_map.get(image_name))
            if len(runs) <= run_idx:
                need_paths.append(test_dir / image_name)

        if not need_paths:
            continue

        print(f"Test第 {run_idx + 1}/{test_runs_per_image} 轮: 需评测 {len(need_paths)} 张")
        for i in tqdm(range(0, len(need_paths), batch_size), desc=f"Test轮次{run_idx + 1}批量评分"):
            chunk = need_paths[i:i + batch_size]
            batch_scores = get_batch_image_scores(
                chunk,
                extra_prompt_text=test_extra_prompt,
                reference_items=test_reference_items,
                retries=HUMAN_EVAL_BATCH_MAX_RETRIES,
            )

            for img_path in chunk:
                score = batch_scores.get(img_path.name)
                if score is None:
                    score = get_image_score(
                        img_path,
                        retries=HUMAN_EVAL_SINGLE_MAX_RETRIES,
                        extra_prompt_text=test_extra_prompt,
                        reference_items=test_reference_items,
                    )
                if score is None:
                    print(f"警告: Test评分失败: {img_path.name}")
                    continue

                runs = _coerce_score_runs(test_runs_map.get(img_path.name))
                runs.append(int(score))
                test_runs_map[img_path.name] = runs

            if HUMAN_EVAL_USE_CACHE:
                cache_obj["train_runs"] = train_runs_map
                cache_obj["test_runs"] = test_runs_map
                _save_human_eval_cache(cache_path, cache_obj)

    # 若批量轮次后仍有样本未达到目标次数，进行单图补评。
    for rec in valid_test_records:
        image_name = rec.get("image_name")
        if not image_name:
            continue

        image_path = test_dir / image_name
        if not image_path.exists():
            continue

        runs = _coerce_score_runs(test_runs_map.get(image_name))
        missing = test_runs_per_image - len(runs)
        if missing <= 0:
            continue

        print(f"Test自动补评: {image_name} 尚缺 {missing} 次")
        for _ in range(missing):
            s = get_image_score(
                image_path,
                retries=HUMAN_EVAL_SINGLE_MAX_RETRIES,
                extra_prompt_text=test_extra_prompt,
                reference_items=test_reference_items,
            )
            if s is None:
                print(f"警告: Test自动补评失败: {image_name}")
                break
            runs.append(int(s))

        test_runs_map[image_name] = runs

    if HUMAN_EVAL_USE_CACHE:
        cache_obj["train_runs"] = train_runs_map
        cache_obj["test_runs"] = test_runs_map
        _save_human_eval_cache(cache_path, cache_obj)

    result_rows = []
    for rec in valid_test_records:
        image_name = rec.get("image_name")
        runs = _coerce_score_runs(test_runs_map.get(image_name))
        if not runs:
            continue

        used_runs = runs[-test_runs_per_image:]
        arr = np.array(used_runs, dtype=float)
        raw_median = int(round(float(np.median(arr))))
        raw_mean = float(np.mean(arr))
        raw_last = int(round(float(arr[-1])))
        raw_std = float(np.std(arr))
        raw_range = int(np.max(arr) - np.min(arr))

        if score_aggregation == "mean":
            raw_aggregate = float(raw_mean)
        elif score_aggregation == "last":
            raw_aggregate = float(raw_last)
        else:
            raw_aggregate = float(raw_median)

        calibrated_rank_score = (
            _apply_calibration_float(raw_aggregate, calibration) if use_posthoc_calibration else float(raw_aggregate)
        )
        calibrated_score = int(round(calibrated_rank_score))

        clean_decision = _decide_cleaning_action(
            score=int(calibrated_score),
            raw_std=raw_std,
            raw_range=raw_range,
            keep_threshold=clean_keep_threshold_used,
            drop_threshold=clean_drop_threshold_used,
            stability_std_threshold=stability_std_threshold,
            stability_range_threshold=stability_range_threshold,
            uncertain_action=clean_uncertain_action,
        )

        human_avg = rec.get("human_avg")
        delta_cal = None if human_avg is None else round(float(calibrated_score) - float(human_avg), 2)
        abs_delta_cal = None if delta_cal is None else round(abs(delta_cal), 2)
        delta_raw = None if human_avg is None else round(float(raw_aggregate) - float(human_avg), 2)
        abs_delta_raw = None if delta_raw is None else round(abs(delta_raw), 2)

        result_rows.append(
            {
                "sample_id": rec.get("sample_id"),
                "relation": rec.get("relation"),
                "dataset": rec.get("dataset"),
                "image_name": image_name,
                "human_avg": None if human_avg is None else round(float(human_avg), 2),
                "model_score": int(calibrated_score),
                "model_rank_score": round(float(calibrated_rank_score), 4),
                "model_raw_score_median": int(raw_median),
                "model_raw_score_mean": round(raw_mean, 4),
                "model_raw_score_last": int(raw_last),
                "model_raw_score_aggregate": round(float(raw_aggregate), 4),
                "model_raw_aggregate_mode": score_aggregation,
                "model_raw_rank_score": round(float(raw_aggregate), 4),
                "model_raw_score_std": round(raw_std, 4),
                "model_raw_score_range": int(raw_range),
                "model_raw_scores": used_runs,
                "model_raw_runs_used": len(used_runs),
                "model_raw_runs_total_cached": len(runs),
                "delta_model_minus_human": delta_cal,
                "abs_delta": abs_delta_cal,
                "delta_raw_minus_human": delta_raw,
                "abs_delta_raw": abs_delta_raw,
                "clean_decision": clean_decision.get("decision"),
                "clean_decision_reason": clean_decision.get("decision_reason"),
                "clean_score_band": clean_decision.get("score_band"),
                "clean_stable": clean_decision.get("stable"),
                "clean_uncertain_before_policy": clean_decision.get("uncertain_before_policy"),
                "clean_confidence": clean_decision.get("confidence"),
                "clean_confidence_level": clean_decision.get("confidence_level"),
                "final_level": rec.get("final_level"),
                "human_reason": rec.get("reason"),
            }
        )

    metrics_calibrated = _compute_test_metrics(result_rows)
    raw_rows_for_metric = []
    for r in result_rows:
        rr = dict(r)
        rr["model_score"] = rr.get("model_raw_score_aggregate", rr.get("model_raw_score_median"))
        rr["model_rank_score"] = rr.get("model_raw_rank_score", rr.get("model_score"))
        raw_rows_for_metric.append(rr)
    metrics_raw = _compute_test_metrics(raw_rows_for_metric)
    clean_summary = _summarize_auto_cleaning(
        result_rows,
        keep_human_threshold=clean_target_keep_human_threshold,
        drop_human_threshold=clean_target_drop_human_threshold,
    )

    target_record = _pick_stability_target(test_records, stability_target)
    target_image_path = test_dir / target_record.get("image_name")
    if not target_image_path.exists():
        raise FileNotFoundError(f"稳定性测试目标图片不存在: {target_image_path}")

    initial_stability_runs = stability_runs
    max_stability_runs = initial_stability_runs + stability_max_extra_runs
    stability_scores_raw = []
    stability_scores_calibrated = []
    for i in range(initial_stability_runs):
        print(f"稳定性测试: 第 {i + 1}/{initial_stability_runs} 次评审 -> {target_record.get('image_name')}")
        s = get_image_score(
            target_image_path,
            retries=HUMAN_EVAL_SINGLE_MAX_RETRIES,
            extra_prompt_text=test_extra_prompt,
            reference_items=test_reference_items,
        )
        if s is not None:
            s_raw = int(s)
            s_cal = _apply_calibration(s_raw, calibration) if use_posthoc_calibration else s_raw
            stability_scores_raw.append(s_raw)
            stability_scores_calibrated.append(s_cal)

    auto_guard_triggered = False
    extra_runs_used = 0
    while stability_scores_calibrated and len(stability_scores_calibrated) < max_stability_runs:
        cur_arr = np.array(stability_scores_calibrated, dtype=float)
        cur_std = float(np.std(cur_arr))
        cur_range = int(np.max(cur_arr) - np.min(cur_arr))

        if cur_std <= stability_std_threshold and cur_range <= stability_range_threshold:
            break

        auto_guard_triggered = True
        extra_runs_used += 1
        run_no = initial_stability_runs + extra_runs_used
        print(
            "稳定性保护触发: "
            f"当前 std={cur_std:.4f}, range={cur_range} 超过阈值，"
            f"追加第 {run_no}/{max_stability_runs} 次评审 -> {target_record.get('image_name')}"
        )

        s = get_image_score(
            target_image_path,
            retries=HUMAN_EVAL_SINGLE_MAX_RETRIES,
            extra_prompt_text=test_extra_prompt,
            reference_items=test_reference_items,
        )
        if s is None:
            print("警告: 稳定性追加评测失败，提前结束追加。")
            break

        s_raw = int(s)
        s_cal = _apply_calibration(s_raw, calibration) if use_posthoc_calibration else s_raw
        stability_scores_raw.append(s_raw)
        stability_scores_calibrated.append(s_cal)

    stability_report = {
        "target_sample_id": target_record.get("sample_id"),
        "target_image_name": target_record.get("image_name"),
        "runs_requested": initial_stability_runs,
        "runs_total": len(stability_scores_calibrated),
        "std_threshold": stability_std_threshold,
        "range_threshold": stability_range_threshold,
        "max_extra_runs": stability_max_extra_runs,
        "extra_runs_used": extra_runs_used,
        "auto_guard_triggered": auto_guard_triggered,
        "scores": stability_scores_calibrated,
        "raw_scores": stability_scores_raw,
        "calibrated_scores": stability_scores_calibrated,
        "passed_threshold": None,
    }
    if stability_scores_calibrated:
        arr = np.array(stability_scores_calibrated, dtype=float)
        std_val = float(np.std(arr))
        range_val = int(np.max(arr) - np.min(arr))
        passed = (std_val <= stability_std_threshold) and (range_val <= stability_range_threshold)
        stability_report.update(
            {
                "mean": round(float(np.mean(arr)), 4),
                "std": round(std_val, 4),
                "min": int(np.min(arr)),
                "max": int(np.max(arr)),
                "range": range_val,
                "passed_threshold": bool(passed),
            }
        )
    if stability_scores_raw:
        arr_raw = np.array(stability_scores_raw, dtype=float)
        stability_report.update(
            {
                "raw_mean": round(float(np.mean(arr_raw)), 4),
                "raw_std": round(float(np.std(arr_raw)), 4),
                "raw_min": int(np.min(arr_raw)),
                "raw_max": int(np.max(arr_raw)),
                "raw_range": int(np.max(arr_raw) - np.min(arr_raw)),
            }
        )

    report = {
        "mode": "human_eval",
        "timestamp": time.strftime("%Y-%m-%d %H:%M:%S"),
        "human_eval_root": str(human_eval_root),
        "eval_config": {
            "test_runs_per_image": test_runs_per_image,
            "score_aggregation": score_aggregation,
            "batch_query_size": batch_size,
            "max_images_per_request": HUMAN_EVAL_MAX_IMAGES_PER_REQUEST,
            "force_rerun": bool(force_rerun),
            "stability_runs": stability_runs,
            "stability_std_threshold": stability_std_threshold,
            "stability_range_threshold": stability_range_threshold,
            "stability_max_extra_runs": stability_max_extra_runs,
            "calibration_mode": calibration_mode,
            "calibration_method": calibration_method,
            "use_prompt_calibration": use_prompt_calibration,
            "use_posthoc_calibration": use_posthoc_calibration,
            "clip_negative_human_to_zero": bool(clip_negative_human_to_zero),
            "clean_keep_threshold": clean_keep_threshold,
            "clean_drop_threshold": clean_drop_threshold,
            "clean_use_auto_threshold": bool(clean_use_auto_threshold),
            "clean_target_keep_human_threshold": clean_target_keep_human_threshold,
            "clean_target_drop_human_threshold": clean_target_drop_human_threshold,
            "clean_target_precision": round(float(clean_target_precision), 4),
            "clean_min_support": clean_min_support,
            "clean_uncertain_action": clean_uncertain_action,
        },
        "auto_cleaning": {
            "policy": clean_policy,
            "summary": clean_summary,
        },
        "calibration": {
            "method": calibration.get("method"),
            "requested_method": calibration.get("requested_method"),
            "a": round(float(calibration.get("a", 1.0)), 6),
            "b": round(float(calibration.get("b", 0.0)), 6),
            "train_count": calibration.get("train_count"),
            "train_pearson": calibration.get("train_raw_r"),
            "train_mae": calibration.get("train_mae"),
            "train_loo_mae": calibration.get("train_loo_mae"),
            "clip_negative_to_zero": calibration.get("clip_negative_to_zero"),
            "candidates": calibration.get("candidates", []),
            "note": calibration.get("note"),
        },
        "train": {
            "xlsx": str(train_xlsx),
            "sample_count": len(train_records),
            "reference_image_count": len(reference_items),
            "calibration_pairs": [
                {
                    "sample_id": p.get("sample_id"),
                    "image_name": p.get("image_name"),
                    "human_avg": None if p.get("human_avg") is None else round(float(p.get("human_avg")), 2),
                    "model_raw": p.get("model_raw"),
                    "model_calibrated": None if p.get("model_raw") is None else (
                        _apply_calibration(p.get("model_raw"), calibration) if use_posthoc_calibration else int(p.get("model_raw"))
                    ),
                }
                for p in train_pairs
            ],
            "samples": [
                {
                    "sample_id": r.get("sample_id"),
                    "dataset": r.get("dataset"),
                    "relation": r.get("relation"),
                    "image_name": r.get("image_name"),
                    "human_avg": None if r.get("human_avg") is None else round(float(r.get("human_avg")), 2),
                    "final_level": r.get("final_level"),
                    "reason": r.get("reason"),
                }
                for r in train_records
            ],
        },
        "test": {
            "xlsx": str(test_xlsx),
            "evaluated_count": len(result_rows),
            "metrics": metrics_calibrated,
            "metrics_raw": metrics_raw,
            "metrics_calibrated": metrics_calibrated,
            "cleaning_summary": clean_summary,
            "results": result_rows,
        },
        "stability_test": stability_report,
    }

    output_path.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")

    print("\n--- 人工校准评测完成 ---")
    print(f"输出文件: {output_path}")
    print(
        "Test指标(校准后): "
        f"count={metrics_calibrated.get('count')} | "
        f"MAE={metrics_calibrated.get('mae')} | RMSE={metrics_calibrated.get('rmse')} | "
        f"Spearman={metrics_calibrated.get('spearman')} | "
        f"Pearson={metrics_calibrated.get('pearson')}"
    )
    print(
        f"Test指标(原始{score_aggregation}): "
        f"count={metrics_raw.get('count')} | "
        f"MAE={metrics_raw.get('mae')} | RMSE={metrics_raw.get('rmse')} | "
        f"Spearman={metrics_raw.get('spearman')} | "
        f"Pearson={metrics_raw.get('pearson')}"
    )

    print("\n--- 评审模型有效性检查 ---")
    train_pearson = calibration.get("train_pearson")
    test_pearson = metrics_calibrated.get("pearson")
    model_valid = True
    
    if train_pearson is not None:
        print(f"Train集皮尔森系数: {train_pearson:.4f} (阈值: {min_pearson_threshold:.2f})")
        if train_pearson < min_pearson_threshold:
            print(f"  [WARNING] Train集皮尔森系数低于阈值！")
            model_valid = False
        else:
            print(f"  [OK] Train集皮尔森系数达标")
    else:
        print("  Train集皮尔森系数不可用，跳过检查")
    
    if test_pearson is not None:
        print(f"Test集皮尔森系数: {test_pearson:.4f} (阈值: {min_pearson_threshold:.2f})")
        if test_pearson < min_pearson_threshold:
            print(f"  [WARNING] Test集皮尔森系数低于阈值！")
            model_valid = False
        else:
            print(f"  [OK] Test集皮尔森系数达标")
    else:
        print("  Test集皮尔森系数不可用，跳过检查")
    
    if model_valid:
        print("[OK] 评审模型有效性检查通过！可以放心用于筛选样本。")
    else:
        print("[FAIL] 评审模型有效性检查未通过！")
        if strict_pearson_check:
            print("  由于启用了严格检查模式，任务中止。")
            raise RuntimeError("评审模型皮尔森系数未达阈值，严格检查模式下中止执行。")
        else:
            print("  警告：虽然检查未通过，但由于未启用严格模式，将继续执行后续操作。")
            print("  建议：优化评审模型或人工复核筛选结果。")
    print(
        "自动清洗汇总: "
        f"keep={clean_summary.get('keep_count')} ({clean_summary.get('keep_ratio')}), "
        f"drop={clean_summary.get('drop_count')} ({clean_summary.get('drop_ratio')}), "
        f"uncertain_before_policy={clean_summary.get('uncertain_before_policy')}, "
        f"unstable={clean_summary.get('unstable_count')}, "
        f"keep_precision={clean_summary.get('keep_precision_vs_human')}, "
        f"drop_precision={clean_summary.get('drop_precision_vs_human')}"
    )
    if stability_scores_calibrated:
        print(
            f"稳定性测试({stability_report.get('target_sample_id')}): "
            f"runs={stability_report.get('runs_total')}, "
            f"scores={stability_scores_calibrated}, "
            f"std={stability_report.get('std')}, range={stability_report.get('range')}, "
            f"passed={stability_report.get('passed_threshold')}"
        )
    else:
        print("稳定性测试未获得有效分数。")


def analyze_dataset(dataset_path: Path):
    """
    对单个数据集的所有图像进行评分、分析和聚合。
    """
    images_path = dataset_path / "images"
    dataset_name = dataset_path.name
    output_path = dataset_path / f"{dataset_name}_analysis.json"
    cache_path = dataset_path / CACHE_FILE_NAME

    if output_path.exists() and not FORCE_RERUN:
        print(f"跳过: {dataset_name} 已存在分析文件 {output_path.name}")
        return
    
    if not images_path.is_dir():
        print(f"警告: 在 {dataset_name} 中未找到 'images' 文件夹，跳过。")
        return

    image_files = sorted(list(images_path.glob("*.png")) + list(images_path.glob("*.jpg")), key=lambda p: p.name)
    if not image_files:
        print(f"警告: 在 {images_path} 中未找到任何图像文件，跳过。")
        return

    print(f"\n--- 正在处理数据集: {dataset_name} ({len(image_files)}张图片) ---")

    # 存储每张图片的分数和长度
    scores_by_length = {}
    image_score_cache = {}

    if RESUME_ENABLED and cache_path.exists():
        try:
            image_score_cache = json.loads(cache_path.read_text(encoding="utf-8"))
            if not isinstance(image_score_cache, dict):
                image_score_cache = {}
        except Exception:
            image_score_cache = {}
    
    def _add_score_by_file_name(img_name: str, sc: int):
        match = re.search(r'_L(\d+)', img_name)
        if not match:
            print(f"警告: 无法从文件名 {img_name} 中提取长度信息，跳过此文件。")
            return
        length = int(match.group(1))
        if length not in scores_by_length:
            scores_by_length[length] = []
        scores_by_length[length].append(sc)

    # 先处理缓存命中
    pending_files = []
    for image_file in image_files:
        cached = image_score_cache.get(image_file.name)
        if isinstance(cached, int) and 0 <= cached <= 100:
            _add_score_by_file_name(image_file.name, int(cached))
        else:
            pending_files.append(image_file)

    # 并发处理未命中缓存的图片
    if pending_files:
        cache_lock = threading.Lock()
        max_workers = max(1, int(MAX_CONCURRENCY))

        with ThreadPoolExecutor(max_workers=max_workers) as executor:
            future_to_file = {executor.submit(get_image_score, img): img for img in pending_files}
            for future in tqdm(as_completed(future_to_file), total=len(future_to_file), desc=f"分析 {dataset_name}"):
                image_file = future_to_file[future]
                score = None
                try:
                    score = future.result()
                except Exception as e:
                    print(f"错误: 并发评分 {image_file.name} 时异常: {e}")
                    score = None

                if score is None:
                    continue

                _add_score_by_file_name(image_file.name, int(score))

                if RESUME_ENABLED:
                    with cache_lock:
                        image_score_cache[image_file.name] = int(score)
                        cache_path.write_text(
                            json.dumps(image_score_cache, ensure_ascii=False, indent=2),
                            encoding="utf-8",
                        )

    if not scores_by_length:
        print(f"错误: 数据集 {dataset_name} 中没有任何图片被成功评分。")
        return

    # 计算聚合结果
    lengths = sorted(scores_by_length.keys())
    mean_scores = [np.mean(scores_by_length[l]) for l in lengths]
    all_scores = [score for l in lengths for score in scores_by_length[l]]
    overall_score = np.mean(all_scores)

    # 计算Spearman相关系数
    rho, p_value = stats.spearmanr(lengths, mean_scores)
    # 如果只有一个长度，spearmanr会返回nan，需要处理
    if np.isnan(rho):
        rho = 0.0

    # 判断关系
    if rho >= 0.2:
        relation = "positive"
    elif rho <= -0.2:
        relation = "negative"
    else:
        relation = "flat"

    # 准备最终的JSON数据
    result = {
        "dataset": dataset_name,
        "lengths": lengths,
        "mean_scores": [round(s, 2) for s in mean_scores],
        "rho_len_score": round(rho, 4),
        "relation": relation,
        "overall_score": round(overall_score, 2)
    }

    # 写入文件
    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump([result], f, ensure_ascii=False, indent=2)
    
    print(f"成功: 已将分析结果写入 {output_path}")


def parse_args():
    parser = argparse.ArgumentParser(description="UEA/UCR 图文一致性自动评测脚本")
    parser.add_argument(
        "--mode",
        choices=["dataset", "human_eval", "api_probe"],
        default="dataset",
        help="dataset: 原全量数据集模式；human_eval: Train校准 + Test对比模式；api_probe: API并发探测",
    )
    parser.add_argument(
        "--runtime_model_id",
        type=str,
        default="all",
        help="运行时生效模型id/name（可逗号分隔），默认 all；用于单模型压测/阈值测试",
    )
    parser.add_argument(
        "--probe_requests",
        type=int,
        default=20,
        help="api_probe 模式下的总请求数，默认 20",
    )
    parser.add_argument(
        "--probe_workers",
        type=int,
        default=max(1, int(MAX_CONCURRENCY)),
        help="api_probe 模式下并发worker数，默认与 MAX_CONCURRENCY 一致",
    )
    parser.add_argument(
        "--probe_timeout",
        type=int,
        default=PRECHECK_TIMEOUT,
        help="api_probe 模式下单请求超时（秒），默认与预检查超时一致",
    )
    parser.add_argument(
        "--probe_model_id",
        type=str,
        default="all",
        help="api_probe 模式下指定模型id/name（可逗号分隔），默认 all",
    )
    parser.add_argument(
        "--probe_batch_images",
        type=int,
        default=1,
        help="api_probe 模式下每个请求携带图片数，默认 1",
    )
    parser.add_argument(
        "--probe_worker_levels",
        type=str,
        default="",
        help="api_probe 模式下worker档位列表（逗号分隔，如 3,5,8,10），用于自动梯度探测",
    )
    parser.add_argument(
        "--probe_min_success_ratio",
        type=float,
        default=1.0,
        help="worker梯度探测判定阈值（0-1），默认 1.0",
    )
    parser.add_argument(
        "--probe_override_model_max_concurrency",
        type=int,
        default=0,
        help="api_probe 模式下临时覆盖模型并发槽位（>0生效，仅用于压测，不会写回配置）",
    )
    parser.add_argument(
        "--human_eval_root",
        type=str,
        default=str(HUMAN_EVAL_ROOT_DEFAULT),
        help="human_eval 模式下人工评审目录（含 Train/Test 子目录）",
    )
    parser.add_argument(
        "--stability_target",
        type=str,
        default=None,
        help="human_eval 模式下稳定性测试目标，可填 sample_id 或 image_name，默认取 Test 第一条",
    )
    parser.add_argument(
        "--stability_runs",
        type=int,
        default=HUMAN_EVAL_STABILITY_RUNS_DEFAULT,
        help="human_eval 模式下稳定性测试初始重复评审次数，默认 3",
    )
    parser.add_argument(
        "--stability_std_threshold",
        type=float,
        default=HUMAN_EVAL_STABILITY_STD_THRESHOLD_DEFAULT,
        help="稳定性判定阈值：标准差上限，默认 2.0",
    )
    parser.add_argument(
        "--stability_range_threshold",
        type=int,
        default=HUMAN_EVAL_STABILITY_RANGE_THRESHOLD_DEFAULT,
        help="稳定性判定阈值：极差上限，默认 5",
    )
    parser.add_argument(
        "--stability_max_extra_runs",
        type=int,
        default=HUMAN_EVAL_STABILITY_MAX_EXTRA_RUNS_DEFAULT,
        help="当稳定性未达标时允许自动追加的最大评审次数，默认 3",
    )
    parser.add_argument(
        "--test_runs_per_image",
        type=int,
        default=HUMAN_EVAL_TEST_RUNS_PER_IMAGE_DEFAULT,
        help="human_eval 模式下 Test 每张图重复评测次数，默认 3（先聚合后再校准）",
    )
    parser.add_argument(
        "--score_aggregation",
        choices=["mean", "median", "last"],
        default=HUMAN_EVAL_SCORE_AGGREGATION_DEFAULT,
        help="多次评测聚合方式：mean=均值(推荐，排序更稳)；median=中位数；last=最后一次",
    )
    parser.add_argument(
        "--batch_query_size",
        type=int,
        default=HUMAN_EVAL_BATCH_QUERY_SIZE_DEFAULT,
        help="human_eval 模式下每次请求的待评图片数（上限5），默认3",
    )
    parser.add_argument(
        "--calibration_mode",
        choices=["hybrid", "posthoc_only", "prompt_only"],
        default=HUMAN_EVAL_CALIBRATION_MODE_DEFAULT,
        help="human_eval 校准模式：hybrid=提示词+后处理；posthoc_only=仅后处理；prompt_only=仅提示词",
    )
    parser.add_argument(
        "--calibration_method",
        choices=["auto", "identity", "bias_only", "linear", "theil_sen"],
        default=HUMAN_EVAL_CALIBRATION_METHOD_DEFAULT,
        help="human_eval 后处理校准方法（mode 含 posthoc 时生效），默认 auto",
    )
    parser.add_argument(
        "--clip_negative_human_to_zero",
        type=int,
        choices=[0, 1],
        default=1 if HUMAN_EVAL_CLIP_NEGATIVE_HUMAN_TO_ZERO_DEFAULT else 0,
        help="是否将 Train 中 [-10,0) 的人工分按 0 参与校准拟合，1=是，0=否",
    )
    parser.add_argument(
        "--clean_keep_threshold",
        type=int,
        default=HUMAN_EVAL_CLEAN_KEEP_THRESHOLD_DEFAULT,
        help="自动清洗保留阈值（含等号），默认 75",
    )
    parser.add_argument(
        "--clean_drop_threshold",
        type=int,
        default=HUMAN_EVAL_CLEAN_DROP_THRESHOLD_DEFAULT,
        help="自动清洗剔除阈值（含等号），默认 60",
    )
    parser.add_argument(
        "--clean_uncertain_action",
        choices=["drop", "keep", "midpoint"],
        default=HUMAN_EVAL_CLEAN_UNCERTAIN_ACTION_DEFAULT,
        help="灰区样本决策策略：drop=直接剔除，keep=直接保留，midpoint=按中点自动二分",
    )
    parser.add_argument(
        "--clean_use_auto_threshold",
        type=int,
        choices=[0, 1],
        default=1 if HUMAN_EVAL_CLEAN_USE_AUTO_THRESHOLD_DEFAULT else 0,
        help="是否在 Train 上自动拟合 keep/drop 阈值，1=是，0=否",
    )
    parser.add_argument(
        "--clean_target_keep_human_threshold",
        type=int,
        default=HUMAN_EVAL_CLEAN_TARGET_KEEP_HUMAN_THRESHOLD_DEFAULT,
        help="自动拟合时对 keep 的人工正例阈值（human>=该值），默认 75",
    )
    parser.add_argument(
        "--clean_target_drop_human_threshold",
        type=int,
        default=HUMAN_EVAL_CLEAN_TARGET_DROP_HUMAN_THRESHOLD_DEFAULT,
        help="自动拟合时对 drop 的人工负例阈值（human<=该值），默认 60",
    )
    parser.add_argument(
        "--clean_target_precision",
        type=float,
        default=HUMAN_EVAL_CLEAN_TARGET_PRECISION_DEFAULT,
        help="自动拟合阈值的目标精度下限（0-1），默认 0.9",
    )
    parser.add_argument(
        "--clean_min_support",
        type=int,
        default=HUMAN_EVAL_CLEAN_MIN_SUPPORT_DEFAULT,
        help="自动拟合阈值时每个候选阈值的最小样本数，默认 2",
    )
    parser.add_argument(
        "--min_pearson_threshold",
        type=float,
        default=HUMAN_EVAL_MIN_PEARSON_THRESHOLD_DEFAULT,
        help="评审模型有效性检查：皮尔森系数最低阈值，默认 0.8",
    )
    parser.add_argument(
        "--strict_pearson_check",
        type=int,
        choices=[0, 1],
        default=1 if HUMAN_EVAL_STRICT_PEARSON_CHECK_DEFAULT else 0,
        help="是否严格执行皮尔森系数检查：1=不达标就中止，0=仅警告但继续，默认 0",
    )
    parser.add_argument(
        "--force_human_eval_rerun",
        action="store_true",
        help="human_eval 模式下忽略 Test 评分缓存并重新评测",
    )
    parser.add_argument(
        "--skip_precheck",
        action="store_true",
        help="跳过运行前模型可用性检查",
    )
    return parser.parse_args()


def main():
    """
    主执行函数
    """
    args = parse_args()

    _apply_runtime_model_filter(args.runtime_model_id)
    if str(args.runtime_model_id or "all").strip().lower() not in {"", "all", "*"}:
        print(f"运行时模型筛选: {args.runtime_model_id}")

    chain_text = " -> ".join([f"{cfg['name']}[{cfg.get('id', 'default')}]" for cfg in MODEL_FALLBACKS])
    print(f"模型调度链: {chain_text}")
    for cfg in MODEL_FALLBACKS:
        print(
            "  - "
            f"{cfg['name']} | id={cfg.get('id', 'default')} | provider={cfg.get('provider', 'unknown')} "
            f"| RPM={cfg.get('rpm', RPM_LIMIT)} TPM={cfg.get('tpm', TPM_LIMIT)}"
        )

    configured_parallel_slots = sum(
        max(1, int(cfg.get("max_concurrency", PER_KEY_MAX_CONCURRENCY))) for cfg in MODEL_FALLBACKS
    )
    print(
        "并发估算: "
        f"每key默认并发={PER_KEY_MAX_CONCURRENCY}, key数量={len(MODEL_FALLBACKS)}, "
        f"按模型配置并发槽位总和={configured_parallel_slots}, "
        f"单图并发上限约={configured_parallel_slots}, "
        f"批量(每请求最多{HUMAN_EVAL_MAX_IMAGES_PER_REQUEST}图)并发上限约="
        f"{configured_parallel_slots * HUMAN_EVAL_MAX_IMAGES_PER_REQUEST}图"
    )
    if _has_doubao_model_configured():
        doubao_cfgs = _find_doubao_models()
        key_summary_items = []
        key_to_model_ids: dict[str, list[str]] = {}
        for cfg in doubao_cfgs:
            cfg_id = str(cfg.get("id") or "default")
            cfg_name = str(cfg.get("name") or "unknown")
            api_key = str(cfg.get("api_key") or "").strip()
            key_summary_items.append(f"{cfg_id}={_masked_key_tail(api_key)}")
            if api_key:
                key_to_model_ids.setdefault(api_key, []).append(cfg_id)
            else:
                print(f"警告: 豆包模型未配置API key，将不可用 -> {cfg_name}[{cfg_id}]")

        print("豆包key摘要: " + ", ".join(key_summary_items))

        for model_ids in key_to_model_ids.values():
            if len(model_ids) > 1:
                print(
                    "提示: 以下豆包模型复用同一API key，请确认是否为预期: "
                    + ", ".join(model_ids)
                )

        concurrency_with_doubao = _sum_configured_concurrency_slots(include_doubao=True)
        concurrency_without_doubao = _sum_configured_concurrency_slots(include_doubao=False)
        print(
            "上游并发策略: "
            f"豆包可用时={concurrency_with_doubao}, "
            f"豆包停用时={concurrency_without_doubao}, "
            f"当前提示并发={MAX_CONCURRENCY}"
        )
    if _find_first_siliconflow_model() is not None:
        print(
            "硅基动态降并发策略: "
            f"单轮重试次数>= {SILICONFLOW_RETRY_DEGRADE_THRESHOLD} 时，当前任务临时降为1并发，"
            "任务结束后自动恢复到配置并发"
        )
    if MODEL_CIRCUIT_BREAKER_ENABLED:
        managed_providers_text = ",".join(sorted(MODEL_CIRCUIT_MANAGED_PROVIDERS)) or "<none>"
        print(
            "模型熔断策略: "
            f"作用provider={managed_providers_text}, "
            f"404冷却={MODEL_CIRCUIT_NOT_FOUND_RECHECK_INTERVAL_SEC}s, "
            f"鉴权冷却={MODEL_CIRCUIT_AUTH_RECHECK_INTERVAL_SEC}s, "
            f"额度冷却={MODEL_CIRCUIT_QUOTA_RECHECK_INTERVAL_SEC}s, "
            f"429限流冷却={MODEL_CIRCUIT_RATE_LIMIT_RECHECK_INTERVAL_SEC}s"
        )

    if PRECHECK_MODELS_BEFORE_RUN and not args.skip_precheck:
        run_model_precheck_or_raise()

    _startup_probe_doubao_quota()
    if _has_doubao_model_configured():
        print(f"豆包探测后当前上游并发提示={MAX_CONCURRENCY}")

    if args.mode == "api_probe":
        worker_levels = sorted(set(_parse_int_list(args.probe_worker_levels)))

        if worker_levels:
            print(
                "api_probe 梯度探测: "
                f"model={args.probe_model_id}, batch_images={args.probe_batch_images}, "
                f"worker_levels={worker_levels}, requests={args.probe_requests}, timeout={args.probe_timeout}s"
            )
            results = []
            for lvl in worker_levels:
                res = run_api_concurrency_probe(
                    total_requests=args.probe_requests,
                    workers=lvl,
                    timeout_sec=args.probe_timeout,
                    probe_model_id=args.probe_model_id,
                    probe_batch_images=args.probe_batch_images,
                    probe_override_model_max_concurrency=args.probe_override_model_max_concurrency,
                )
                results.append(res)

            threshold = min(1.0, max(0.0, float(args.probe_min_success_ratio)))
            eligible = [r for r in results if float(r.get("success_ratio", 0.0)) >= threshold]
            recommended = max(eligible, key=lambda x: int(x.get("workers", 0))) if eligible else None

            print("\n--- 梯度探测汇总 ---")
            for r in results:
                print(
                    f"workers={r.get('workers')} | success_ratio={r.get('success_ratio')} "
                    f"| fail={r.get('fail_count')} | p95={r.get('p95_latency_sec')}s "
                    f"| estimated_parallel_images_current_probe={r.get('estimated_parallel_images_current_probe')}"
                )

            if recommended is not None:
                print(
                    "推荐稳定worker上限: "
                    f"{recommended.get('workers')} (阈值 success_ratio>={threshold})"
                )
            else:
                best_effort = max(
                    results,
                    key=lambda x: (float(x.get("success_ratio", 0.0)), int(x.get("workers", 0))),
                )
                print(
                    "未达到设定阈值，best-effort结果: "
                    f"workers={best_effort.get('workers')}, success_ratio={best_effort.get('success_ratio')}"
                )
        else:
            run_api_concurrency_probe(
                total_requests=args.probe_requests,
                workers=args.probe_workers,
                timeout_sec=args.probe_timeout,
                probe_model_id=args.probe_model_id,
                probe_batch_images=args.probe_batch_images,
                probe_override_model_max_concurrency=args.probe_override_model_max_concurrency,
            )
        return

    if args.mode == "human_eval":
        run_human_eval(
            human_eval_root=Path(args.human_eval_root),
            stability_target=args.stability_target,
            stability_runs=args.stability_runs,
            stability_std_threshold=args.stability_std_threshold,
            stability_range_threshold=args.stability_range_threshold,
            stability_max_extra_runs=args.stability_max_extra_runs,
            force_rerun=args.force_human_eval_rerun,
            test_runs_per_image=args.test_runs_per_image,
            score_aggregation=args.score_aggregation,
            batch_query_size=args.batch_query_size,
            calibration_mode=args.calibration_mode,
            calibration_method=args.calibration_method,
            clip_negative_human_to_zero=bool(args.clip_negative_human_to_zero),
            clean_keep_threshold=args.clean_keep_threshold,
            clean_drop_threshold=args.clean_drop_threshold,
            clean_uncertain_action=args.clean_uncertain_action,
            clean_use_auto_threshold=bool(args.clean_use_auto_threshold),
            clean_target_keep_human_threshold=args.clean_target_keep_human_threshold,
            clean_target_drop_human_threshold=args.clean_target_drop_human_threshold,
            clean_target_precision=args.clean_target_precision,
            clean_min_support=args.clean_min_support,
            min_pearson_threshold=args.min_pearson_threshold,
            strict_pearson_check=bool(args.strict_pearson_check),
        )
        return

    # 获取所有数据集路径
    base_path = Path(".") # 脚本在根目录，所以是当前目录
    multivariate_path = base_path / "Multivariate"
    univariate_path = base_path / "Univariate"

    dataset_paths = []
    if multivariate_path.is_dir():
        dataset_paths.extend([d for d in multivariate_path.iterdir() if d.is_dir()])
    if univariate_path.is_dir():
        dataset_paths.extend([d for d in univariate_path.iterdir() if d.is_dir()])
    dataset_paths = sorted(dataset_paths, key=lambda p: str(p).lower())

    if not dataset_paths:
        print("错误: 在 'Multivariate' 或 'Univariate' 文件夹中未找到任何数据集目录。")
        return
        
    print(f"总共找到 {len(dataset_paths)} 个数据集。开始处理...")

    # 遍历并分析每个数据集
    for dataset_path in dataset_paths:
        analyze_dataset(dataset_path)
        
    print("\n--- 所有数据集处理完毕！ ---")


if __name__ == "__main__":
    main()
